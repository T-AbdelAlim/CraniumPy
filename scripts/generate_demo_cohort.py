"""generates the shipped synthetic demo cohort under
src/craniumpy_core/demo_cohort/ - a one-time script, not run at app startup
(same "committed package data" pattern as template_registry.py's shipped
templates, or resources/test_mesh/test_mesh_holes.ply). re-run by hand only
if the demo cohort itself needs to change; the output is committed, not
regenerated on every install or test run.

~150 synthetic cases across 5 diagnoses (control + 4 real craniofacial
conditions this app's own measurements target: sagittal/scaphocephaly,
metopic/trigonocephaly, unicoronal synostosis, and non-synostotic
positional/deformational plagiocephaly - kept as a genuinely separate
diagnosis from unicoronal synostosis, not conflated with it, since the two
have very different treatment (surgery vs. helmet-only) and severity
profiles). surgical cases get a second, post-op row - see PRE_OP_AGE /
POST_OP_OFFSETS below - so the demo cohort can actually demonstrate
"stratify by image_timing within one treatment group" (see the Cohort
workspace's FilterBar + Stratify tab), not just a single cross-sectional
snapshot per patient.

clinical parameters (treatment age/type per diagnosis, expected CI/frontal-
angle/asymmetry direction and rough magnitude per subtype) are loosely
grounded in published craniosynostosis literature and Dutch/European
treatment norms (Erasmus MC's metopic cohort, van der Meulen's interfrontal
angle work, CVAI/CVA severity banding, standard endoscopic-vs-open surgical
timing windows) - approximate ballpark figures for demo realism, NOT
validated clinical reference ranges. every row is otherwise shaped exactly
like a real exported cohort row (see api/results_bundle.py's
_metrics_row/COLUMNS below) so nothing downstream has to know this cohort
is synthetic.

a subset of rows also get a synthetic NICP-topology mesh (smooth randomized
deformations of the real shipped templates - same vertex count/face
connectivity as the template by construction, so they're vertex-
correspondent exactly like a real NICP fit's output would be - see
craniumpy_core.cohort.mean_shape), which is what makes the Mean shape tab
demoable without needing real scans or an actual NICP run.

run from the repo root:
    python scripts/generate_demo_cohort.py
"""

from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

import numpy as np
import trimesh
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "src"))

from craniumpy_core.io import load_mesh, save_mesh  # noqa: E402
from craniumpy_core.template_registry import TEMPLATES_DIR  # noqa: E402

OUTPUT_DIR = REPO_ROOT / "src" / "craniumpy_core" / "demo_cohort"
MESHES_DIR = OUTPUT_DIR / "meshes"

RNG_SEED = 20260213  # fixed so the committed output is reproducible across regenerations
N_CASES = 150  # patients/cases - surgical cases add a 2nd (post-op) row, so total rows > N_CASES
N_NICP_PER_TARGET = 25  # -> ~50 NICP-fitted rows total, split across the 2 template groups

# mirrors api/results_bundle.py's _METOPIC_ROW_KEYS exactly - keep in sync.
METOPIC_KEYS = (
    "frontal_angle_deg", "midline_curvature_concentration", "midline_max_curvature",
    "midline_max_curvature_position", "ridge_protrusion_mm", "ridge_area_mm2",
    "ridge_area_normalized", "left_temporal_hollowing", "right_temporal_hollowing",
    "mean_temporal_hollowing", "left_max_temporal_depth_mm", "right_max_temporal_depth_mm",
    "parabolic_deviation_index",
)

# mirrors _metrics_row's own column set/order (minus patient_id, which
# real cohort exports drop too - see _upsert_cohort_xlsx).
COLUMNS = [
    "cohort_id", "file_name", "file_path", "diagnosis", "sex", "date_imaging", "age_imaging",
    "image_timing", "treatment", "age_surgery_months", "free_variable", "target", "com_correction",
    "nicp_used", "nicp_template", "nicp_mesh_path", "depth_mm", "breadth_mm", "cephalic_index",
    "circumference_cm", "mesh_volume_cc", "cranial_asymmetry_index", "mean_asymmetry_index",
    "frontal_bossing_angle_deg",
] + [f"metopic_{key}" for key in METOPIC_KEYS]

NUMERIC_COLUMNS = {
    "age_imaging", "age_surgery_months", "depth_mm", "breadth_mm", "cephalic_index", "circumference_cm",
    "mesh_volume_cc", "cranial_asymmetry_index", "mean_asymmetry_index", "frontal_bossing_angle_deg",
} | {f"metopic_{key}" for key in METOPIC_KEYS}

DIAGNOSES = ("control", "scaphocephaly", "trigonocephaly", "unicoronal_synostosis", "positional_plagiocephaly")
DIAGNOSIS_WEIGHTS = {
    "control": 0.30,
    "scaphocephaly": 0.20,
    "trigonocephaly": 0.20,
    "unicoronal_synostosis": 0.15,
    "positional_plagiocephaly": 0.15,
}

# diagnosis -> probability this row's target is "face" rather than
# "cranium" - trigonocephaly (metopic synostosis) is a forehead condition,
# so it's mostly processed as "face"; the skull-shape conditions
# (scaphocephaly, unicoronal synostosis, positional plagiocephaly) mostly
# as "cranium". control cases split roughly evenly.
DIAGNOSIS_FACE_BIAS = {
    "control": 0.5,
    "scaphocephaly": 0.1,
    "trigonocephaly": 0.85,
    "unicoronal_synostosis": 0.25,
    "positional_plagiocephaly": 0.15,
}

CRANIUM_NICP_TEMPLATE = "clipped_template_xy_com"
FACE_NICP_TEMPLATE = "template_face"

# --- treatment & timing, per diagnosis ------------------------------------
# see the module docstring - approximate figures from published
# craniosynostosis literature / Dutch-European treatment norms, not
# clinical reference values.

# fraction of each surgical-eligible diagnosis that actually goes to
# surgery vs. observation/conservative management. contemporary Dutch
# metopic practice in particular is now majority-observation (roughly 2/3
# conservative per Erasmus MC's own cohort). positional plagiocephaly and
# control are handled separately below (never "surgical" in this schema).
SURGICAL_FRACTION = {
    "scaphocephaly": 0.95,
    "trigonocephaly": 0.35,
    "unicoronal_synostosis": 0.90,
}

# diagnosis -> [(probability, procedure label, age_mean, age_sd, age_lo,
# age_hi, helmet_after_surgery)] - sagittal synostosis is the one condition
# here with a real fork between an early endoscopic approach (~1.6-5.6mo,
# must be done before ~4-6mo, commonly followed by helmet therapy to guide
# ongoing molding) and later open cranial vault remodeling (~3-8mo); the
# other two surgical diagnoses are open fronto-orbital advancement only,
# typically 8-14mo.
SURGICAL_TIMING = {
    "scaphocephaly": [
        (0.55, "endoscopic strip craniectomy", 3.3, 1.0, 1.6, 5.6, True),
        (0.45, "open cranial vault remodeling", 5.0, 1.3, 3.0, 8.0, False),
    ],
    "trigonocephaly": [
        (1.0, "fronto-orbital advancement", 10.5, 1.3, 8.0, 14.0, False),
    ],
    "unicoronal_synostosis": [
        (1.0, "fronto-orbital advancement", 10.0, 1.4, 7.0, 14.0, False),
    ],
}

# post-op follow-up scan timing (months since surgery) - baseline, 6 weeks,
# 6 months, 12 months is a commonly reported 3D-photogrammetry follow-up
# schedule for craniosynostosis surgery (exact intervals vary by center).
# only one follow-up is generated per surgical case (not the full
# schedule), weighted toward the 6-month point as the single most
# representative post-op scan.
POST_OP_OFFSETS_MONTHS = {"6w": 1.5, "6mo": 6.0, "1y": 12.0}
POST_OP_OFFSET_WEIGHTS = [0.25, 0.5, 0.25]


def _fmt(value: float) -> str:
    return f"{value:.2f}"


def _draw(rng: np.random.Generator, mean: float, sd: float, lo: float, hi: float) -> float:
    return float(np.clip(rng.normal(mean, sd), lo, hi))


def _draw_param(rng: np.random.Generator, table: dict, diagnosis: str, phase: str) -> float:
    """looks up a (mean, sd, lo, hi) tuple for (diagnosis, phase), falling
    back to that diagnosis's "pre" entry (a diagnosis with no distinct
    post-op numbers modeled - it never actually reaches a post-op row -
    just reuses its own pre-op distribution), then to control's "pre" entry
    (a diagnosis/target combination with no clinical signal modeled at all,
    e.g. a rare cranium-target trigonocephaly row) - draws from whichever
    is found."""
    for key in ((diagnosis, phase), (diagnosis, "pre"), ("control", "pre")):
        if key in table:
            return _draw(rng, *table[key])
    raise KeyError(f"no fallback found for {diagnosis!r}/{phase!r}")  # pragma: no cover - table always has control/pre


# (mean, sd, lo, hi) per (diagnosis, phase). "post" entries only exist
# where a post-op improvement is actually modeled - surgical correction
# rarely fully normalizes the shape, so post-op numbers move toward (not
# all the way to) the control range.
CEPHALIC_INDEX_PARAMS = {
    ("control", "pre"): (81, 3.5, 72, 90),
    ("scaphocephaly", "pre"): (64, 3.0, 55, 69),  # dolichocephalic, markedly low
    ("scaphocephaly", "post"): (73, 3.0, 66, 80),
    ("trigonocephaly", "pre"): (80, 4.0, 70, 90),  # metopic synostosis - CI itself isn't the abnormal axis
    ("unicoronal_synostosis", "pre"): (93, 4.0, 86, 102),  # brachycephalic, markedly high
    ("unicoronal_synostosis", "post"): (85, 3.5, 78, 92),
    ("positional_plagiocephaly", "pre"): (84, 4.0, 75, 94),  # mildly elevated, not the defining feature
}

FRONTAL_ANGLE_PARAMS = {
    ("control", "pre"): (144, 8, 125, 160),
    ("trigonocephaly", "pre"): (116, 9, 95, 135),  # acute/pointed - the defining metopic feature
    ("trigonocephaly", "post"): (136, 7, 122, 152),
}

RIDGE_PROTRUSION_PARAMS = {
    ("control", "pre"): (1.0, 0.6, -0.5, 2.8),
    ("trigonocephaly", "pre"): (3.6, 1.3, 0.5, 7.0),
    ("trigonocephaly", "post"): (1.6, 0.7, 0.0, 3.5),
}

FRONTAL_BOSSING_PARAMS = {
    ("control", "pre"): (152, 10, 125, 172),
    ("trigonocephaly", "pre"): (145, 10, 120, 165),
    ("trigonocephaly", "post"): (150, 9, 125, 170),
}

# app-native units (mean per-vertex left/right deviation, mm - see
# craniumpy_core.asymmetry) - NOT the CVAI% figures reported in the
# literature, which use a different formula/scale entirely. only the
# relative ordering between diagnoses (positional plagiocephaly and
# unicoronal synostosis clearly most asymmetric; scaphocephaly/
# trigonocephaly mildly-moderately so; control least) is meant to track
# the research, not the absolute numbers.
ASYMMETRY_PARAMS = {
    ("control", "pre"): (3.0, 0.9, 0.5, 5.5),
    ("scaphocephaly", "pre"): (3.8, 1.1, 1.0, 7.0),
    ("scaphocephaly", "post"): (3.0, 0.9, 1.0, 5.5),
    ("trigonocephaly", "pre"): (3.4, 1.1, 1.0, 6.5),
    ("trigonocephaly", "post"): (2.6, 0.9, 0.8, 5.0),
    ("unicoronal_synostosis", "pre"): (6.5, 1.8, 3.0, 12.0),
    ("unicoronal_synostosis", "post"): (3.8, 1.3, 1.5, 7.5),
    ("positional_plagiocephaly", "pre"): (7.5, 2.2, 3.5, 14.0),  # most asymmetric of all groups
}


def _decide_treatment(rng: np.random.Generator, diagnosis: str) -> tuple[str, float | None]:
    """-> (treatment label, age at surgery in months | None). the only
    diagnoses that ever get a real age_surgery_months (and therefore a
    post-op follow-up row - see _generate_case) are the 3 synostosis
    conditions with a surgical fraction above; positional plagiocephaly is
    helmet-or-observation only (never "surgery" under this schema's own
    meaning of the field), and control is always observation."""
    if diagnosis == "control":
        return "observation", None
    if diagnosis == "positional_plagiocephaly":
        return ("helmet therapy" if rng.random() < 0.6 else "repositioning / observation"), None

    if rng.random() >= SURGICAL_FRACTION[diagnosis]:
        return "observation", None

    options = SURGICAL_TIMING[diagnosis]
    probs = np.array([o[0] for o in options], dtype=float)
    label, mean, sd, lo, hi, helmet_after = options[rng.choice(len(options), p=probs / probs.sum())][1:]
    age_surgery = _draw(rng, mean, sd, lo, hi)
    treatment = f"{label} + helmet" if helmet_after else label
    return treatment, age_surgery


def _pre_op_age_imaging(rng: np.random.Generator, diagnosis: str, age_surgery: float | None) -> float:
    if age_surgery is not None:
        # imaged shortly (2-6 weeks) before surgery, not at some unrelated
        # earlier point - this is the pre-op baseline scan.
        return max(0.5, age_surgery - _draw(rng, 1.0, 0.5, 0.3, 2.0))
    if diagnosis == "positional_plagiocephaly":
        # imaged when helmet therapy is being considered - consensus window.
        return _draw(rng, 6.0, 2.5, 2.0, 14.0)
    if diagnosis == "control":
        return _draw(rng, 8.0, 5.0, 1.0, 30.0)
    # non-surgical (observation) synostosis case - imaged across a broader
    # spread, matching ongoing conservative follow-up rather than a single
    # pre-op baseline tied to an operation date.
    return _draw(rng, 9.0, 5.0, 1.0, 30.0)


def _shift_date(base: date, months: float) -> date:
    total_months = base.month - 1 + round(months)
    year = base.year + total_months // 12
    month = total_months % 12 + 1
    return date(year, month, min(base.day, 28))


def _cranium_metrics(rng: np.random.Generator, diagnosis: str, phase: str) -> dict[str, str]:
    """depth/breadth/cephalic-index/circumference/volume + cranial
    asymmetry - see CEPHALIC_INDEX_PARAMS/ASYMMETRY_PARAMS for the
    per-diagnosis/phase distributions this draws from."""
    cephalic_index = _draw_param(rng, CEPHALIC_INDEX_PARAMS, diagnosis, phase)
    asymmetry = _draw_param(rng, ASYMMETRY_PARAMS, diagnosis, phase)

    breadth_mm = _draw(rng, 130, 8, 100, 160)
    depth_mm = breadth_mm / (cephalic_index / 100)
    circumference_cm = _draw(rng, 44, 3, 36, 52)
    mesh_volume_cc = _draw(rng, 900, 120, 550, 1300)

    return {
        "depth_mm": _fmt(depth_mm),
        "breadth_mm": _fmt(breadth_mm),
        "cephalic_index": _fmt(cephalic_index),
        "circumference_cm": _fmt(circumference_cm),
        "mesh_volume_cc": _fmt(mesh_volume_cc),
        "cranial_asymmetry_index": _fmt(asymmetry),
    }


def _face_metrics(rng: np.random.Generator, diagnosis: str, phase: str) -> dict[str, str]:
    """metopic/forehead measurements + facial asymmetry + frontal bossing -
    see FRONTAL_ANGLE_PARAMS/RIDGE_PROTRUSION_PARAMS/FRONTAL_BOSSING_PARAMS/
    ASYMMETRY_PARAMS for the per-diagnosis/phase distributions."""
    frontal_angle = _draw_param(rng, FRONTAL_ANGLE_PARAMS, diagnosis, phase)
    ridge_protrusion = _draw_param(rng, RIDGE_PROTRUSION_PARAMS, diagnosis, phase)
    frontal_bossing_angle = _draw_param(rng, FRONTAL_BOSSING_PARAMS, diagnosis, phase)
    asymmetry = _draw_param(rng, ASYMMETRY_PARAMS, diagnosis, phase)

    midline_curvature_concentration = _draw(rng, 0.35, 0.08, 0.1, 0.6)
    midline_max_curvature = _draw(rng, 0.012, 0.004, 0.003, 0.025)
    midline_max_curvature_position = _draw(rng, 0.5, 0.05, 0.35, 0.65)
    ridge_area_mm2 = _draw(rng, 110, 35, 30, 220)
    ridge_area_normalized = _draw(rng, 0.045, 0.015, 0.01, 0.09)
    left_hollow = _draw(rng, 0.12, 0.05, 0.0, 0.3)
    right_hollow = _draw(rng, 0.12, 0.05, 0.0, 0.3)
    left_depth = _draw(rng, 1.3, 0.5, 0.1, 3.0)
    right_depth = _draw(rng, 1.3, 0.5, 0.1, 3.0)
    parabolic_deviation_index = _draw(rng, 0.9, 0.3, 0.2, 2.0)

    return {
        "mean_asymmetry_index": _fmt(asymmetry),
        "frontal_bossing_angle_deg": _fmt(frontal_bossing_angle),
        "metopic_frontal_angle_deg": _fmt(frontal_angle),
        "metopic_midline_curvature_concentration": _fmt(midline_curvature_concentration),
        "metopic_midline_max_curvature": _fmt(midline_max_curvature),
        "metopic_midline_max_curvature_position": _fmt(midline_max_curvature_position),
        "metopic_ridge_protrusion_mm": _fmt(ridge_protrusion),
        "metopic_ridge_area_mm2": _fmt(ridge_area_mm2),
        "metopic_ridge_area_normalized": _fmt(ridge_area_normalized),
        "metopic_left_temporal_hollowing": _fmt(left_hollow),
        "metopic_right_temporal_hollowing": _fmt(right_hollow),
        "metopic_mean_temporal_hollowing": _fmt((left_hollow + right_hollow) / 2),
        "metopic_left_max_temporal_depth_mm": _fmt(left_depth),
        "metopic_right_max_temporal_depth_mm": _fmt(right_depth),
        "metopic_parabolic_deviation_index": _fmt(parabolic_deviation_index),
    }


def _blank_row() -> dict[str, str]:
    return {c: "" for c in COLUMNS}


def _build_row(
    rng: np.random.Generator, cohort_id: str, diagnosis: str, sex: str, target: str, treatment: str,
    age_surgery: float | None, age_imaging: float, image_timing: str, base_date: date, phase: str,
) -> dict[str, str]:
    row = _blank_row()
    row["cohort_id"] = cohort_id
    row["file_name"] = f"{cohort_id}_{'F' if target == 'face' else 'C'}.ply"
    row["diagnosis"] = diagnosis
    row["sex"] = sex
    row["date_imaging"] = base_date.isoformat()
    row["age_imaging"] = _fmt(age_imaging)
    row["image_timing"] = image_timing
    row["treatment"] = treatment
    if age_surgery is not None:
        row["age_surgery_months"] = _fmt(age_surgery)
    row["target"] = target
    row["com_correction"] = "yes"
    row.update(_cranium_metrics(rng, diagnosis, phase) if target == "cranium" else _face_metrics(rng, diagnosis, phase))
    return row


def _generate_case(rng: np.random.Generator, cohort_id_counter: list[int]) -> list[dict[str, str]]:
    """one synthetic patient - 1 row (pre-op), or 2 (pre-op + a single
    post-op follow-up) when the case actually went to surgery. every field
    that describes the patient rather than the visit (diagnosis, sex,
    target, treatment) is shared across both rows, same as a real patient
    scanned twice would have identical patient-level fields and only the
    visit-level ones (age_imaging, image_timing, date_imaging, the outcome
    metrics) differ."""
    diagnosis = str(rng.choice(DIAGNOSES, p=[DIAGNOSIS_WEIGHTS[d] for d in DIAGNOSES]))
    target = "face" if rng.random() < DIAGNOSIS_FACE_BIAS[diagnosis] else "cranium"
    sex = str(rng.choice(["M", "F"]))
    treatment, age_surgery = _decide_treatment(rng, diagnosis)
    pre_op_age = _pre_op_age_imaging(rng, diagnosis, age_surgery)
    base_date = date(int(rng.integers(2022, 2026)), int(rng.integers(1, 13)), int(rng.integers(1, 29)))

    def next_id() -> str:
        cohort_id_counter[0] += 1
        return f"D{cohort_id_counter[0]:04d}"

    rows = [
        _build_row(
            rng, next_id(), diagnosis, sex, target, treatment, age_surgery, pre_op_age, "pre-op", base_date, "pre",
        )
    ]

    if age_surgery is not None:
        offset_label = str(rng.choice(list(POST_OP_OFFSETS_MONTHS), p=POST_OP_OFFSET_WEIGHTS))
        offset_months = POST_OP_OFFSETS_MONTHS[offset_label]
        post_op_age = age_surgery + offset_months
        post_op_date = _shift_date(base_date, (post_op_age - pre_op_age))
        rows.append(
            _build_row(
                rng, next_id(), diagnosis, sex, target, treatment, age_surgery, post_op_age,
                f"post_op_{offset_label}", post_op_date, "post",
            )
        )

    return rows


def _deform_template(template: trimesh.Trimesh, rng: np.random.Generator, n_bumps: int = 6, max_amplitude_mm: float = 2.5) -> trimesh.Trimesh:
    """a smooth, randomized per-patient variant of a shipped template -
    stands in for a real NICP fit's output without needing an actual scan
    or an actual NICP run. displaces vertices along their own normals by a
    sum of a few broad Gaussian "bumps" (not per-vertex independent noise,
    which would look like sensor noise rather than a plausible shape
    variation) - critically, keeps the exact same vertex count and face
    array as the input template, so the result is vertex-correspondent
    with every other deformation of the same template, exactly like real
    NICP-fitted patients sharing a template are (see
    craniumpy_core.cohort.mean_shape's own validation)."""
    vertices = np.asarray(template.vertices, dtype=float)
    normals = np.asarray(template.vertex_normals, dtype=float)
    sigma = float(template.bounding_box.extents.max()) * 0.18

    displacement = np.zeros(len(vertices))
    for center_idx in rng.integers(0, len(vertices), size=n_bumps):
        dist = np.linalg.norm(vertices - vertices[center_idx], axis=1)
        amplitude = rng.uniform(-max_amplitude_mm, max_amplitude_mm)
        displacement += amplitude * np.exp(-(dist**2) / (2 * sigma**2))

    deformed = vertices + normals * displacement[:, None]
    return trimesh.Trimesh(vertices=deformed, faces=template.faces, process=False)


def main() -> None:
    rng = np.random.default_rng(RNG_SEED)
    MESHES_DIR.mkdir(parents=True, exist_ok=True)

    rows: list[dict[str, str]] = []
    cohort_id_counter = [0]
    for _ in range(N_CASES):
        rows.extend(_generate_case(rng, cohort_id_counter))

    # NICP mesh assignment - a random subset of rows per target, independent
    # of diagnosis/phase (this feature is about template-topology grouping,
    # not clinical subgroup), so pre-op and post-op rows are equally likely
    # to end up NICP-fitted.
    cranium_template = load_mesh(TEMPLATES_DIR / f"{CRANIUM_NICP_TEMPLATE}.ply")
    face_template = load_mesh(TEMPLATES_DIR / f"{FACE_NICP_TEMPLATE}.ply")
    for target, template_mesh, template_name in (
        ("cranium", cranium_template, CRANIUM_NICP_TEMPLATE),
        ("face", face_template, FACE_NICP_TEMPLATE),
    ):
        candidates = [r for r in rows if r["target"] == target]
        chosen = rng.choice(len(candidates), size=min(N_NICP_PER_TARGET, len(candidates)), replace=False)
        for idx in chosen:
            row = candidates[idx]
            deformed = _deform_template(template_mesh, rng)
            relative_path = f"meshes/{row['file_name'].removesuffix('.ply')}_rg_N.ply"
            save_mesh(deformed, OUTPUT_DIR / relative_path)
            row["nicp_used"] = "yes"
            row["nicp_template"] = template_name
            row["nicp_mesh_path"] = relative_path
    for row in rows:
        if not row["nicp_used"]:
            row["nicp_used"] = "no"

    _write_xlsx(rows)
    nicp_count = sum(1 for r in rows if r["nicp_used"] == "yes")
    post_op_count = sum(1 for r in rows if r["image_timing"].startswith("post_op_"))
    print(f"wrote {len(rows)} rows ({N_CASES} cases, {post_op_count} with a post-op follow-up) to {OUTPUT_DIR / 'demo_cohort.xlsx'}")
    print(f"wrote {nicp_count} synthetic NICP-fitted meshes to {MESHES_DIR}")


def _write_xlsx(rows: list[dict[str, str]]) -> None:
    """same table formatting as api/results_bundle.py's _write_xlsx_rows -
    not reused directly (that function lives behind the API layer, and
    duplicating ~15 lines here keeps this script standalone/runnable
    without importing api.*)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    display_values = []
    for row in rows:
        cells = []
        for name in COLUMNS:
            value = row.get(name, "")
            if name in NUMERIC_COLUMNS and value != "":
                cells.append(float(value))
            elif name in NUMERIC_COLUMNS:
                cells.append(None)
            else:
                cells.append(value)
        ws.append(cells)
        display_values.append([str(c) if c is not None else "" for c in cells])

    for i, name in enumerate(COLUMNS, start=1):
        if name in NUMERIC_COLUMNS:
            for cell in ws[get_column_letter(i)][1:]:
                cell.number_format = "0.00"
        widest = max([len(name)] + [len(values[i - 1]) for values in display_values])
        ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 10), 40)

    last_col = get_column_letter(len(COLUMNS))
    table = Table(displayName="demo_cohort", ref=f"A1:{last_col}{len(rows) + 1}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
    ws.add_table(table)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_DIR / "demo_cohort.xlsx")


if __name__ == "__main__":
    main()
