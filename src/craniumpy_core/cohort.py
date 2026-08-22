"""cohort-level analysis: reading an accumulated cohort spreadsheet back in
(see api/results_bundle.py's _upsert_cohort_xlsx, the thing that writes
one), and computing a mean 3D shape across patients who share the same NICP
template topology.

mean_shape relies on registration.nicp.register_template's own guarantee
(see that module) that a template fit always returns the SAME vertex count,
face connectivity, and vertex order as the template itself - so two
patients fit to the identical template are exactly vertex-correspondent,
and averaging position-by-position is meaningful. patients fit to
*different* templates are not correspondent at all; nothing here is a
generic mesh-registration tool, it only ever averages meshes the caller has
already grouped by template (see api/results_bundle.py's nicp_template
column, the grouping key the cohort workspace actually uses).

kept deliberately dependency-light (no pandas) - openpyxl (already a
dependency for the cohort spreadsheet writer) is enough to read the sheet
back into plain {column: value} rows; anything past that (grouping,
descriptive stats, plotting) is the cohort workspace's own job, not this
module's.
"""

from __future__ import annotations

import sys
from dataclasses import dataclass
from pathlib import Path
from typing import BinaryIO

import numpy as np
import openpyxl
import trimesh

from .asymmetry import AsymmetryResult, calculate_asymmetry
from .craniometrics import (
    CranioMeasurements,
    FrontalBossingResult,
    _select_forehead_half,
    extract_measurements,
    find_hc_slice_height,
    frontal_bossing,
    hc_slice_polygon,
)
from .io import load_mesh
from .metopic import MetopicResult, analyze_forehead
from .registration.rigid import FACE_REFERENCE_TRIANGLE, REFERENCE_TRIANGLE

# same frozen-exe-vs-source-tree split as template_registry.py's
# TEMPLATES_DIR - see that module's comment for why __file__ alone isn't
# enough once this is running out of a pyinstaller exe. demo_cohort/ is
# package data (committed under src/craniumpy_core/, generated once by
# scripts/generate_demo_cohort.py - not regenerated at runtime), not a
# top-level docs resource, since it needs this exact same resolution.
if getattr(sys, "frozen", False):
    DEMO_COHORT_DIR = Path(sys._MEIPASS) / "craniumpy_core" / "demo_cohort"
else:
    DEMO_COHORT_DIR = Path(__file__).resolve().parent / "demo_cohort"


def load_cohort_xlsx(source: str | Path | BinaryIO) -> tuple[list[str], list[dict[str, str]]]:
    """(columns, rows) from a cohort spreadsheet - source is anything
    openpyxl.load_workbook accepts: a real path (desktop, see
    api/routers/cohort.py's /load) or a file-like object such as an
    uploaded file's BytesIO (browser, see that router's /upload). every
    cell comes back as a plain string, even the numeric ones - the cohort
    workspace does its own per-column numeric parsing rather than trusting
    openpyxl's per-cell typing, since a column can be sparse (some rows
    genuinely blank, others real numbers) and a lone blank cell shouldn't
    make an otherwise-numeric column look non-numeric. rows that are
    entirely blank (a stray trailing empty row some spreadsheet tools leave
    behind) are skipped."""
    wb = openpyxl.load_workbook(source, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(cell) if cell is not None else "" for cell in next(rows_iter)]
    rows = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        rows.append({header[i]: ("" if v is None else str(v)) for i, v in enumerate(values)})
    return header, rows


@dataclass
class MeanShapeResult:
    mesh: trimesh.Trimesh
    # per-vertex mean Euclidean distance from that vertex's own mean
    # position, across the input meshes (mm) - how much a given point on
    # the surface actually wandered from patient to patient, directly
    # usable as a heatmap (same units/shape the asymmetry heatmap already
    # uses - see craniumpy_core.asymmetry).
    variability: np.ndarray
    source_count: int


def _validate_same_topology(mesh: trimesh.Trimesh, reference: trimesh.Trimesh, label: str, reference_label: str) -> None:
    """same vertex count AND face connectivity as `reference` - the two
    things register_template's own guarantee (see this module's docstring)
    promises are identical for any two meshes fit to the same template.
    shared by mean_shape (checking each input against the first) and
    reference_diff (checking the computed mean against a reference
    template) - both are the same underlying question ("are these two
    meshes actually vertex-correspondent"), just applied to a different
    pair."""
    if len(mesh.vertices) != len(reference.vertices):
        raise ValueError(
            f"{label} has {len(mesh.vertices)} vertices, {reference_label} has {len(reference.vertices)} - "
            "these aren't the same template topology, so they can't be compared vertex-by-vertex"
        )
    reference_faces = np.asarray(reference.faces)
    if mesh.faces.shape != reference_faces.shape or not np.array_equal(np.asarray(mesh.faces), reference_faces):
        raise ValueError(
            f"{label} has different face connectivity than {reference_label} despite matching vertex count - "
            "these aren't the same template topology"
        )


def mean_shape(mesh_paths: list[Path]) -> MeanShapeResult:
    """vertex-by-vertex average of a set of same-topology (same-template)
    meshes. raises ValueError, naming the offending file, if any mesh
    doesn't match the first one's vertex count or face connectivity - the
    real-world failure mode of accidentally mixing two different
    templates' outputs, which would otherwise silently average unrelated
    vertex indices together and produce a meaningless mesh with no error
    at all."""
    if not mesh_paths:
        raise ValueError("no mesh paths given")

    meshes = [load_mesh(p) for p in mesh_paths]
    reference = meshes[0]

    for path, mesh in zip(mesh_paths, meshes):
        _validate_same_topology(mesh, reference, str(path), str(mesh_paths[0]))

    stacked = np.stack([np.asarray(m.vertices) for m in meshes])  # (n_meshes, n_vertices, 3)
    mean_vertices = stacked.mean(axis=0)
    variability = np.linalg.norm(stacked - mean_vertices, axis=2).mean(axis=0)

    mean_mesh = trimesh.Trimesh(vertices=mean_vertices, faces=reference.faces, process=False)
    return MeanShapeResult(mesh=mean_mesh, variability=variability, source_count=len(meshes))


@dataclass
class ExcludedMesh:
    """one mesh mean_shape_with_outliers left out of the average, and why -
    surfaced back to the caller (see api/routers/cohort.py's /mean-shape-qc)
    so a freeform group of meshes (the Mean Shape workspace's own picker,
    unlike a cohort group that's already known to share one template) never
    silently drops a file without the user finding out."""

    path: str
    reason: str


def mean_shape_with_outliers(mesh_paths: list[Path]) -> tuple[MeanShapeResult, list[ExcludedMesh]]:
    """same vertex-by-vertex average as mean_shape, but never aborts on a
    topology mismatch - excludes the offending mesh instead and reports why,
    for a freeform group of meshes with no pre-designated reference template
    (unlike mean_shape's own meshes[0]-is-reference assumption, which is
    fine when every input is already known-good, e.g. a cohort group already
    keyed by nicp_template).

    the REFERENCE topology here is whichever one the MAJORITY of the
    successfully-loaded meshes actually share (grouped by a cheap signature:
    vertex count + a hash of the face array) - not simply the first file.
    trusting meshes[0] the way mean_shape does would mean a single bad first
    file incorrectly excludes every genuinely-good one; majority vote is
    robust to that, and correctly identifies the true minority as the
    outliers. ties (two equally-sized topology groups) resolve to whichever
    group was encountered first while loading - deterministic, but not
    meaningful beyond that; a real tie is already an ambiguous input the
    caller should look at directly.

    a mesh that fails to even load (corrupt file, wrong format) is reported
    as excluded too, rather than raising and aborting every other mesh's
    chance to be averaged - same "one bad file never aborts the whole batch"
    principle api/routers/facial.py's own batch processing already follows.

    raises ValueError only if literally nothing could be averaged (every
    mesh failed to load, or no majority group could be formed at all)."""
    if not mesh_paths:
        raise ValueError("no mesh paths given")

    loaded: list[tuple[Path, trimesh.Trimesh]] = []
    excluded: list[ExcludedMesh] = []
    for path in mesh_paths:
        try:
            loaded.append((path, load_mesh(path)))
        except Exception as exc:  # noqa: BLE001 - any load failure excludes this one file, nothing more
            excluded.append(ExcludedMesh(path=str(path), reason=f"could not load: {exc}"))

    if not loaded:
        raise ValueError("no mesh could be loaded")

    def _signature(mesh: trimesh.Trimesh) -> tuple[int, int]:
        return len(mesh.vertices), hash(np.asarray(mesh.faces).tobytes())

    signatures = [_signature(mesh) for _, mesh in loaded]
    groups: dict[tuple[int, int], list[int]] = {}
    for i, sig in enumerate(signatures):
        groups.setdefault(sig, []).append(i)
    majority_sig = max(groups, key=lambda sig: len(groups[sig]))
    majority_vertex_count = majority_sig[0]

    kept_meshes: list[trimesh.Trimesh] = []
    reference: trimesh.Trimesh | None = None
    for (path, mesh), sig in zip(loaded, signatures):
        if sig == majority_sig:
            kept_meshes.append(mesh)
            if reference is None:
                reference = mesh
            continue
        if sig[0] != majority_vertex_count:
            reason = f"{sig[0]} vertices vs {majority_vertex_count} in the majority group"
        else:
            reason = "different face connectivity than the majority group despite matching vertex count"
        excluded.append(ExcludedMesh(path=str(path), reason=reason))

    if not kept_meshes:
        raise ValueError("no majority topology group could be formed - every loaded mesh has a different topology")

    stacked = np.stack([np.asarray(m.vertices) for m in kept_meshes])
    mean_vertices = stacked.mean(axis=0)
    variability = np.linalg.norm(stacked - mean_vertices, axis=2).mean(axis=0)
    mean_mesh = trimesh.Trimesh(vertices=mean_vertices, faces=reference.faces, process=False)
    return MeanShapeResult(mesh=mean_mesh, variability=variability, source_count=len(kept_meshes)), excluded


def reference_diff(mean_mesh: trimesh.Trimesh, reference_mesh: trimesh.Trimesh) -> np.ndarray:
    """signed per-vertex distance (mm) of mean_mesh from reference_mesh,
    projected onto the reference surface's own vertex normals - positive
    where the mean shape sits OUTWARD of the reference (the direction its
    surface normal points), negative where it sits INWARD. same sign
    convention and units as craniumpy_core.asymmetry's heatmap, so the
    cohort workspace's existing diverging (blue/red) heatmap rendering
    applies to this unchanged - unlike mean_shape's own variability array,
    which is an unsigned magnitude and gets a different (sequential) color
    scale in the frontend (see three/measurementsLayer.js's
    applySequentialHeatmap vs applyHeatmap).

    reference_mesh has to be the exact template mean_mesh's own patients
    were NICP-fitted to (or any other mesh with identical vertex count/face
    connectivity) - raises ValueError otherwise, same "a clear error beats
    a silently meaningless comparison" reasoning as mean_shape's own
    validation."""
    _validate_same_topology(mean_mesh, reference_mesh, "mean shape", "reference template")
    displacement = np.asarray(mean_mesh.vertices) - np.asarray(reference_mesh.vertices)
    normals = np.asarray(reference_mesh.vertex_normals)
    return np.einsum("ij,ij->i", displacement, normals)


@dataclass
class GroupMeasurements:
    craniometrics: CranioMeasurements | None
    asymmetry: AsymmetryResult
    metopic: MetopicResult | None
    frontal_bossing: FrontalBossingResult | None
    # the HC-equivalent slice height metopic was computed at - always set,
    # even though CranioMeasurements already carries its own copy of this
    # (as .slice_height) for the cranium branch, since MetopicResult itself
    # doesn't carry one (see api/routers/mesh.py's get_results, which pulls
    # the same value from session.hc_slice_height instead for that reason).
    slice_height: float


def measure_mean_shape(mesh: trimesh.Trimesh, target: str) -> GroupMeasurements:
    """runs the same measurement suite the single-patient pipeline does
    (see pipeline.measure_cranial/measure_facial) directly on a computed
    cohort mean shape - the "with the same metrics as the Patients
    workspace" view of the Mean shape tab.

    this works with NO per-group landmark tracking because every mesh
    averaged into a mean shape was itself already rigidly registered onto
    registration.rigid.REFERENCE_TRIANGLE before NICP fitting (see
    mean_shape's own docstring) - sellion/left_tragus/right_tragus sit at
    that exact fixed position for every such mesh, cranium or face target
    alike (face target subtracts an extra sellion-at-origin offset - see
    register() in pipeline.py - hence FACE_REFERENCE_TRIANGLE instead),
    and the mean of several meshes in the same frame is still in that
    frame. calculate_asymmetry needs no landmarks at all (it's a property
    of the mesh's own left/right shape - see that function).

    NOTE this is a close approximation, not exact, for patients whose
    export used CoM correction: a small per-patient depth-axis recenter is
    applied after registration (see pipeline.py's _recenter_com_z), which
    varies slightly per patient and isn't reproduced here - the practical
    effect is a few mm of slice-height uncertainty, not a wrong answer."""
    landmarks = REFERENCE_TRIANGLE if target == "cranium" else FACE_REFERENCE_TRIANGLE
    sellion = landmarks[0]

    asymmetry = calculate_asymmetry(mesh)

    if target == "cranium":
        metrics = extract_measurements(mesh, landmarks=landmarks)
        bossing = frontal_bossing(mesh, sellion, metrics.slice_height)
        return GroupMeasurements(
            craniometrics=metrics, asymmetry=asymmetry, metopic=None, frontal_bossing=bossing,
            slice_height=metrics.slice_height,
        )

    slice_height = find_hc_slice_height(mesh, landmarks)
    metopic = analyze_forehead(mesh, slice_height)
    bossing = frontal_bossing(mesh, sellion, slice_height)
    return GroupMeasurements(
        craniometrics=None, asymmetry=asymmetry, metopic=metopic, frontal_bossing=bossing, slice_height=slice_height,
    )


def _forehead_arc_yz(profile: np.ndarray, sellion: np.ndarray) -> np.ndarray:
    """the sellion-to-vertex forehead arc from a sagittal profile (see
    craniometrics.frontal_bossing's own `profile` field, an ordered walk
    of the whole x=sellion.x section) - the exact same "largest contiguous
    run above sellion, then split at the crown and keep the forehead-side
    half" extraction frontal_bossing uses internally to find its own
    frontal_point (see that function and craniometrics._select_forehead_half),
    factored out here since frontal_bossing doesn't expose the arc alone.
    the _select_forehead_half step matters, not just the above-sellion
    mask: on a closed cranial cap, the "above sellion" run sweeps in one
    unbroken arc from the forehead's base, over the vertex, all the way
    down to the OCCIPUT's base too (both sit above sellion's height) -
    without splitting at the crown, this would silently average forehead
    points from one patient against occipital points from another
    wherever their y-ranges happen to overlap.

    returns (y, z) pairs sorted by y (ascending, sellion-height first)."""
    above = profile[:, 1] > sellion[1]
    if not above.any():
        return np.empty((0, 2))
    indices = np.where(above)[0]
    splits = np.where(np.diff(indices) > 1)[0] + 1
    arc = profile[max(np.split(indices, splits), key=len)]
    arc = _select_forehead_half(arc, sellion)
    return arc[:, 1:]  # (y, z) - already ascending in y, see _select_forehead_half's own docstring


@dataclass
class SagittalMidlineBand:
    y: np.ndarray  # common height grid (mm), ascending
    mean_z: np.ndarray  # mean depth (mm) at each y, across the group
    sd_z: np.ndarray  # sample SD of depth (mm) at each y, across the group
    source_count: int


def sagittal_midline_band(mesh_paths: list[Path], target: str, n_samples: int = 60) -> SagittalMidlineBand:
    """mean +/- SD of the sagittal (midline) forehead-to-vertex depth
    profile across a group of same-template NICP-fitted meshes - "how much
    does the surface bulge or recede at each height, on average, and how
    much does that vary from patient to patient" (see
    craniometrics.frontal_bossing's own `profile` field, which this reuses
    per-mesh).

    each patient's own arc is resampled (linear interpolation) onto a
    single, group-wide height grid before averaging - unlike mean_shape's
    vertex-by-vertex average, there's no shared per-vertex index to average
    directly here: the sagittal section plane cuts through different mesh
    EDGES for each patient (even though the underlying mesh topology is
    identical), so each patient's raw profile has a different length and
    parametrization. the grid spans only the height range every patient's
    own arc actually covers (the tallest patient's forehead-to-vertex arc
    almost never starts as low, or ends as high, as the shortest one's) -
    extrapolating past what a given patient's data actually shows would be
    inventing that patient's shape past where their arc ends.

    raises ValueError if the meshes aren't the same template topology (via
    the same check mean_shape uses), if no mesh's own forehead arc could be
    extracted at all, or if the group's own arcs don't overlap enough in
    height to average meaningfully (e.g. one mesh's clip is much shorter
    than the rest)."""
    landmarks = REFERENCE_TRIANGLE if target == "cranium" else FACE_REFERENCE_TRIANGLE
    sellion = landmarks[0]

    if not mesh_paths:
        raise ValueError("no mesh paths given")
    meshes = [load_mesh(p) for p in mesh_paths]
    reference = meshes[0]
    for path, mesh in zip(mesh_paths, meshes):
        _validate_same_topology(mesh, reference, str(path), str(mesh_paths[0]))

    arcs = []
    for mesh in meshes:
        slice_height = (
            extract_measurements(mesh, landmarks=landmarks).slice_height
            if target == "cranium"
            else find_hc_slice_height(mesh, landmarks)
        )
        bossing = frontal_bossing(mesh, sellion, slice_height)
        if bossing is None:
            continue
        arc = _forehead_arc_yz(bossing.profile, bossing.sellion)
        if len(arc) >= 2:
            arcs.append(arc)

    if not arcs:
        raise ValueError("no sagittal forehead profile could be extracted from any mesh in this group")

    y_lo = max(arc[0, 0] for arc in arcs)
    y_hi = min(arc[-1, 0] for arc in arcs)
    if y_hi <= y_lo:
        raise ValueError("these meshes' forehead profiles don't overlap enough in height to average")

    y_grid = np.linspace(y_lo, y_hi, n_samples)
    z_values = np.stack([np.interp(y_grid, arc[:, 0], arc[:, 1]) for arc in arcs])
    return SagittalMidlineBand(
        y=y_grid, mean_z=z_values.mean(axis=0), sd_z=z_values.std(axis=0), source_count=len(arcs),
    )


@dataclass
class SpreadBand:
    """a +/-1 SD ribbon around some mean curve on the mean shape's own
    surface - the shared 3D shape hc_ring_band/metopic_band/
    sagittal_band_to_spread_band all return, so the frontend's live 3D
    viewer overlay and the PDF report only need ONE rendering path for all
    three bands (see three/spreadBandOverlay.js's addSpreadBandRibbon and
    api/results_bundle.py's _draw_measurements/_draw_metopic/
    _draw_frontal_bossing, which all take one of these as an optional
    "spread_band"/"sagittal_band" kwarg)."""

    mean: np.ndarray  # (N, 3) - the mean curve itself, same points the plain (non-banded) overlay already draws
    inner: np.ndarray  # (N, 3) - mean shifted by -1 SD along this band's own "spread direction"
    outer: np.ndarray  # (N, 3) - mean shifted by +1 SD
    closed: bool  # True for the HC ring (a closed loop), False for the sagittal/metopic arcs
    source_count: int


def sagittal_band_to_spread_band(band: SagittalMidlineBand, target: str) -> SpreadBand:
    """embeds sagittal_midline_band's own (y, mean_z, sd_z) into 3D points,
    at x = sellion.x (~0 for either target - see registration.rigid's
    REFERENCE_TRIANGLE/FACE_REFERENCE_TRIANGLE) - the sagittal plane every
    frontal_bossing profile is already drawn in. kept separate from
    sagittal_midline_band itself since that function's own (y, mean_z,
    sd_z) return shape is what the Mean shape tab's 2D profile chart
    needs; this is only for the 3D ribbon overlay/report, which needs
    real points instead."""
    x = float((REFERENCE_TRIANGLE if target == "cranium" else FACE_REFERENCE_TRIANGLE)[0][0])
    n = len(band.y)
    mean = np.column_stack([np.full(n, x), band.y, band.mean_z])
    inner = np.column_stack([np.full(n, x), band.y, band.mean_z - band.sd_z])
    outer = np.column_stack([np.full(n, x), band.y, band.mean_z + band.sd_z])
    return SpreadBand(mean=mean, inner=inner, outer=outer, closed=False, source_count=band.source_count)


def hc_ring_band(mesh_paths: list[Path], target: str, n_samples: int = 72) -> SpreadBand:
    """+/-1 SD ribbon around the HC ring (see craniometrics.hc_slice_polygon,
    the same red ring _draw_measurements already plots), across a group of
    same-template NICP-fitted meshes.

    each patient's own ring is resampled (circular interpolation, since
    the ring wraps around) onto a common angle grid before averaging -
    same "different mesh, different section-plane intersection points, no
    shared per-vertex index to average directly" reasoning
    sagittal_midline_band uses for its own arc. the ring is drawn at the
    GROUP's mean slice height (each patient's own height varies slightly
    - see extract_measurements' own docstring on why the height search is
    height-dependent) rather than any one patient's - a single flat
    reference height for the whole band, since the underlying radius
    values it draws already average over that same variation.

    raises ValueError if the meshes aren't the same template topology, or
    if no mesh's own HC ring could be extracted at all."""
    landmarks = REFERENCE_TRIANGLE if target == "cranium" else FACE_REFERENCE_TRIANGLE

    if not mesh_paths:
        raise ValueError("no mesh paths given")
    meshes = [load_mesh(p) for p in mesh_paths]
    reference = meshes[0]
    for path, mesh in zip(mesh_paths, meshes):
        _validate_same_topology(mesh, reference, str(path), str(mesh_paths[0]))

    angle_grid = np.linspace(-np.pi, np.pi, n_samples, endpoint=False)
    rho_samples = []
    heights = []
    for mesh in meshes:
        slice_height = (
            extract_measurements(mesh, landmarks=landmarks).slice_height
            if target == "cranium"
            else find_hc_slice_height(mesh, landmarks)
        )
        polygon = hc_slice_polygon(mesh, slice_height)
        if polygon is None or len(polygon) < 3:
            continue
        phi = np.arctan2(polygon[:, 2], polygon[:, 0])
        rho = np.hypot(polygon[:, 0], polygon[:, 2])
        order = np.argsort(phi)
        phi, rho = phi[order], rho[order]
        # wrap the ring around both ends so interpolation across the +/-pi
        # seam (where a plain np.interp would otherwise just clamp to the
        # nearest real endpoint instead of blending across the wraparound)
        # sees real neighboring data on both sides.
        phi_ext = np.concatenate([phi - 2 * np.pi, phi, phi + 2 * np.pi])
        rho_ext = np.tile(rho, 3)
        rho_samples.append(np.interp(angle_grid, phi_ext, rho_ext))
        heights.append(slice_height)

    if not rho_samples:
        raise ValueError("no HC slice could be extracted from any mesh in this group")

    rho_stack = np.stack(rho_samples)
    mean_rho = rho_stack.mean(axis=0)
    sd_rho = rho_stack.std(axis=0)
    mean_height = float(np.mean(heights))

    def to_points(rho: np.ndarray) -> np.ndarray:
        return np.column_stack(
            [rho * np.cos(angle_grid), np.full(n_samples, mean_height), rho * np.sin(angle_grid)]
        )

    return SpreadBand(
        mean=to_points(mean_rho), inner=to_points(mean_rho - sd_rho), outer=to_points(mean_rho + sd_rho),
        closed=True, source_count=len(rho_samples),
    )


def metopic_band(mesh_paths: list[Path], target: str, n_samples: int = 60) -> SpreadBand:
    """+/-1 SD ribbon around the metopic forehead contour (see
    metopic.analyze_forehead's own `contour` field, the same curve
    _draw_metopic already plots), across a group of same-template
    NICP-fitted meshes.

    each patient's own contour is resampled (linear interpolation) onto a
    common x grid before averaging, covering only the x-range every
    patient's own contour actually spans - same "don't invent a patient's
    shape past where their own data ends" reasoning sagittal_midline_band
    uses for its own height range.

    raises ValueError if the meshes aren't the same template topology, if
    no mesh's own metopic contour could be extracted at all, or if the
    group's own contours don't overlap enough in x to average
    meaningfully."""
    landmarks = REFERENCE_TRIANGLE if target == "cranium" else FACE_REFERENCE_TRIANGLE

    if not mesh_paths:
        raise ValueError("no mesh paths given")
    meshes = [load_mesh(p) for p in mesh_paths]
    reference = meshes[0]
    for path, mesh in zip(mesh_paths, meshes):
        _validate_same_topology(mesh, reference, str(path), str(mesh_paths[0]))

    contours = []
    heights = []
    for mesh in meshes:
        slice_height = (
            extract_measurements(mesh, landmarks=landmarks).slice_height
            if target == "cranium"
            else find_hc_slice_height(mesh, landmarks)
        )
        metopic_result = analyze_forehead(mesh, slice_height)
        if metopic_result is None:
            continue
        x, z = metopic_result.contour[:, 0], metopic_result.contour[:, 1]
        order = np.argsort(x)
        contours.append((x[order], z[order]))
        heights.append(slice_height)

    if not contours:
        raise ValueError("no metopic contour could be extracted from any mesh in this group")

    x_lo = max(x[0] for x, _ in contours)
    x_hi = min(x[-1] for x, _ in contours)
    if x_hi <= x_lo:
        raise ValueError("these meshes' forehead contours don't overlap enough in x to average")

    x_grid = np.linspace(x_lo, x_hi, n_samples)
    z_values = np.stack([np.interp(x_grid, x, z) for x, z in contours])
    mean_z = z_values.mean(axis=0)
    sd_z = z_values.std(axis=0)
    mean_height = float(np.mean(heights))

    def to_points(z: np.ndarray) -> np.ndarray:
        return np.column_stack([x_grid, np.full(n_samples, mean_height), z])

    return SpreadBand(
        mean=to_points(mean_z), inner=to_points(mean_z - sd_z), outer=to_points(mean_z + sd_z),
        closed=False, source_count=len(contours),
    )


def load_demo_cohort() -> tuple[list[str], list[dict[str, str]]]:
    """the shipped synthetic demo cohort (see scripts/generate_demo_cohort.py
    for how it's built) - same shape as load_cohort_xlsx, except
    nicp_mesh_path is stored in the sheet as a path relative to
    DEMO_COHORT_DIR (so the committed asset is relocatable/reproducible
    across machines) and gets rewritten here into a real absolute path,
    exactly like every other cohort row's nicp_mesh_path already is (see
    api/results_bundle.py's _nicp_mesh_path) - so nothing downstream (the
    Mean shape tab, /api/cohort/mean-shape) has to know this cohort is
    synthetic at all."""
    xlsx_path = DEMO_COHORT_DIR / "demo_cohort.xlsx"
    if not xlsx_path.is_file():
        raise FileNotFoundError(
            f"no demo cohort at {xlsx_path} - run scripts/generate_demo_cohort.py once to generate it"
        )
    columns, rows = load_cohort_xlsx(xlsx_path)
    for row in rows:
        relative = row.get("nicp_mesh_path", "")
        if relative:
            row["nicp_mesh_path"] = str((DEMO_COHORT_DIR / relative).resolve())
    return columns, rows
