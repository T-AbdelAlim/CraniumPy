"""reads back the Facial Anthropometrics workspace's own batch measurement
export (see api/routers/facial.py's export_batch) for loading into the
Cohort workspace as an attached dataset. mirrors craniumpy_core.cohort's
own load_cohort_xlsx shape/style - openpyxl-only, every cell a plain
string, blank trailing rows skipped - just reading two sheets instead of
the cohort loader's one.
"""

from __future__ import annotations

from pathlib import Path
from typing import BinaryIO

import openpyxl


def _read_sheet_rows(ws) -> tuple[list[str], list[dict[str, str]]]:
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(cell) if cell is not None else "" for cell in next(rows_iter)]
    rows = []
    for values in rows_iter:
        if all(v is None for v in values):
            continue
        rows.append({header[i]: ("" if v is None else str(v)) for i, v in enumerate(values)})
    return header, rows


def load_measurement_export_xlsx(source: str | Path | BinaryIO) -> tuple[list[str], list[dict[str, str]], list[dict[str, str]]]:
    """(measurement_columns, measurement_rows, legend_rows) - the
    "measurements" and "legend" sheets api/routers/facial.py's export_batch
    writes. a measurement row's own "identifier" column is the source
    filename it was measured from - the join key
    facial_cohort_link.resolve_cohort_ids_by_filename matches against.
    raises ValueError (not some openpyxl-internal error) if the file isn't
    actually one of these exports - the "measurements" sheet is mandatory,
    "legend" is read if present but not required."""
    wb = openpyxl.load_workbook(source, data_only=True)
    if "measurements" not in wb.sheetnames:
        raise ValueError("not a Facial Anthropometrics measurement export - no 'measurements' sheet found")
    columns, rows = _read_sheet_rows(wb["measurements"])
    legend_rows = _read_sheet_rows(wb["legend"])[1] if "legend" in wb.sheetnames else []
    return columns, rows, legend_rows
