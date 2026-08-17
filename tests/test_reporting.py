"""unit tests for the summary-spreadsheet/cohort-spreadsheet pieces of
api/results_bundle.py - pure function tests, no FastAPI/mesh pipeline
needed, same style as test_craniometrics.py/test_metopic.py testing the
underlying math directly.

_report_pdf's own layout isn't covered here beyond the line-advance
arithmetic below - the API-level tests in test_api.py exercise the whole
thing by opening the PDF that comes out of a real export."""

from __future__ import annotations

import io
from pathlib import Path

import numpy as np
import pytest
import trimesh
from openpyxl import Workbook, load_workbook

from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.figure import Figure
from shapely.geometry import MultiPoint, Polygon

from api.results_bundle import (
    PAGE_H_IN,
    PAGE_W_IN,
    _PDF_METRIC_FIELDS,
    _draw_asymmetry,
    _draw_frontal_bossing,
    _draw_measurements,
    _draw_metopic,
    _id_mapping_path,
    _metrics_row,
    _silhouette_polygon,
    _summary_xlsx,
    _upsert_cohort_xlsx,
)
from craniumpy_core.asymmetry import AsymmetryResult
from craniumpy_core.craniometrics import CranioMeasurements, FrontalBossingResult
from craniumpy_core.metopic import MetopicResult


def _config(com_translation: bool = True, nicp: dict | None = None) -> dict:
    return {"com_translation": com_translation, "nicp": nicp}


def _craniometrics() -> CranioMeasurements:
    return CranioMeasurements(
        slice_height=10.0,
        depth_mm=59.1234,
        breadth_mm=67.001,
        cephalic_index=113.4,
        circumference_cm=19.7,
        mesh_volume_cc=9.05,
        front_opt=np.zeros(3),
        occ_opt=np.zeros(3),
        lh_opt=np.zeros(3),
        rh_opt=np.zeros(3),
    )


def _asymmetry() -> AsymmetryResult:
    return AsymmetryResult(heatmap=np.zeros(5), mean_asymmetry_index=4.401)


def _frontal_bossing() -> FrontalBossingResult:
    return FrontalBossingResult(
        angle_deg=159.026, sellion=np.zeros(3), frontal_point=np.zeros(3), profile=np.zeros((3, 3))
    )


def _metopic() -> MetopicResult:
    return MetopicResult(
        contour=np.zeros((3, 2)),
        arc_length=np.zeros(3),
        normalized_arc_length=np.zeros(3),
        midline_u=0.5,
        parabola_a=-0.01,
        parabola_c=30.0,
        deviation_profile=np.zeros(3),
        gradient_profile=np.zeros(3),
        curvature_profile=np.zeros(3),
        frontal_angle_deg=120.456,
        frontal_angle_points=(np.zeros(2), np.zeros(2), np.zeros(2)),
        forehead_width_mm=80.0,
        midline_curvature_concentration=0.3,
        midline_max_curvature=0.01,
        midline_max_curvature_position=0.5,
        ridge_protrusion_mm=2.5,
        ridge_protrusion_position=0.5,
        ridge_area_mm2=120.0,
        ridge_area_normalized=0.05,
        left_temporal_hollowing=0.1,
        right_temporal_hollowing=0.05,
        mean_temporal_hollowing=0.075,
        left_max_temporal_depth_mm=1.2,
        right_max_temporal_depth_mm=0.8,
        parabolic_deviation_index=0.9,
        central_window=(0.4, 0.6),
        left_temporal_window=(0.1, 0.3),
        right_temporal_window=(0.7, 0.9),
    )


_METADATA_KEYS = (
    "file_name", "file_path", "patient_id", "diagnosis", "sex", "date_imaging", "age_imaging",
    "image_timing", "treatment", "age_surgery_months", "free_variable",
)
_SETTINGS_KEYS = ("com_correction", "nicp_used", "nicp_template")


# --- _metrics_row -----------------------------------------------------


def test_metrics_row_blank_metadata_and_missing_metrics_come_through_empty_not_omitted():
    row = _metrics_row("cranium", {}, _config(com_translation=False), None, None, None, None)

    for key in _METADATA_KEYS:
        assert row[key] == ""
    assert row["target"] == "cranium"
    for key in ("depth_mm", "breadth_mm", "cephalic_index", "circumference_cm", "mesh_volume_cc",
                "mean_asymmetry_index", "frontal_bossing_angle_deg"):
        assert row[key] == ""
    assert all(v == "" for k, v in row.items() if k.startswith("metopic_"))


def test_metrics_row_echoes_metadata_fields_verbatim():
    metadata = {key: f"value-{key}" for key in _METADATA_KEYS}
    row = _metrics_row("face", metadata, _config(), None, None, None, None)
    for key in _METADATA_KEYS:
        assert row[key] == f"value-{key}"


def test_metrics_row_only_populates_applicable_metric_groups():
    # cranial-style session: craniometrics + frontal_bossing present,
    # asymmetry/metopic absent (as a real facial-only pair would be) - the
    # row must still carry every column, just blank for the absent ones.
    row = _metrics_row("cranium", {}, _config(), _craniometrics(), None, None, _frontal_bossing())

    assert row["depth_mm"] == "59.12"
    assert row["breadth_mm"] == "67.00"
    assert row["frontal_bossing_angle_deg"] == "159.03"
    assert row["mean_asymmetry_index"] == ""
    assert all(v == "" for k, v in row.items() if k.startswith("metopic_"))


def test_metrics_row_populates_metopic_and_asymmetry_when_given():
    row = _metrics_row("face", {}, _config(), None, _asymmetry(), _metopic(), None)

    assert row["mean_asymmetry_index"] == "4.40"
    assert row["metopic_frontal_angle_deg"] == "120.46"
    assert row["metopic_ridge_protrusion_mm"] == "2.50"
    assert row["depth_mm"] == ""
    assert row["frontal_bossing_angle_deg"] == ""


def test_metrics_row_settings_reflect_com_and_nicp_off():
    row = _metrics_row("cranium", {}, _config(com_translation=False, nicp=None), None, None, None, None)

    assert row["com_correction"] == "no"
    assert row["nicp_used"] == "no"
    assert row["nicp_template"] == ""
    # defaults blank when the caller has no real dest_dir to resolve a
    # written mesh path against (e.g. the browser/session-summary path -
    # see results_bundle._build_analysis_files, which never passes this)
    assert row["nicp_mesh_path"] == ""


def test_metrics_row_nicp_mesh_path_passed_through_verbatim():
    # the two desktop writers that DO know a real dest_dir resolve this
    # themselves (see results_bundle._nicp_mesh_path) and hand it in - this
    # just confirms _metrics_row doesn't touch/reformat it.
    row = _metrics_row(
        "cranium", {}, _config(nicp={"template": "cranium_com"}), None, None, None, None,
        nicp_mesh_path="/data/results/patient_rg_CN.ply",
    )
    assert row["nicp_mesh_path"] == "/data/results/patient_rg_CN.ply"


def test_metrics_row_settings_reflect_com_and_nicp_on_with_shipped_template():
    row = _metrics_row(
        "cranium", {}, _config(com_translation=True, nicp={"template": "cranium_com", "custom_template_path": None}),
        None, None, None, None,
    )

    assert row["com_correction"] == "yes"
    assert row["nicp_used"] == "yes"
    assert row["nicp_template"] == "cranium_com"


def test_metrics_row_settings_nicp_template_uses_filename_not_full_path():
    row = _metrics_row(
        "cranium", {},
        _config(nicp={"template": None, "custom_template_path": "/data/templates/my_template.ply"}),
        None, None, None, None,
    )

    assert row["nicp_used"] == "yes"
    assert row["nicp_template"] == "my_template.ply"


# --- _summary_xlsx -------------------------------------------------------


def _read_sheet(xlsx_bytes: bytes) -> list[dict[str, str]]:
    """mirrors what a numeric cell actually displays in Excel (2 decimals,
    per the number_format _write_xlsx_rows applies), not Python's own
    str() - which drops the decimals entirely for a whole number (openpyxl
    hands a whole-valued numeric cell back as int, e.g. 61 not 61.0) or
    drops a trailing zero for a fraction (str(70.5) == "70.5", not
    "70.50") - either way not what a person actually sees in the sheet,
    and not what an exact round-trip comparison against a _fmt(...) string
    like "61.00" should be checked against."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    header = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        cells = {}
        for i, v in enumerate(excel_row):
            if v is None:
                cells[header[i]] = ""
            elif isinstance(v, (int, float)):
                cells[header[i]] = f"{v:.2f}"
            else:
                cells[header[i]] = str(v)
        rows.append(cells)
    return rows


def test_summary_xlsx_round_trips_the_row():
    row = _metrics_row("cranium", {"file_name": "patient1.ply", "sex": "female"}, _config(), _craniometrics(), None, None, None)
    xlsx_bytes = _summary_xlsx(row)

    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
    assert [c.value for c in header_cells] == list(row.keys())
    assert header_cells[0].font.bold  # sanity: the header row is styled, not plain text

    rows = _read_sheet(xlsx_bytes)
    assert len(rows) == 1
    assert rows[0] == row


# --- _upsert_cohort_xlsx ---------------------------------------------------


def _read_cohort(path: Path) -> list[dict[str, str]]:
    return _read_sheet(path.read_bytes())


def _read_mapping(cohort_path: Path) -> list[dict[str, str]]:
    return _read_sheet(_id_mapping_path(cohort_path).read_bytes())


def test_upsert_cohort_xlsx_creates_file_with_header_and_row(tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    row = _metrics_row(
        "cranium", {"file_name": "a.ply", "file_path": "/data/a.ply", "patient_id": "MRN-1"},
        _config(), _craniometrics(), None, None, None,
    )

    _upsert_cohort_xlsx(cohort_path, row)

    assert cohort_path.exists()
    rows = _read_cohort(cohort_path)
    assert len(rows) == 1
    # the shared cohort file gets a cohort_id in place of the local
    # patient_id - see test_upsert_cohort_xlsx_writes_local_id_mapping_file
    # for where patient_id actually ends up
    assert rows[0]["cohort_id"] == "C00001"
    assert "patient_id" not in rows[0]
    assert rows[0]["file_path"] == "/data/a.ply"
    assert rows[0]["depth_mm"] == row["depth_mm"]


def test_upsert_cohort_xlsx_writes_local_id_mapping_file(tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    row = _metrics_row(
        "cranium", {"file_name": "a.ply", "file_path": "/data/a.ply", "patient_id": "MRN-1"},
        _config(), _craniometrics(), None, None, None,
    )

    _upsert_cohort_xlsx(cohort_path, row)

    mapping_path = _id_mapping_path(cohort_path)
    assert mapping_path.exists()
    assert mapping_path != cohort_path
    mapping_rows = _read_mapping(cohort_path)
    assert len(mapping_rows) == 1
    assert mapping_rows[0]["cohort_id"] == "C00001"
    assert mapping_rows[0]["patient_id"] == "MRN-1"
    assert mapping_rows[0]["file_path"] == "/data/a.ply"


def test_upsert_cohort_xlsx_appends_a_genuinely_new_row(tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    row_a = _metrics_row("cranium", {"file_name": "a.ply", "file_path": "/data/a.ply"}, _config(), _craniometrics(), None, None, None)
    row_b = _metrics_row("face", {"file_name": "b.ply", "file_path": "/data/b.ply"}, _config(), None, _asymmetry(), None, None)

    _upsert_cohort_xlsx(cohort_path, row_a)
    _upsert_cohort_xlsx(cohort_path, row_b)

    rows = _read_cohort(cohort_path)
    assert len(rows) == 2
    assert rows[0]["file_path"] == "/data/a.ply"
    assert rows[1]["file_path"] == "/data/b.ply"
    # each distinct file gets its own, sequential cohort_id
    assert rows[0]["cohort_id"] == "C00001"
    assert rows[1]["cohort_id"] == "C00002"


def test_upsert_cohort_xlsx_replaces_row_with_matching_file_path_instead_of_duplicating(tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    first = _metrics_row("cranium", {"file_name": "a.ply", "file_path": "/data/a.ply"}, _config(), _craniometrics(), None, None, None)
    _upsert_cohort_xlsx(cohort_path, first)

    updated_craniometrics = _craniometrics()
    updated_craniometrics.depth_mm = 61.0
    second = _metrics_row("cranium", {"file_name": "a.ply", "file_path": "/data/a.ply"}, _config(), updated_craniometrics, None, None, None)
    _upsert_cohort_xlsx(cohort_path, second)

    rows = _read_cohort(cohort_path)
    assert len(rows) == 1
    assert rows[0]["depth_mm"] == "61.00"
    # re-exporting the same file reuses its existing cohort_id rather than
    # handing out a new one
    assert rows[0]["cohort_id"] == "C00001"


def test_upsert_cohort_xlsx_falls_back_to_file_name_when_file_path_blank(tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    first = _metrics_row("cranium", {"file_name": "a.ply"}, _config(), _craniometrics(), None, None, None)
    _upsert_cohort_xlsx(cohort_path, first)

    updated_craniometrics = _craniometrics()
    updated_craniometrics.depth_mm = 70.0
    second = _metrics_row("cranium", {"file_name": "a.ply"}, _config(), updated_craniometrics, None, None, None)
    _upsert_cohort_xlsx(cohort_path, second)

    rows = _read_cohort(cohort_path)
    assert len(rows) == 1
    assert rows[0]["depth_mm"] == "70.00"


def test_upsert_cohort_xlsx_unions_columns_from_an_older_narrower_schema(tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    # simulate an older cohort file with a column today's schema no longer
    # writes (e.g. a since-removed metric) plus a normal row
    wb = Workbook()
    ws = wb.active
    ws.append(["file_path", "file_name", "an_old_removed_column"])
    ws.append(["/data/old.ply", "old.ply", "legacy-value"])
    wb.save(cohort_path)

    new_row = _metrics_row("cranium", {"file_name": "b.ply", "file_path": "/data/b.ply"}, _config(), _craniometrics(), None, None, None)
    _upsert_cohort_xlsx(cohort_path, new_row)

    wb = load_workbook(cohort_path)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "an_old_removed_column" in header
    assert "depth_mm" in header

    rows = _read_cohort(cohort_path)
    old_row = next(r for r in rows if r["file_path"] == "/data/old.ply")
    assert old_row["an_old_removed_column"] == "legacy-value"
    assert old_row["depth_mm"] == ""  # new column, blank on the untouched old row

    new_row_out = next(r for r in rows if r["file_path"] == "/data/b.ply")
    assert new_row_out["an_old_removed_column"] == ""  # old column, blank on the new row
    assert new_row_out["depth_mm"] == "59.12"
    assert new_row_out["cohort_id"] == "C00001"


# --- PDF figure layout ---------------------------------------------------
#
# regression tests for text overlapping between a plot's own x-axis label
# (matplotlib-placed, using real font metrics - not something these
# functions control directly the way the fixed line-advance PDF text
# elsewhere in this module is) and the fig.legend/caption text manually
# positioned below it. rendered at actual PDF page scale (the report page
# rect, not a standalone-PNG-sized one) since that's the scale the
# overlap only showed up at - see _draw_measurements/_draw_frontal_bossing.


def _text_block_extents(fig: Figure) -> tuple[float, float, float, float, float, float]:
    """(xlabel_bottom, legend_top, legend_bottom, caption_top), all as a
    fraction of the whole figure's height (0 = figure bottom), plus the
    two gaps between them - overlap means the gap is negative. assumes
    exactly one axes, one legend, and the caption is whichever fig.text is
    lowest (both figures under test have just the one)."""
    canvas = FigureCanvasAgg(fig)
    canvas.draw()
    renderer = canvas.get_renderer()
    fig_h = fig.bbox.height

    xlabel_bottom = fig.axes[0].xaxis.label.get_window_extent(renderer).y0 / fig_h
    legend_bb = fig.legends[0].get_window_extent(renderer)
    legend_top, legend_bottom = legend_bb.y1 / fig_h, legend_bb.y0 / fig_h
    caption = min(fig.texts, key=lambda t: t.get_window_extent(renderer).y0)
    caption_top = caption.get_window_extent(renderer).y1 / fig_h

    return (
        xlabel_bottom, legend_top, legend_bottom, caption_top,
        xlabel_bottom - legend_top, legend_bottom - caption_top,
    )


def test_measurements_figure_xlabel_legend_caption_dont_overlap():
    mesh = trimesh.creation.icosphere(subdivisions=3, radius=1.0)
    mesh.vertices = np.asarray(mesh.vertices) * np.array([70.0, 90.0, 60.0])
    measurements = _craniometrics()

    page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
    _draw_measurements(page, (0.04, 0.46, 0.92, 0.50), mesh, measurements)

    _, _, _, _, xlabel_legend_gap, legend_caption_gap = _text_block_extents(page)
    assert xlabel_legend_gap >= 0
    assert legend_caption_gap >= 0


def test_frontal_bossing_figure_xlabel_legend_caption_dont_overlap():
    y = np.linspace(0.0, 100.0, 300)
    z = -0.005 * (y - 40.0) ** 2 + 20.0
    profile = np.column_stack([np.zeros_like(y), y, z])
    bossing = FrontalBossingResult(
        angle_deg=68.3, sellion=np.array([0.0, 0.0, 20.0]), frontal_point=np.array([0.0, 40.0, 20.0]),
        profile=profile, horizontal=np.array([0.0, 0.0, 1.0]), slice_height=40.0,
    )

    page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
    _draw_frontal_bossing(page, (0.04, 0.46, 0.92, 0.50), bossing)

    _, _, _, _, xlabel_legend_gap, legend_caption_gap = _text_block_extents(page)
    assert xlabel_legend_gap >= 0
    assert legend_caption_gap >= 0


def test_metopic_figure_xlabel_legend_caption_dont_overlap():
    # 4 axes (main + 3 profile panels), not just the 1 the helper above
    # assumes - checked directly rather than through _text_block_extents.
    # ax_dev is the one that actually collides in practice: its own xlabel
    # sits much closer to the legend than ax_main's, whose aspect="equal"
    # shrinks its plotted area (and pushes its xlabel well clear) in a way
    # ax_dev's plain axes doesn't share.
    u = np.linspace(0.0, 1.0, 100)
    metopic = MetopicResult(
        contour=np.column_stack([np.linspace(-40, 40, 100), np.full(100, 20.0)]),
        arc_length=u * 100.0,
        normalized_arc_length=u,
        midline_u=0.5, parabola_a=-0.01, parabola_c=30.0,
        deviation_profile=np.sin(u * 6.0),
        gradient_profile=np.cos(u * 6.0),
        curvature_profile=np.sin(u * 3.0) * 0.01,
        frontal_angle_deg=125.0,
        frontal_angle_points=(np.array([0.0, 30.0]), np.array([-30.0, 20.0]), np.array([30.0, 20.0])),
        forehead_width_mm=90.0, midline_curvature_concentration=0.25, midline_max_curvature=0.02,
        midline_max_curvature_position=0.5, ridge_protrusion_mm=2.0, ridge_protrusion_position=0.5,
        ridge_area_mm2=50.0, ridge_area_normalized=0.02, left_temporal_hollowing=0.1, right_temporal_hollowing=0.08,
        mean_temporal_hollowing=0.09, left_max_temporal_depth_mm=5.0, right_max_temporal_depth_mm=4.0,
        parabolic_deviation_index=3.0, central_window=(0.4, 0.6), left_temporal_window=(0.1, 0.3),
        right_temporal_window=(0.7, 0.9),
    )

    page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
    _draw_metopic(page, (0.04, 0.46, 0.92, 0.50), metopic)

    canvas = FigureCanvasAgg(page)
    canvas.draw()
    renderer = canvas.get_renderer()
    fig_h = page.bbox.height

    ax_main, ax_phi, ax_kappa, ax_dev = page.axes
    legend_bb = page.legends[0].get_window_extent(renderer)
    legend_top, legend_bottom = legend_bb.y1 / fig_h, legend_bb.y0 / fig_h
    caption = min(page.texts, key=lambda t: t.get_window_extent(renderer).y0)
    caption_top = caption.get_window_extent(renderer).y1 / fig_h

    for ax in (ax_main, ax_dev):
        xlabel_bottom = ax.xaxis.label.get_window_extent(renderer).y0 / fig_h
        assert xlabel_bottom >= legend_top
    assert legend_bottom >= caption_top

    # the 3 stacked side panels (ax_phi above ax_kappa above ax_dev) - each
    # one's title must clear the tick labels of the panel directly above it.
    for upper, lower in ((ax_phi, ax_kappa), (ax_kappa, ax_dev)):
        tick_bottom = min(
            t.get_window_extent(renderer).y0 for t in upper.get_xticklabels() if t.get_text()
        ) / fig_h
        title_top = lower.title.get_window_extent(renderer).y1 / fig_h
        assert tick_bottom >= title_top


# --- asymmetry figure (silhouette + sagittal view) ------------------------


def test_silhouette_polygon_follows_concavities_not_just_convex_hull():
    # a horseshoe ("C") band of points, open on one side and hollow in the
    # middle - strongly concave, so an outline that actually traces the
    # band should enclose noticeably less area than the convex hull would,
    # which fills in both the open mouth and the hollow center.
    theta = np.linspace(np.deg2rad(20), np.deg2rad(340), 150)
    outer = np.column_stack([np.cos(theta), np.sin(theta)]) * 50.0
    inner = np.column_stack([np.cos(theta), np.sin(theta)]) * 25.0
    end_radii = np.linspace(25.0, 50.0, 10)[:, None]
    end_a = np.array([np.cos(theta[0]), np.sin(theta[0])]) * end_radii
    end_b = np.array([np.cos(theta[-1]), np.sin(theta[-1])]) * end_radii
    points = np.vstack([outer, inner, end_a, end_b])

    silhouette = _silhouette_polygon(points, ratio=0.3)
    assert tuple(silhouette[0]) == pytest.approx(tuple(silhouette[-1]))  # closed ring

    concave_area = Polygon(silhouette).area
    convex_area = MultiPoint(points).convex_hull.area
    assert concave_area < convex_area * 0.8


def _asymmetry_mesh_and_heatmap():
    # icosphere stretched into a head-ish blob, with a heatmap zeroed on
    # x < 0 - the same convention craniumpy_core.asymmetry.calculate_asymmetry
    # actually produces, which is what view="sagittal" relies on to know
    # which half of the mesh to keep.
    mesh = trimesh.creation.icosphere(subdivisions=3, radius=1.0)
    mesh.vertices = np.asarray(mesh.vertices) * np.array([70.0, 90.0, 60.0])
    vertices = np.asarray(mesh.vertices)
    radius = np.linalg.norm(vertices, axis=1)
    heatmap = radius - radius.mean()
    heatmap[vertices[:, 0] < 0] = 0.0
    asymmetry = AsymmetryResult(heatmap=heatmap, mean_asymmetry_index=float(np.abs(heatmap).mean()))
    return mesh, asymmetry


def test_draw_asymmetry_top_and_frontal_views_draw_a_silhouette():
    mesh, asymmetry = _asymmetry_mesh_and_heatmap()
    for view in ("top", "frontal"):
        page = Figure(figsize=(6, 6))
        _draw_asymmetry(page, (0, 0, 1, 1), mesh, asymmetry, label="cranial", view=view)
        ax = page.axes[0]
        assert len(ax.lines) == 1  # the silhouette outline, and nothing else


def test_draw_asymmetry_sagittal_view_has_no_silhouette_and_only_the_data_half():
    mesh, asymmetry = _asymmetry_mesh_and_heatmap()
    vertices = np.asarray(mesh.vertices)

    page = Figure(figsize=(6, 6))
    _draw_asymmetry(page, (0, 0, 1, 1), mesh, asymmetry, label="cranial", view="sagittal")
    ax = page.axes[0]
    assert len(ax.lines) == 0  # no silhouette for the sagittal view

    # the triangulation actually plotted should only reference vertices on
    # the x >= 0 (data-carrying) half - collapsing left/right without this
    # would overlay the empty mirror half's faces into the same footprint.
    triangulation = ax.collections[0]._triangulation
    used_vertex_indices = np.unique(triangulation.triangles)
    assert np.all(vertices[used_vertex_indices, 0] >= 0)


def test_pdf_metric_fields_orders_asymmetry_sections_last():
    # both the primary and sagittal group for each target's asymmetry
    # section must be among the last entries, so a session that computes
    # any of them always ends its PDF report there regardless of which
    # other groups (craniometrics, frontal_bossing, metopic) also ran.
    keys = list(_PDF_METRIC_FIELDS.keys())
    asymmetry_keys = {"cranial_asymmetry", "cranial_asymmetry_sagittal", "asymmetry", "asymmetry_sagittal"}
    other_keys = [k for k in keys if k not in asymmetry_keys]
    last_other_index = max(keys.index(k) for k in other_keys)
    first_asymmetry_index = min(keys.index(k) for k in asymmetry_keys)
    assert first_asymmetry_index > last_other_index
    # sagittal always comes right after its own target's primary section
    assert keys.index("cranial_asymmetry_sagittal") == keys.index("cranial_asymmetry") + 1
    assert keys.index("asymmetry_sagittal") == keys.index("asymmetry") + 1
