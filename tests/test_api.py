"""API tests - upload -> analyze -> poll status -> results -> mesh export
-> results bundle.

landmarks are always manual now (see pipeline.py). using REFERENCE_TRIANGLE
as landmarks on template_xy_com.ply since it's already in the frame the
clip constants assume, so harmonize() has something real to work with
without needing an actual scan.
"""

import json
import time
import zipfile
from io import BytesIO
from pathlib import Path

import numpy as np
import pytest
import trimesh
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import app
from api.results_bundle import _id_mapping_path
from craniumpy_core.io import load_mesh
from craniumpy_core.registration.rigid import REFERENCE_TRIANGLE

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATES_DIR = REPO_ROOT / "src" / "craniumpy_core" / "templates"
TEMPLATE_PATH = TEMPLATES_DIR / "template_xy_com.ply"


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


@pytest.fixture
def landmarks_payload() -> list[dict]:
    return [{"x": float(p[0]), "y": float(p[1]), "z": float(p[2])} for p in REFERENCE_TRIANGLE]


def _upload(client: TestClient) -> str:
    with open(TEMPLATE_PATH, "rb") as f:
        response = client.post(
            "/api/sessions", files=[("files", ("template_xy_com.ply", f, "application/octet-stream"))]
        )
    assert response.status_code == 200, response.text
    return response.json()["session_id"]


def _poll_status(client: TestClient, session_id: str, timeout: float) -> str:
    deadline = time.time() + timeout
    status = "running"
    while time.time() < deadline:
        status = client.get(f"/api/sessions/{session_id}/status").json()["status"]
        if status != "running":
            break
        time.sleep(0.2)
    return status


def _run_clip(client: TestClient, session_id: str, body: dict, timeout: float = 30) -> str:
    response = client.post(f"/api/sessions/{session_id}/clip", json=body)
    assert response.status_code == 200, response.text
    return _poll_status(client, session_id, timeout)


def _run_run(client: TestClient, session_id: str, body: dict | None = None, timeout: float = 30) -> str:
    response = client.post(f"/api/sessions/{session_id}/run", json=body or {})
    assert response.status_code == 200, response.text
    return _poll_status(client, session_id, timeout)


def _clip_and_run(
    client: TestClient, session_id: str, clip_body: dict, run_body: dict | None = None, timeout: float = 30
) -> str:
    """most tests just want the whole staged flow to finish - clip then
    run, same net effect the old one-call /analyze used to have."""
    status = _run_clip(client, session_id, clip_body, timeout)
    if status != "done":
        return status
    return _run_run(client, session_id, run_body, timeout)


def test_upload_returns_session_and_vertex_count(client):
    session_id = _upload(client)
    assert session_id

    response = client.get(f"/api/sessions/{session_id}/mesh/original")
    assert response.status_code == 200
    assert response.headers["content-type"] == "model/gltf-binary"
    assert len(response.content) > 0


def test_upload_rejects_unsupported_extension(client):
    response = client.post(
        "/api/sessions", files=[("files", ("mesh.xyz", b"not a mesh", "application/octet-stream"))]
    )
    assert response.status_code == 400


def test_unknown_session_returns_404(client):
    response = client.get("/api/sessions/does-not-exist/status")
    assert response.status_code == 404


def test_list_templates(client):
    response = client.get("/api/templates")
    assert response.status_code == 200
    names = [t["name"] for t in response.json()]
    assert "clipped_template_xy_com" in names
    assert len(response.json()) >= 1


def test_get_template_mesh(client):
    response = client.get("/api/templates/clipped_template_xy_com/mesh")
    assert response.status_code == 200
    assert response.headers["content-type"] == "model/gltf-binary"
    assert len(response.content) > 0


def test_get_unknown_template_mesh_404s(client):
    response = client.get("/api/templates/not-a-real-template/mesh")
    assert response.status_code == 404


def test_get_template_mesh_served_unmodified(client):
    # templates are served exactly as stored on disk - no live clipping or
    # any other processing, so vertex count should match the raw file.
    response = client.get("/api/templates/clipped_template_xy_com/mesh")
    on_disk = load_mesh(TEMPLATES_DIR / "clipped_template_xy_com.ply")
    served = trimesh.load(BytesIO(response.content), file_type="glb", process=False, force="mesh")
    assert len(served.vertices) == len(on_disk.vertices)


def test_get_custom_template_mesh_from_path(client):
    response = client.get(f"/api/templates/custom/mesh?path={TEMPLATE_PATH}")
    assert response.status_code == 200
    assert response.headers["content-type"] == "model/gltf-binary"
    assert len(response.content) > 0


def test_get_custom_template_mesh_missing_path_400s(client):
    response = client.get("/api/templates/custom/mesh?path=C:/nope/not-a-real-file.ply")
    assert response.status_code == 400


def test_upload_custom_template_mesh(client):
    with open(TEMPLATE_PATH, "rb") as f:
        response = client.post(
            "/api/templates/custom/upload",
            files=[("files", ("custom_template.ply", f, "application/octet-stream"))],
        )
    assert response.status_code == 200
    assert response.headers["content-type"] == "model/gltf-binary"
    assert len(response.content) > 0


def test_full_analysis_flow_rigid(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(
        client,
        session_id,
        {"target": "cranium", "landmarks": landmarks_payload, "com_translation": True},
    )
    assert status == "done"

    results = client.get(f"/api/sessions/{session_id}/results")
    assert results.status_code == 200, results.text
    body = results.json()
    assert len(body["landmarks"]) == 3
    assert body["craniometrics"] is not None
    assert body["craniometrics"]["depth_mm"] > 0
    assert len(body["craniometrics"]["hc_slice_polygon"]) > 3
    assert body["vertex_count"] > 0

    mesh_response = client.get(f"/api/sessions/{session_id}/mesh/result")
    assert mesh_response.status_code == 200
    assert len(mesh_response.content) > 0

    registered_response = client.get(f"/api/sessions/{session_id}/mesh/registered")
    assert registered_response.status_code == 200
    assert len(registered_response.content) > 0

    clipped_response = client.get(f"/api/sessions/{session_id}/mesh/clipped")
    assert clipped_response.status_code == 200
    assert len(clipped_response.content) > 0


def test_analyze_cranial_with_alt_frontal_landmark(client, landmarks_payload):
    # stands in for "subnasale instead of sellion" - a different point,
    # clearly outside the real landmark triangle
    alt_frontal = {
        "x": landmarks_payload[0]["x"],
        "y": landmarks_payload[0]["y"] - 15.0,
        "z": landmarks_payload[0]["z"] - 5.0,
    }
    run_body = {"n_vertices": 3000}

    session_plain = _upload(client)
    assert _clip_and_run(
        client, session_plain,
        {"target": "cranium", "landmarks": landmarks_payload},
        run_body,
    ) == "done"
    plain_results = client.get(f"/api/sessions/{session_plain}/results").json()

    session_alt = _upload(client)
    assert _clip_and_run(
        client, session_alt,
        {"target": "cranium", "landmarks": landmarks_payload, "alt_frontal_landmark": alt_frontal},
        run_body,
    ) == "done"
    alt_results = client.get(f"/api/sessions/{session_alt}/results").json()

    assert plain_results["used_alt_frontal"] is False
    assert alt_results["used_alt_frontal"] is True

    # the numbers themselves never change based on which frame gets shown -
    # they always come from the mandatory sellion landmark
    assert alt_results["craniometrics"]["depth_mm"] == pytest.approx(plain_results["craniometrics"]["depth_mm"])
    assert alt_results["craniometrics"]["breadth_mm"] == pytest.approx(plain_results["craniometrics"]["breadth_mm"])
    assert alt_results["craniometrics"]["circumference_cm"] == pytest.approx(
        plain_results["craniometrics"]["circumference_cm"]
    )

    # but the displayed/downloadable mesh actually is a different pose
    plain_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_plain}/mesh/result").content), file_type="glb", force="mesh"
    )
    alt_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_alt}/mesh/result").content), file_type="glb", force="mesh"
    )
    assert not np.allclose(
        np.asarray(plain_mesh.vertices).mean(axis=0), np.asarray(alt_mesh.vertices).mean(axis=0), atol=1.0
    )

    # the saved report documents both frames when they differ
    bundle = client.get(f"/api/sessions/{session_alt}/bundle")
    assert bundle.status_code == 200
    zf = zipfile.ZipFile(BytesIO(bundle.content))
    report_name = next(n for n in zf.namelist() if n.endswith("_report_cranial.json"))
    report = json.loads(zf.read(report_name))
    assert "display_frontal_landmark" in report
    # com_translation nudges Z (a few mm, on this fixture) - not an exact
    # match to REFERENCE_TRIANGLE, but nowhere near alt_frontal's z either
    # (5mm further off) - just confirms the report's "sellion" entry is the
    # sellion-frame landmark, not the alt one.
    np.testing.assert_allclose(report["landmarks"]["sellion"], list(REFERENCE_TRIANGLE[0]), atol=4.0)


def test_bundle_analysis_nests_report_under_mesh_folder(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}, {"n_vertices": 3000})
    assert status == "done"

    response = client.get(f"/api/sessions/{session_id}/bundle/analysis")
    assert response.status_code == 200
    names = set(zipfile.ZipFile(BytesIO(response.content)).namelist())
    mesh_names = {n for n in names if "/analysis/" not in n}
    analysis_names = {n for n in names if "/analysis/" in n}
    assert len(mesh_names) == 2
    assert any(n.endswith("_report_cranial.json") for n in analysis_names)
    assert any(n.endswith("_measurements.png") for n in analysis_names)


def test_clip_progress_is_reported(client, landmarks_payload):
    session_id = _upload(client)
    client.post(
        f"/api/sessions/{session_id}/clip",
        json={"target": "cranium", "landmarks": landmarks_payload},
    )
    # at least one poll should show a real stage, not just idle - proves
    # progress is actually wired through and not just a static "running" flag
    seen_stages = set()
    deadline = time.time() + 30
    while time.time() < deadline:
        status = client.get(f"/api/sessions/{session_id}/status").json()
        if status["progress"]:
            seen_stages.add(status["progress"]["stage"])
        if status["status"] != "running":
            break
        time.sleep(0.05)
    assert seen_stages - {"idle"}


def test_manual_clip_mode(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(
        client,
        session_id,
        {
            "target": "cranium",
            "landmarks": landmarks_payload,
            "clipping": {"mode": "manual", "manual_plane_normal": [0, 1, 0], "manual_plane_origin": [0, 0, 0]},
            "repair": False,
        },
        {"n_vertices": 3000},
    )
    assert status == "done"
    results = client.get(f"/api/sessions/{session_id}/results").json()
    assert results["vertex_count"] > 0


def test_run_with_nicp_cranial_produces_template_topology(client, landmarks_payload):
    # "fit template" only ever runs after a completed plain run (matches
    # the frontend's own gating - see App.jsx's pipelineRan) - the plain
    # run's own result_mesh/craniometrics must survive a fit untouched.
    from craniumpy_core.template_registry import load_shipped_template

    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}, timeout=60)
    assert status == "done"

    plain_result_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_id}/mesh/result").content), file_type="glb", process=False, force="mesh"
    )
    plain_craniometrics = client.get(f"/api/sessions/{session_id}/results").json()["craniometrics"]
    assert plain_craniometrics is not None

    status = _run_run(
        client,
        session_id,
        # small alpha schedule / relaxed threshold - fast is all this test
        # needs, real tuning is exercised in test_nicp.py.
        {"nicp": {"template": "clipped_template_xy", "alpha_start": 50, "alpha_end": 1, "alpha_steps": 3, "inner_iters": 1, "dist_threshold": 50.0}},
        timeout=60,
    )
    assert status == "done"

    template = load_shipped_template("clipped_template_xy")
    nicp_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_id}/mesh/nicp-result").content), file_type="glb", process=False, force="mesh"
    )
    assert len(nicp_mesh.vertices) == len(template.vertices)

    # result_mesh/craniometrics are exactly what the plain run left them as
    # - a fit is a pure addition, not a replacement.
    result_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_id}/mesh/result").content), file_type="glb", process=False, force="mesh"
    )
    assert len(result_mesh.vertices) == len(plain_result_mesh.vertices)
    assert client.get(f"/api/sessions/{session_id}/results").json()["craniometrics"] == plain_craniometrics


def test_run_with_nicp_facial_produces_template_topology(client, landmarks_payload):
    from craniumpy_core.template_registry import load_shipped_template

    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "face", "landmarks": landmarks_payload}, timeout=60)
    assert status == "done"

    plain_result_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_id}/mesh/result").content), file_type="glb", process=False, force="mesh"
    )
    plain_asymmetry = client.get(f"/api/sessions/{session_id}/results").json()["asymmetry"]
    assert plain_asymmetry is not None

    status = _run_run(
        client,
        session_id,
        {"nicp": {"template": "template_face", "alpha_start": 50, "alpha_end": 1, "alpha_steps": 3, "inner_iters": 1, "dist_threshold": 50.0}},
        timeout=60,
    )
    assert status == "done"

    template = load_shipped_template("template_face")
    nicp_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_id}/mesh/nicp-result").content), file_type="glb", process=False, force="mesh"
    )
    assert len(nicp_mesh.vertices) == len(template.vertices)

    result_mesh = trimesh.load(
        BytesIO(client.get(f"/api/sessions/{session_id}/mesh/result").content), file_type="glb", process=False, force="mesh"
    )
    assert len(result_mesh.vertices) == len(plain_result_mesh.vertices)
    assert client.get(f"/api/sessions/{session_id}/results").json()["asymmetry"] == plain_asymmetry


def test_facial_run_includes_metopic_analysis(client, landmarks_payload):
    # metopic/frontal-angle analysis rides along with every facial-target
    # run automatically (no separate opt-in) - see craniumpy_core.metopic
    # and pipeline.hc_slice_height_facial_frame.
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "face", "landmarks": landmarks_payload}, timeout=60)
    assert status == "done"

    results = client.get(f"/api/sessions/{session_id}/results").json()
    assert results["asymmetry"] is not None
    metopic = results["metopic"]
    assert metopic is not None

    for key in (
        "frontal_angle_deg",
        "midline_curvature_concentration",
        "midline_max_curvature",
        "midline_max_curvature_position",
        "ridge_protrusion_mm",
        "ridge_area_mm2",
        "ridge_area_normalized",
        "left_temporal_hollowing",
        "right_temporal_hollowing",
        "mean_temporal_hollowing",
        "left_max_temporal_depth_mm",
        "right_max_temporal_depth_mm",
        "parabolic_deviation_index",
    ):
        assert key in metopic
        assert np.isfinite(metopic[key])

    assert 0.0 < metopic["frontal_angle_deg"] < 180.0
    assert len(metopic["contour"]) > 5
    assert len(metopic["gradient_profile"]) == len(metopic["curvature_profile"]) == len(metopic["deviation_profile"]) == len(metopic["contour"])
    # cranial-target sessions stay completely unaffected by this feature
    cranial_session = _upload(client)
    status = _clip_and_run(client, cranial_session, {"target": "cranium", "landmarks": landmarks_payload}, timeout=60)
    assert status == "done"
    assert client.get(f"/api/sessions/{cranial_session}/results").json()["metopic"] is None


def test_bundle_analysis_includes_metopic_figure_for_facial_target(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "face", "landmarks": landmarks_payload}, timeout=60)
    assert status == "done"

    response = client.get(f"/api/sessions/{session_id}/bundle/analysis")
    assert response.status_code == 200
    names = set(zipfile.ZipFile(BytesIO(response.content)).namelist())
    analysis_names = {n for n in names if "/analysis/" in n}
    assert any(n.endswith("_asymmetry.png") for n in analysis_names)
    assert any(n.endswith("_metopic.png") for n in analysis_names)

    report_name = next(n for n in analysis_names if n.endswith("_report_frontal.json"))
    report = json.loads(zipfile.ZipFile(BytesIO(response.content)).read(report_name))
    assert "metopic" in report
    assert "frontal_angle_deg" in report["metopic"]


def test_run_includes_frontal_bossing_for_both_targets(client, landmarks_payload):
    # unlike metopic (facial-only), frontal bossing is computed for cranial
    # and facial targets alike - see craniumpy_core.craniometrics.frontal_bossing
    for target in ("cranium", "face"):
        session_id = _upload(client)
        status = _clip_and_run(client, session_id, {"target": target, "landmarks": landmarks_payload}, timeout=60)
        assert status == "done"

        results = client.get(f"/api/sessions/{session_id}/results").json()
        bossing = results["frontal_bossing"]
        assert bossing is not None, f"target={target}"
        assert np.isfinite(bossing["angle_deg"])
        assert 0.0 <= bossing["angle_deg"] <= 180.0
        for point_key in ("sellion", "frontal_point"):
            for axis in ("x", "y", "z"):
                assert np.isfinite(bossing[point_key][axis])
        assert len(bossing["profile"]) > 0


def test_bundle_analysis_includes_frontal_bossing_figure(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}, timeout=60)
    assert status == "done"

    response = client.get(f"/api/sessions/{session_id}/bundle/analysis")
    assert response.status_code == 200
    names = set(zipfile.ZipFile(BytesIO(response.content)).namelist())
    analysis_names = {n for n in names if "/analysis/" in n}
    assert any(n.endswith("_frontal_bossing.png") for n in analysis_names)

    report_name = next(n for n in analysis_names if n.endswith("_report_cranial.json"))
    report = json.loads(zipfile.ZipFile(BytesIO(response.content)).read(report_name))
    assert "frontal_bossing" in report
    assert "angle_deg" in report["frontal_bossing"]


def test_bundle_analysis_includes_asymmetry_sagittal_figure_for_both_targets(client, landmarks_payload):
    # the sagittal (side-view) asymmetry figure rides along as its own PNG,
    # in addition to the existing top/frontal-view one - see
    # api/results_bundle.py's _build_analysis_files.
    for target, primary_suffix in (("cranium", "cranial_asymmetry"), ("face", "asymmetry")):
        session_id = _upload(client)
        status = _clip_and_run(client, session_id, {"target": target, "landmarks": landmarks_payload}, timeout=60)
        assert status == "done"

        response = client.get(f"/api/sessions/{session_id}/bundle/analysis")
        assert response.status_code == 200
        names = set(zipfile.ZipFile(BytesIO(response.content)).namelist())
        analysis_names = {n for n in names if "/analysis/" in n}
        assert any(n.endswith(f"_{primary_suffix}.png") for n in analysis_names), f"target={target}"
        assert any(n.endswith(f"_{primary_suffix}_sagittal.png") for n in analysis_names), f"target={target}"


def test_bundle_meshes_before_run_returns_409(client, landmarks_payload):
    session_id = _upload(client)
    status = _run_clip(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"
    # clipped, but never run - save-meshes needs result_mesh, which only /run sets
    response = client.get(f"/api/sessions/{session_id}/bundle/meshes")
    assert response.status_code == 409


def test_bundle_meshes_contains_only_mesh_files(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}, {"n_vertices": 3000})
    assert status == "done"

    response = client.get(f"/api/sessions/{session_id}/bundle/meshes")
    assert response.status_code == 200
    names = zipfile.ZipFile(BytesIO(response.content)).namelist()
    assert len(names) == 2
    assert all(n.endswith("_rg.ply") or n.endswith("_rg_C.ply") for n in names)


def test_run_before_clip_returns_409(client):
    session_id = _upload(client)
    response = client.post(f"/api/sessions/{session_id}/run", json={"n_vertices": 3000})
    assert response.status_code == 409


def test_undo_with_nothing_to_undo_returns_reverted_false(client):
    session_id = _upload(client)
    response = client.post(f"/api/sessions/{session_id}/clip/undo")
    assert response.status_code == 200
    assert response.json() == {"reverted": False}


def test_clip_then_undo_reverts_to_registered_mesh(client, landmarks_payload):
    session_id = _upload(client)
    status = _run_clip(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"
    assert client.get(f"/api/sessions/{session_id}/mesh/clipped").status_code == 200

    undo_response = client.post(f"/api/sessions/{session_id}/clip/undo")
    assert undo_response.status_code == 200
    assert undo_response.json() == {"reverted": True}

    # the clip is gone, but the registration it started from is still there
    assert client.get(f"/api/sessions/{session_id}/mesh/clipped").status_code == 409
    assert client.get(f"/api/sessions/{session_id}/mesh/registered").status_code == 200

    # and run can't proceed until a fresh clip commits again
    assert client.post(f"/api/sessions/{session_id}/run", json={"n_vertices": 3000}).status_code == 409


def test_reclip_invalidates_previous_run_result(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"
    assert client.get(f"/api/sessions/{session_id}/results").status_code == 200

    # a fresh /clip - even with the same settings - invalidates whatever
    # /run already produced, since it might no longer match
    status = _run_clip(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"
    assert client.get(f"/api/sessions/{session_id}/results").status_code == 409


def test_switch_target_never_visited_reports_not_restored_and_stays_blank(client, landmarks_payload):
    session_id = _upload(client)
    assert _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}) == "done"

    response = client.post(f"/api/sessions/{session_id}/switch-target", json={"target": "face"})
    assert response.status_code == 200
    assert response.json() == {
        "restored": False,
        "align_succeeded": False,
        "pipeline_ran": False,
        "has_nicp_result": False,
        "used_alt_frontal": False,
    }
    # a genuinely blank facial scene, not the cranial one left showing
    assert client.get(f"/api/sessions/{session_id}/mesh/registered").status_code == 409
    assert client.get(f"/api/sessions/{session_id}/results").status_code == 409


def test_switch_target_restores_a_previously_processed_target_without_recomputation(client, landmarks_payload, monkeypatch):
    # proves the "snapshot, don't recompute" contract end to end: switching
    # back to cranial after visiting facial must reuse exactly what the
    # first cranial run produced, without ever calling back into the
    # register/clip pipeline for it a second time.
    import craniumpy_core.pipeline as pipeline_module

    real_register_and_clip = pipeline_module.register_and_clip_cranial
    call_count = {"n": 0}

    def counting_register_and_clip(*args, **kwargs):
        call_count["n"] += 1
        return real_register_and_clip(*args, **kwargs)

    monkeypatch.setattr(pipeline_module, "register_and_clip_cranial", counting_register_and_clip)

    session_id = _upload(client)
    assert _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}) == "done"
    assert call_count["n"] == 1
    cranial_results = client.get(f"/api/sessions/{session_id}/results").json()

    switch_away = client.post(f"/api/sessions/{session_id}/switch-target", json={"target": "face"})
    assert switch_away.json()["restored"] is False

    assert _clip_and_run(client, session_id, {"target": "face", "landmarks": landmarks_payload}) == "done"
    assert client.get(f"/api/sessions/{session_id}/results").json()["craniometrics"] is None  # facial now active

    switch_back = client.post(f"/api/sessions/{session_id}/switch-target", json={"target": "cranium"})
    body = switch_back.json()
    assert body["restored"] is True
    assert body["pipeline_ran"] is True
    assert body["align_succeeded"] is True

    # register_and_clip_cranial was never called again to get here
    assert call_count["n"] == 1
    restored_results = client.get(f"/api/sessions/{session_id}/results").json()
    assert restored_results["craniometrics"] == cranial_results["craniometrics"]

    # the restored state is genuinely usable, not just present - exporting
    # the (correctly restored) cranial data works immediately, no re-run
    assert client.get(f"/api/sessions/{session_id}/bundle/analysis").status_code == 200


def test_switch_target_while_job_running_returns_409(client, landmarks_payload):
    from api.sessions import store

    session_id = _upload(client)
    store.get(session_id).job_status = "running"
    response = client.post(f"/api/sessions/{session_id}/switch-target", json={"target": "face"})
    assert response.status_code == 409


def test_repair_only_runs_once_across_repeated_clips(client, landmarks_payload, monkeypatch):
    import craniumpy_core.pipeline as pipeline_module

    real_repair = pipeline_module.repair_mesh
    call_count = {"n": 0}

    def counting_repair(*args, **kwargs):
        call_count["n"] += 1
        return real_repair(*args, **kwargs)

    monkeypatch.setattr(pipeline_module, "repair_mesh", counting_repair)

    session_id = _upload(client)
    assert _run_clip(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}) == "done"
    assert call_count["n"] == 1

    # repeating the exact same clip on the same session should reuse the
    # cached repair, not re-run pymeshfix
    assert _run_clip(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}) == "done"
    assert call_count["n"] == 1

    # manual mode is excluded from the landmark-based rough pre-clip (its
    # plane is arbitrary, not derived from the landmarks - see
    # pipeline.rough_bounding_clip), so what actually gets repaired differs
    # from the cranial pass above and the cache correctly misses, re-running
    # repair on the full mesh instead of reusing the cranial pass's
    # rough-clipped repair.
    assert (
        _run_clip(
            client,
            session_id,
            {
                "target": "cranium",
                "landmarks": landmarks_payload,
                "clipping": {"mode": "manual", "manual_plane_normal": [0, 1, 0], "manual_plane_origin": [0, 0, 0]},
            },
        )
        == "done"
    )
    assert call_count["n"] == 2


def test_results_before_analysis_returns_409(client):
    session_id = _upload(client)
    response = client.get(f"/api/sessions/{session_id}/results")
    assert response.status_code == 409


def test_measure_registered_skips_straight_to_results(client):
    # "preprocessing already performed" - no /align, /clip, or /run at all,
    # just the uploaded mesh treated as already registered.
    session_id = _upload(client)
    response = client.post(f"/api/sessions/{session_id}/measure-registered", json={"target": "cranium"})
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["craniometrics"] is not None
    assert body["asymmetry"] is not None
    assert body["metopic"] is None  # cranium target never computes metopic, same as a real /run

    # get_results now works too, without a separate poll - job_status was
    # set straight to "done".
    results_response = client.get(f"/api/sessions/{session_id}/results")
    assert results_response.status_code == 200
    assert results_response.json()["craniometrics"] is not None


def test_measure_registered_face_target_computes_metopic(client):
    session_id = _upload(client)
    response = client.post(f"/api/sessions/{session_id}/measure-registered", json={"target": "face"})
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["craniometrics"] is None  # face target never computes craniometrics
    assert body["asymmetry"] is not None
    assert body["metopic"] is not None


def test_measure_registered_then_save_meshes_and_analysis(client, tmp_path):
    import shutil

    tmp_mesh = tmp_path / "patient.ply"
    shutil.copy(TEMPLATE_PATH, tmp_mesh)
    open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
    session_id = open_response.json()["session_id"]

    response = client.post(f"/api/sessions/{session_id}/measure-registered", json={"target": "cranium"})
    assert response.status_code == 200, response.text

    save_meshes_response = client.post(f"/api/sessions/{session_id}/save/meshes")
    assert save_meshes_response.status_code == 200, save_meshes_response.text
    meshes_dir = Path(save_meshes_response.json()["saved_to"])
    # plain CP_C_{stem}/ - no landmark-count/CoM suffix, since there's no
    # real clip config behind a skipped-preprocessing session.
    assert meshes_dir == tmp_path / "CP_C_patient" / "meshes"
    assert any(n.endswith("_rg.ply") for n in (p.name for p in meshes_dir.iterdir()))

    save_analysis_response = client.post(f"/api/sessions/{session_id}/save/analysis")
    assert save_analysis_response.status_code == 200, save_analysis_response.text
    analysis_dir = Path(save_analysis_response.json()["saved_to"])
    assert analysis_dir == tmp_path / "CP_C_patient" / "analysis"
    assert any(n.endswith("_report_cranial.json") for n in (p.name for p in analysis_dir.iterdir()))


def test_clip_rejects_wrong_landmark_count(client):
    session_id = _upload(client)
    response = client.post(
        f"/api/sessions/{session_id}/clip",
        json={"target": "cranium", "landmarks": [{"x": 0, "y": 0, "z": 0}]},
    )
    assert response.status_code == 422


def test_results_bundle_download(client, landmarks_payload):
    session_id = _upload(client)
    status = _clip_and_run(
        client,
        session_id,
        {"target": "cranium", "landmarks": landmarks_payload},
        {"n_vertices": 3000},
    )
    assert status == "done"

    bundle = client.get(f"/api/sessions/{session_id}/bundle")
    assert bundle.status_code == 200
    assert bundle.headers["content-type"] == "application/zip"
    assert "CP_C_template_xy_com_CoM.zip" in bundle.headers["content-disposition"]

    zf = zipfile.ZipFile(BytesIO(bundle.content))
    names = zf.namelist()
    assert any(n.endswith("_rg.ply") for n in names)
    assert any(n.endswith("_rg_C.ply") for n in names)
    assert any(n.endswith("_report_cranial.json") for n in names)
    assert any(n.endswith("_measurements.png") for n in names)


def test_open_mesh_from_paths(client):
    response = client.post("/api/sessions/from-paths", json={"paths": [str(TEMPLATE_PATH)]})
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["vertex_count"] > 0


def test_open_mesh_from_paths_missing_file_400s(client):
    response = client.post("/api/sessions/from-paths", json={"paths": ["C:/nope/not-a-real-file.ply"]})
    assert response.status_code == 400


def test_save_results_without_source_dir_400s(client, landmarks_payload):
    # opened via the plain-bytes /api/sessions upload, so there's no known
    # source folder to save into
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"

    response = client.post(f"/api/sessions/{session_id}/save")
    assert response.status_code == 400


def test_save_results_to_source_folder(client, landmarks_payload, tmp_path):
    import shutil

    tmp_mesh = tmp_path / "1016510_20210730.000112_edited.ply"
    shutil.copy(TEMPLATE_PATH, tmp_mesh)

    open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
    session_id = open_response.json()["session_id"]

    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"

    save_response = client.post(f"/api/sessions/{session_id}/save")
    assert save_response.status_code == 200, save_response.text
    saved_to = Path(save_response.json()["saved_to"])
    assert saved_to == tmp_path / "CP_C_1016510_20210730_edited_CoM"
    assert (saved_to / "1016510_20210730_edited_rg.ply").exists()
    assert (saved_to / "1016510_20210730_edited_rg_C.ply").exists()
    assert (saved_to / "1016510_20210730_edited_report_cranial.json").exists()


def test_save_results_dest_dir_override(client, landmarks_payload, tmp_path):
    # dest_dir, when given, wins over the session's own source_dir - the
    # desktop app's "change save folder..." control (see api/schemas.py's
    # SaveRequest).
    import shutil

    source_dir = tmp_path / "source"
    source_dir.mkdir()
    override_dir = tmp_path / "chosen"
    override_dir.mkdir()
    tmp_mesh = source_dir / "1016510_20210730.000112_edited.ply"
    shutil.copy(TEMPLATE_PATH, tmp_mesh)

    open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
    session_id = open_response.json()["session_id"]

    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"

    save_response = client.post(f"/api/sessions/{session_id}/save", json={"dest_dir": str(override_dir)})
    assert save_response.status_code == 200, save_response.text
    saved_to = Path(save_response.json()["saved_to"])
    assert saved_to == override_dir / "CP_C_1016510_20210730_edited_CoM"
    assert not (source_dir / "CP_C_1016510_20210730_edited_CoM").exists()


def test_save_results_dest_dir_not_a_real_folder_400s(client, landmarks_payload, tmp_path):
    import shutil

    tmp_mesh = tmp_path / "1016510_20210730.000112_edited.ply"
    shutil.copy(TEMPLATE_PATH, tmp_mesh)

    open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
    session_id = open_response.json()["session_id"]

    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"

    save_response = client.post(f"/api/sessions/{session_id}/save", json={"dest_dir": "C:/nope/not-a-real-folder"})
    assert save_response.status_code == 400


def test_save_meshes_gains_a_third_file_after_nicp_fit(client, landmarks_payload, tmp_path):
    # a plain run saves the usual two mesh files - no third one, even
    # though a template is available, since none was ever fit (see
    # api/sessions.py's Session.nicp_result_mesh).
    import shutil

    tmp_mesh = tmp_path / "patient.ply"
    shutil.copy(TEMPLATE_PATH, tmp_mesh)
    open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
    session_id = open_response.json()["session_id"]

    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"

    save_response = client.post(f"/api/sessions/{session_id}/save/meshes")
    assert save_response.status_code == 200, save_response.text
    saved_to = Path(save_response.json()["saved_to"])
    names = {p.name for p in saved_to.iterdir()}
    assert len(names) == 2
    assert any(n.endswith("_rg.ply") for n in names)
    assert any(n.endswith("_rg_C.ply") for n in names)
    assert not any(n.endswith("_rg_CN.ply") for n in names)

    # fitting a template (no /clip needed - same clipped mesh) adds a third
    # file on the next save, without disturbing the first two.
    status = _run_run(
        client,
        session_id,
        {
            "nicp": {
                "template": "clipped_template_xy",
                "alpha_start": 50,
                "alpha_end": 1,
                "alpha_steps": 3,
                "inner_iters": 1,
                "dist_threshold": 50.0,
            }
        },
    )
    assert status == "done"

    save_response = client.post(f"/api/sessions/{session_id}/save/meshes")
    assert save_response.status_code == 200, save_response.text
    saved_to = Path(save_response.json()["saved_to"])
    names = {p.name for p in saved_to.iterdir()}
    assert len(names) == 3
    assert any(n.endswith("_rg.ply") for n in names)
    assert any(n.endswith("_rg_C.ply") for n in names)
    assert any(n.endswith("_rg_CN.ply") for n in names)


def test_analysis_export_records_nicp_settings_after_a_fit(client, landmarks_payload):
    # the JSON report's settings block and the summary xlsx row both have
    # to reflect that a template was actually fit, and which one, even
    # though "fit template" never touches craniometrics/asymmetry itself
    # (see api/sessions.py's Session.nicp_result_mesh) - see
    # api/routers/mesh.py's _config_with_nicp for the wiring under test.
    session_id = _upload(client)
    status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
    assert status == "done"

    status = _run_run(
        client,
        session_id,
        {
            "nicp": {
                "template": "clipped_template_xy",
                "alpha_start": 50,
                "alpha_end": 1,
                "alpha_steps": 3,
                "inner_iters": 1,
                "dist_threshold": 50.0,
            }
        },
    )
    assert status == "done"

    response = client.get(f"/api/sessions/{session_id}/bundle/analysis")
    assert response.status_code == 200, response.text
    zf = zipfile.ZipFile(BytesIO(response.content))
    analysis_names = [n for n in zf.namelist() if "/analysis/" in n]

    report_name = next(n for n in analysis_names if n.endswith("_report_cranial.json"))
    report = json.loads(zf.read(report_name))
    assert report["settings"]["nicp"]["template"] == "clipped_template_xy"

    xlsx_name = next(n for n in analysis_names if n.endswith("_summary_cranial.xlsx"))
    row = _read_xlsx_rows(zf.read(xlsx_name))[0]
    assert row["nicp_used"] == "yes"
    assert row["nicp_template"] == "clipped_template_xy"


def _read_xlsx_rows(xlsx_bytes: bytes) -> list[dict]:
    # a numeric cell (see results_bundle._write_xlsx_rows) comes back from
    # openpyxl as a bare int/float, formatted here to the same 2 decimals
    # the sheet itself displays (int for a whole number, dropped trailing
    # zeros for a fraction - neither matches what's actually shown/what a
    # _fmt(...)-style "X.00" string comparison expects otherwise).
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    header = [str(c.value) if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(
            {
                header[i]: ("" if v is None else f"{v:.2f}" if isinstance(v, (int, float)) else str(v))
                for i, v in enumerate(excel_row)
            }
        )
    return rows


def test_bundle_analysis_includes_summary_xlsx_and_pdf_report(client, landmarks_payload):
    # both targets - the spreadsheet/PDF ride along on every export
    # regardless of which metric groups actually computed (see
    # _build_analysis_files).
    for target in ("cranium", "face"):
        session_id = _upload(client)
        status = _clip_and_run(client, session_id, {"target": target, "landmarks": landmarks_payload}, timeout=60)
        assert status == "done"

        response = client.get(
            f"/api/sessions/{session_id}/bundle/analysis",
            params={"sex": "female", "date_imaging": "2024-01-15", "age_imaging": "6"},
        )
        assert response.status_code == 200, response.text
        zf = zipfile.ZipFile(BytesIO(response.content))
        analysis_names = [n for n in zf.namelist() if "/analysis/" in n]

        target_suffix = "cranial" if target == "cranium" else "frontal"
        xlsx_name = next(n for n in analysis_names if n.endswith(f"_summary_{target_suffix}.xlsx"))
        pdf_name = next(n for n in analysis_names if n.endswith(f"_report_{target_suffix}.pdf"))

        rows = _read_xlsx_rows(zf.read(xlsx_name))
        assert len(rows) == 1
        row = rows[0]
        assert row["sex"] == "female"
        assert row["date_imaging"] == "2024-01-15"
        assert row["age_imaging"] == "6"
        assert row["target"] == target
        # settings columns - com_translation defaults on, no nicp was
        # requested for this run (see _clip_and_run's plain clip/run body)
        assert row["com_correction"] == "yes"
        assert row["nicp_used"] == "no"
        assert row["nicp_template"] == ""
        # a metadata field that was never sent still appears, just blank
        assert row["treatment"] == ""

        results = client.get(f"/api/sessions/{session_id}/results").json()
        if target == "cranium":
            assert float(row["depth_mm"]) == pytest.approx(results["craniometrics"]["depth_mm"], abs=0.05)
            assert row["mean_asymmetry_index"] == ""
            # cranial asymmetry - same mirror-and-ICP method as facial's,
            # now computed for cranial sessions too (see
            # pipeline.measure_cranial) - its own column, not the shared
            # "mean_asymmetry_index" one facial populates below.
            assert float(row["cranial_asymmetry_index"]) == pytest.approx(
                results["asymmetry"]["mean_asymmetry_index"], abs=0.05
            )
        else:
            assert float(row["mean_asymmetry_index"]) == pytest.approx(
                results["asymmetry"]["mean_asymmetry_index"], abs=0.05
            )
            assert row["depth_mm"] == ""
            assert row["cranial_asymmetry_index"] == ""

        pdf_bytes = zf.read(pdf_name)
        assert pdf_bytes.startswith(b"%PDF")
        assert len(pdf_bytes) > 1000


def test_bundle_analysis_export_selection_excludes_unchecked_sections(client, landmarks_payload):
    # the "measurements"/"asymmetry"/"meshes" export checkboxes (see
    # AnalysisPanel.jsx) - unticking one drops that section from the JSON
    # report/PDF/xlsx entirely (not just hides it), or skips the mesh files.
    session_id = _upload(client)
    assert _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}, timeout=60) == "done"

    response = client.get(
        f"/api/sessions/{session_id}/bundle/analysis",
        params={"include_measurements": "false"},
    )
    assert response.status_code == 200
    zf = zipfile.ZipFile(BytesIO(response.content))
    names = zf.namelist()
    analysis_names = [n for n in names if "/analysis/" in n]
    assert not any(n.endswith("_measurements.png") for n in analysis_names)

    report_name = next(n for n in analysis_names if n.endswith("_report_cranial.json"))
    report = json.loads(zf.read(report_name))
    assert "craniometrics" not in report

    xlsx_name = next(n for n in analysis_names if n.endswith("_summary_cranial.xlsx"))
    rows = _read_xlsx_rows(zf.read(xlsx_name))
    assert rows[0]["depth_mm"] == ""
    # asymmetry (still ticked, default) is unaffected
    assert rows[0]["cranial_asymmetry_index"] != ""

    # meshes ticked (default) - still present
    assert any(n.endswith("_rg.ply") for n in names)


def test_bundle_analysis_export_selection_excludes_asymmetry_and_meshes(client, landmarks_payload):
    session_id = _upload(client)
    assert _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}, timeout=60) == "done"

    response = client.get(
        f"/api/sessions/{session_id}/bundle/analysis",
        params={"include_asymmetry": "false", "include_meshes": "false"},
    )
    assert response.status_code == 200
    zf = zipfile.ZipFile(BytesIO(response.content))
    names = zf.namelist()
    analysis_names = [n for n in names if "/analysis/" in n]

    assert not any(n.endswith("_cranial_asymmetry.png") for n in analysis_names)
    report_name = next(n for n in analysis_names if n.endswith("_report_cranial.json"))
    report = json.loads(zf.read(report_name))
    assert "asymmetry" not in report
    assert "craniometrics" in report  # measurements still ticked (default)

    # no mesh files anywhere in the zip - only the analysis/ subfolder
    assert not any("/analysis/" not in n for n in names)


def test_save_analysis_include_meshes_false_skips_mesh_files(client, landmarks_payload, tmp_path):
    session_id = _upload(client)
    assert _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}, timeout=60) == "done"

    response = client.post(
        f"/api/sessions/{session_id}/save/analysis",
        json={"dest_dir": str(tmp_path), "include_meshes": False},
    )
    assert response.status_code == 200, response.text
    results_dir = Path(response.json()["saved_to"]).parent
    assert (results_dir / "analysis").is_dir()
    assert not any(results_dir.glob("*.ply"))  # no mesh files written into a folder that never had any


def test_save_analysis_cohort_xlsx_create_append_and_replace(client, landmarks_payload, tmp_path):
    import shutil

    def _open_run_and_save(filename: str, cohort_path: Path, extra_metadata: dict | None = None):
        tmp_mesh = tmp_path / filename
        shutil.copy(TEMPLATE_PATH, tmp_mesh)
        open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
        session_id = open_response.json()["session_id"]
        status = _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload})
        assert status == "done"

        metadata = {"file_name": filename, "file_path": str(tmp_mesh)}
        metadata.update(extra_metadata or {})
        save_response = client.post(
            f"/api/sessions/{session_id}/save/analysis",
            json={"metadata": metadata, "cohort_xlsx_path": str(cohort_path)},
        )
        assert save_response.status_code == 200, save_response.text
        return tmp_mesh

    cohort_path = tmp_path / "cohort" / "cohort.xlsx"
    cohort_path.parent.mkdir()

    # first export creates the cohort file with one row
    mesh_a = _open_run_and_save("patient_a.ply", cohort_path, {"sex": "male", "patient_id": "MRN-A"})
    rows = _read_xlsx_rows(cohort_path.read_bytes())
    assert len(rows) == 1
    assert rows[0]["sex"] == "male"
    # the shared cohort file gets a cohort_id instead of every
    # patient-identifying field (patient_id, file_name/file_path,
    # date_of_birth/date_of_intervention) - none of them ever appear in
    # it, see the id-mapping assertions below for where they do end up
    assert rows[0]["cohort_id"] == "C00001"
    for key in ("patient_id", "file_name", "file_path", "date_of_birth", "date_of_intervention"):
        assert key not in rows[0]

    mapping_rows = _read_xlsx_rows(_id_mapping_path(cohort_path).read_bytes())
    assert len(mapping_rows) == 1
    assert mapping_rows[0]["cohort_id"] == "C00001"
    assert mapping_rows[0]["patient_id"] == "MRN-A"
    assert mapping_rows[0]["file_path"] == str(mesh_a)

    # exporting a different file_path appends a second row, with its own id
    mesh_b = _open_run_and_save("patient_b.ply", cohort_path, {"sex": "female", "patient_id": "MRN-B"})
    rows = _read_xlsx_rows(cohort_path.read_bytes())
    assert len(rows) == 2
    assert {r["cohort_id"] for r in rows} == {"C00001", "C00002"}
    mapping_rows = _read_xlsx_rows(_id_mapping_path(cohort_path).read_bytes())
    assert {r["cohort_id"]: r["file_path"] for r in mapping_rows} == {
        "C00001": str(mesh_a), "C00002": str(mesh_b),
    }

    # re-exporting the SAME file_path updates the row in place, no
    # duplicate, and reuses rather than reassigns its cohort_id - matched
    # via the id-mapping file's own file_path column, since the cohort
    # file itself no longer carries one to match on directly
    _open_run_and_save("patient_a.ply", cohort_path, {"sex": "male", "treatment": "helmet", "patient_id": "MRN-A"})
    rows = _read_xlsx_rows(cohort_path.read_bytes())
    assert len(rows) == 2
    updated = next(r for r in rows if r["cohort_id"] == "C00001")
    assert updated["treatment"] == "helmet"


def test_save_analysis_cohort_xlsx_records_nicp_mesh_path(client, landmarks_payload, tmp_path):
    # the cohort workspace's mean-shape feature has to be able to find each
    # patient's NICP-fitted mesh from the cohort spreadsheet alone - this is
    # the column that makes that possible (see results_bundle._nicp_mesh_path).
    import shutil

    tmp_mesh = tmp_path / "patient.ply"
    shutil.copy(TEMPLATE_PATH, tmp_mesh)
    open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
    session_id = open_response.json()["session_id"]
    assert _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}) == "done"

    status = _run_run(
        client,
        session_id,
        {
            "nicp": {
                "template": "clipped_template_xy",
                "alpha_start": 50,
                "alpha_end": 1,
                "alpha_steps": 3,
                "inner_iters": 1,
                "dist_threshold": 50.0,
            }
        },
    )
    assert status == "done"

    cohort_path = tmp_path / "cohort.xlsx"
    save_response = client.post(
        f"/api/sessions/{session_id}/save/analysis",
        json={"metadata": {"file_name": "patient.ply", "file_path": str(tmp_mesh)}, "cohort_xlsx_path": str(cohort_path)},
    )
    assert save_response.status_code == 200, save_response.text

    rows = _read_xlsx_rows(cohort_path.read_bytes())
    assert len(rows) == 1
    nicp_path = Path(rows[0]["nicp_mesh_path"])
    assert nicp_path.is_file()
    assert nicp_path.name.endswith("N.ply")


def test_save_analysis_cohort_xlsx_nicp_mesh_path_blank_without_a_fit(client, landmarks_payload, tmp_path):
    import shutil

    tmp_mesh = tmp_path / "patient.ply"
    shutil.copy(TEMPLATE_PATH, tmp_mesh)
    open_response = client.post("/api/sessions/from-paths", json={"paths": [str(tmp_mesh)]})
    session_id = open_response.json()["session_id"]
    assert _clip_and_run(client, session_id, {"target": "cranium", "landmarks": landmarks_payload}) == "done"

    cohort_path = tmp_path / "cohort.xlsx"
    save_response = client.post(
        f"/api/sessions/{session_id}/save/analysis",
        json={"metadata": {"file_name": "patient.ply", "file_path": str(tmp_mesh)}, "cohort_xlsx_path": str(cohort_path)},
    )
    assert save_response.status_code == 200, save_response.text
    rows = _read_xlsx_rows(cohort_path.read_bytes())
    assert rows[0]["nicp_mesh_path"] == ""


# --- /api/cohort/* -------------------------------------------------------
# math itself (load_cohort_xlsx, mean_shape) is covered directly in
# test_cohort.py - these just exercise the request/response plumbing.


def _write_cohort_xlsx(path, header, rows) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for row in rows:
        ws.append(row)
    wb.save(path)


def test_cohort_load_reads_a_real_path(client, tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    _write_cohort_xlsx(cohort_path, ["patient_id", "diagnosis"], [["P001", "metopic"]])

    response = client.post("/api/cohort/load", json={"path": str(cohort_path)})

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["columns"] == ["patient_id", "diagnosis"]
    assert body["rows"] == [{"patient_id": "P001", "diagnosis": "metopic"}]


def test_cohort_load_missing_file_is_a_400(client, tmp_path):
    response = client.post("/api/cohort/load", json={"path": str(tmp_path / "does_not_exist.xlsx")})
    assert response.status_code == 400


def test_cohort_demo_returns_the_shipped_demo_cohort(client):
    response = client.get("/api/cohort/demo")

    assert response.status_code == 200, response.text
    body = response.json()
    assert len(body["rows"]) > 0
    assert "nicp_mesh_path" in body["columns"]
    nicp_rows = [r for r in body["rows"] if r["nicp_used"] == "yes"]
    assert len(nicp_rows) > 0
    assert Path(nicp_rows[0]["nicp_mesh_path"]).is_file()


def test_cohort_upload_reads_uploaded_bytes(client, tmp_path):
    cohort_path = tmp_path / "cohort.xlsx"
    _write_cohort_xlsx(cohort_path, ["patient_id"], [["P001"]])

    with open(cohort_path, "rb") as f:
        response = client.post(
            "/api/cohort/upload", files=[("file", ("cohort.xlsx", f, "application/octet-stream"))]
        )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["columns"] == ["patient_id"]
    assert body["rows"] == [{"patient_id": "P001"}]


def test_cohort_stats_test_two_groups_returns_parametric_and_rank_based(client):
    response = client.post(
        "/api/cohort/stats-test",
        json={"values": {"control": [1.0, 2.0, 3.0, 4.0], "metopic": [5.0, 6.0, 7.0, 8.0]}},
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["n_groups"] == 2
    assert body["test_name"] == "Welch's t-test"
    assert body["alternative_test_name"] == "Mann-Whitney U"
    assert body["group_sizes"] == {"control": 4, "metopic": 4}


def test_cohort_stats_test_three_groups_uses_anova_and_kruskal(client):
    response = client.post(
        "/api/cohort/stats-test",
        json={"values": {"a": [1.0, 2.0, 3.0], "b": [4.0, 5.0, 6.0], "c": [7.0, 8.0, 9.0]}},
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["n_groups"] == 3
    assert body["test_name"] == "One-way ANOVA"
    assert body["alternative_test_name"] == "Kruskal-Wallis H"


def test_cohort_stats_test_requires_at_least_two_groups(client):
    response = client.post("/api/cohort/stats-test", json={"values": {"only_one": [1.0, 2.0]}})
    assert response.status_code == 400


def _write_tetrahedron(path, offset=(0.0, 0.0, 0.0)) -> None:
    vertices = np.array(
        [[0.0, 0.0, 0.0], [1.0, 0.0, 0.0], [0.0, 1.0, 0.0], [0.0, 0.0, 1.0]]
    ) + np.array(offset)
    faces = np.array([[0, 1, 2], [0, 1, 3], [0, 2, 3], [1, 2, 3]])
    trimesh.Trimesh(vertices=vertices, faces=faces, process=False).export(path)


def test_cohort_mean_shape_computes_and_serves_the_mesh(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))

    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": [str(path_a), str(path_b)]})

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["vertex_count"] == 4
    assert body["source_count"] == 2
    assert len(body["heatmap"]) == 4
    assert all(v == pytest.approx(1.0) for v in body["heatmap"])

    mesh_response = client.get(f"/api/cohort/mean-shape/{body['result_id']}/mesh")
    assert mesh_response.status_code == 200
    assert mesh_response.headers["content-type"] == "model/gltf-binary"
    assert len(mesh_response.content) > 0


def test_cohort_mean_shape_rejects_mismatched_topology(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a)
    trimesh.creation.box().export(path_b)

    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": [str(path_a), str(path_b)]})
    assert response.status_code == 400


def test_cohort_mean_shape_missing_file_is_a_400(client, tmp_path):
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": [str(tmp_path / "nope.ply")]})
    assert response.status_code == 400


def test_cohort_mean_shape_reference_diff_matches_the_template_exactly(client):
    # averaging two identical copies of the same shipped template gives
    # back that template's own vertices exactly (no floating-point drift -
    # (x + x) / 2 == x bit-for-bit), so a reference-diff against that same
    # template should be exactly (or extremely nearly) zero everywhere.
    template_path = TEMPLATES_DIR / "clipped_template_xy_com.ply"
    response = client.post(
        "/api/cohort/mean-shape", json={"mesh_paths": [str(template_path), str(template_path)]}
    )
    assert response.status_code == 200, response.text
    result_id = response.json()["result_id"]

    diff_response = client.get(
        f"/api/cohort/mean-shape/{result_id}/reference-diff", params={"template": "clipped_template_xy_com"}
    )
    assert diff_response.status_code == 200, diff_response.text
    heatmap = diff_response.json()["heatmap"]
    assert len(heatmap) == response.json()["vertex_count"]
    assert all(abs(v) < 1e-6 for v in heatmap)


def test_cohort_mean_shape_reference_diff_rejects_mismatched_topology(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": [str(path_a), str(path_b)]})
    result_id = response.json()["result_id"]

    diff_response = client.get(
        f"/api/cohort/mean-shape/{result_id}/reference-diff", params={"template": "clipped_template_xy_com"}
    )
    assert diff_response.status_code == 400


def test_cohort_mean_shape_reference_diff_unknown_template_is_a_400(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a)
    _write_tetrahedron(path_b)
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": [str(path_a), str(path_b)]})
    result_id = response.json()["result_id"]

    diff_response = client.get(
        f"/api/cohort/mean-shape/{result_id}/reference-diff", params={"template": "not_a_real_template"}
    )
    assert diff_response.status_code == 400


def test_cohort_mean_shape_reference_diff_unknown_result_id_is_a_404(client):
    response = client.get(
        "/api/cohort/mean-shape/not-a-real-id/reference-diff", params={"template": "clipped_template_xy_com"}
    )
    assert response.status_code == 404


def test_cohort_export_xlsx_multiple_sheets(client):
    response = client.post(
        "/api/cohort/export-xlsx",
        json={
            "sheets": [
                {
                    "title": "cohort data",
                    "columns": ["patient_id", "diagnosis", "cephalic_index"],
                    "rows": [
                        {"patient_id": "P001", "diagnosis": "control", "cephalic_index": "81.20"},
                        {"patient_id": "0042", "diagnosis": "scaphocephaly", "cephalic_index": "64.50"},
                    ],
                },
                {
                    "title": "test result",
                    "columns": ["test", "statistic", "p_value"],
                    "rows": [{"test": "Welch's t-test", "statistic": "3.21", "p_value": "0.012"}],
                },
            ],
            "filename": "my export!!.xlsx",
        },
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert response.headers["content-disposition"] == 'attachment; filename="my-export-.xlsx"'

    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["cohort data", "test result"]
    ws = wb["cohort data"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("patient_id", "diagnosis", "cephalic_index")
    # patient_id stays text (mixed "P001"/"0042" - not every cell parses as
    # a number) with its leading zero intact; cephalic_index becomes a real
    # numeric cell since every cell in it does parse.
    assert rows[2][0] == "0042"
    assert rows[1][2] == pytest.approx(81.2)
    assert ws["A1"].font.bold is True


def test_cohort_export_xlsx_deduplicates_sheet_names(client):
    response = client.post(
        "/api/cohort/export-xlsx",
        json={
            "sheets": [
                {"title": "data", "columns": ["a"], "rows": [{"a": "1"}]},
                {"title": "data", "columns": ["a"], "rows": [{"a": "2"}]},
            ],
        },
    )
    assert response.status_code == 200, response.text
    wb = load_workbook(BytesIO(response.content))
    assert wb.sheetnames == ["data", "data (2)"]


def test_cohort_report_cranium_target(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "clipped_template_xy_com"][:6]

    response = client.post(
        "/api/cohort/report",
        json={
            "mesh_paths": mesh_paths, "target": "cranium", "group_label": "scaphocephaly, pre-op",
            "include_spread_bands": True,
        },
    )

    assert response.status_code == 200, response.text
    assert response.headers["content-type"] == "application/pdf"
    assert response.headers["content-disposition"] == 'attachment; filename="scaphocephaly-pre-op_report.pdf"'
    assert response.content[:4] == b"%PDF"


def test_cohort_report_face_target_without_spread_bands(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "template_face"][:6]

    response = client.post(
        "/api/cohort/report",
        json={"mesh_paths": mesh_paths, "target": "face", "group_label": "trigonocephaly", "include_spread_bands": False},
    )

    assert response.status_code == 200, response.text
    assert response.content[:4] == b"%PDF"


def test_cohort_report_face_target_with_spread_bands_includes_metopic_band(client):
    # a face-target report with spread bands on should compute the metopic
    # band too (not just sagittal) - a bigger PDF than the same report with
    # bands off is a simple, robust way to check the extra page's content
    # actually got drawn without parsing PDF internals.
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "template_face"][:6]

    with_bands = client.post(
        "/api/cohort/report",
        json={"mesh_paths": mesh_paths, "target": "face", "group_label": "trigonocephaly", "include_spread_bands": True},
    )
    without_bands = client.post(
        "/api/cohort/report",
        json={"mesh_paths": mesh_paths, "target": "face", "group_label": "trigonocephaly", "include_spread_bands": False},
    )

    assert with_bands.status_code == 200, with_bands.text
    assert without_bands.status_code == 200, without_bands.text
    assert len(with_bands.content) != len(without_bands.content)


def test_cohort_report_invalid_target_is_a_400(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))

    response = client.post(
        "/api/cohort/report", json={"mesh_paths": [str(path_a), str(path_b)], "target": "nonsense"}
    )
    assert response.status_code == 400


def test_cohort_report_missing_file_is_a_400(client, tmp_path):
    response = client.post(
        "/api/cohort/report", json={"mesh_paths": [str(tmp_path / "nope.ply")], "target": "cranium"}
    )
    assert response.status_code == 400


def test_cohort_mean_shape_download_serves_a_named_ply(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": [str(path_a), str(path_b)]})
    result_id = response.json()["result_id"]

    download = client.get(f"/api/cohort/mean-shape/{result_id}/download", params={"filename": "trigonocephaly_pre-op_mean.ply"})

    assert download.status_code == 200, download.text
    assert download.headers["content-disposition"] == 'attachment; filename="trigonocephaly_pre-op_mean.ply"'
    assert len(download.content) > 0
    loaded = trimesh.load(BytesIO(download.content), file_type="ply", process=False)
    assert len(loaded.vertices) == 4


def test_cohort_mean_shape_download_sanitizes_the_filename(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": [str(path_a), str(path_b)]})
    result_id = response.json()["result_id"]

    download = client.get(
        f"/api/cohort/mean-shape/{result_id}/download", params={"filename": '../../etc/passwd"; evil.ply'}
    )

    assert download.status_code == 200, download.text
    disposition = download.headers["content-disposition"]
    assert disposition.startswith('attachment; filename="') and disposition.endswith('.ply"')
    filename = disposition[len('attachment; filename="') : -1]
    assert ".." not in filename
    assert "/" not in filename
    assert '"' not in filename
    assert ";" not in filename


def test_cohort_mean_shape_download_unknown_result_id_is_a_404(client):
    response = client.get("/api/cohort/mean-shape/not-a-real-id/download")
    assert response.status_code == 404


def test_cohort_sagittal_band_cranium_target(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "clipped_template_xy_com"]

    response = client.post("/api/cohort/sagittal-band", json={"mesh_paths": mesh_paths, "target": "cranium"})

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["source_count"] == len(mesh_paths)
    assert len(body["y"]) == len(body["mean_z"]) == len(body["sd_z"])
    assert all(v >= 0 for v in body["sd_z"])


def test_cohort_sagittal_band_invalid_target_is_a_400(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))

    response = client.post(
        "/api/cohort/sagittal-band", json={"mesh_paths": [str(path_a), str(path_b)], "target": "nonsense"}
    )
    assert response.status_code == 400


def test_cohort_sagittal_band_missing_file_is_a_400(client, tmp_path):
    response = client.post(
        "/api/cohort/sagittal-band", json={"mesh_paths": [str(tmp_path / "nope.ply")], "target": "cranium"}
    )
    assert response.status_code == 400


def test_cohort_sagittal_band_rejects_mismatched_topology(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    trimesh.creation.box().export(path_b)

    response = client.post(
        "/api/cohort/sagittal-band", json={"mesh_paths": [str(path_a), str(path_b)], "target": "cranium"}
    )
    assert response.status_code == 400


def test_cohort_hc_ring_band_cranium_target(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "clipped_template_xy_com"]

    response = client.post("/api/cohort/hc-ring-band", json={"mesh_paths": mesh_paths, "target": "cranium"})

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["source_count"] == len(mesh_paths)
    assert body["closed"] is True
    assert len(body["mean"]) == len(body["inner"]) == len(body["outer"]) == 72
    assert set(body["mean"][0].keys()) == {"x", "y", "z"}


def test_cohort_hc_ring_band_invalid_target_is_a_400(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))

    response = client.post(
        "/api/cohort/hc-ring-band", json={"mesh_paths": [str(path_a), str(path_b)], "target": "nonsense"}
    )
    assert response.status_code == 400


def test_cohort_hc_ring_band_missing_file_is_a_400(client, tmp_path):
    response = client.post(
        "/api/cohort/hc-ring-band", json={"mesh_paths": [str(tmp_path / "nope.ply")], "target": "cranium"}
    )
    assert response.status_code == 400


def test_cohort_hc_ring_band_rejects_mismatched_topology(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    trimesh.creation.box().export(path_b)

    response = client.post(
        "/api/cohort/hc-ring-band", json={"mesh_paths": [str(path_a), str(path_b)], "target": "cranium"}
    )
    assert response.status_code == 400


def test_cohort_metopic_band_face_target(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "template_face"]

    response = client.post("/api/cohort/metopic-band", json={"mesh_paths": mesh_paths, "target": "face"})

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["source_count"] == len(mesh_paths)
    assert body["closed"] is False
    assert len(body["mean"]) == len(body["inner"]) == len(body["outer"])
    assert set(body["mean"][0].keys()) == {"x", "y", "z"}


def test_cohort_metopic_band_invalid_target_is_a_400(client, tmp_path):
    path_a = tmp_path / "a.ply"
    path_b = tmp_path / "b.ply"
    _write_tetrahedron(path_a, offset=(-1.0, 0.0, 0.0))
    _write_tetrahedron(path_b, offset=(1.0, 0.0, 0.0))

    response = client.post(
        "/api/cohort/metopic-band", json={"mesh_paths": [str(path_a), str(path_b)], "target": "nonsense"}
    )
    assert response.status_code == 400


def test_cohort_metopic_band_missing_file_is_a_400(client, tmp_path):
    response = client.post(
        "/api/cohort/metopic-band", json={"mesh_paths": [str(tmp_path / "nope.ply")], "target": "face"}
    )
    assert response.status_code == 400


def test_cohort_mean_shape_measurements_cranium_target(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "clipped_template_xy_com"][:5]
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": mesh_paths})
    result_id = response.json()["result_id"]

    measurements = client.get(f"/api/cohort/mean-shape/{result_id}/measurements", params={"target": "cranium"})

    assert measurements.status_code == 200, measurements.text
    body = measurements.json()
    assert body["craniometrics"] is not None
    assert body["craniometrics"]["cephalic_index"] > 0
    assert body["metopic"] is None
    assert body["asymmetry"]["mean_asymmetry_index"] >= 0
    assert body["frontal_bossing"] is not None


def test_cohort_mean_shape_measurements_face_target(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "template_face"][:5]
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": mesh_paths})
    result_id = response.json()["result_id"]

    measurements = client.get(f"/api/cohort/mean-shape/{result_id}/measurements", params={"target": "face"})

    assert measurements.status_code == 200, measurements.text
    body = measurements.json()
    assert body["metopic"] is not None
    assert body["craniometrics"] is None
    assert body["asymmetry"]["mean_asymmetry_index"] >= 0


def test_cohort_mean_shape_measurements_invalid_target_is_a_400(client):
    from craniumpy_core.cohort import load_demo_cohort

    _columns, rows = load_demo_cohort()
    mesh_paths = [r["nicp_mesh_path"] for r in rows if r["nicp_template"] == "clipped_template_xy_com"][:5]
    response = client.post("/api/cohort/mean-shape", json={"mesh_paths": mesh_paths})
    result_id = response.json()["result_id"]

    measurements = client.get(f"/api/cohort/mean-shape/{result_id}/measurements", params={"target": "nonsense"})
    assert measurements.status_code == 400


def test_cohort_mean_shape_measurements_unknown_result_id_is_a_404(client):
    response = client.get("/api/cohort/mean-shape/not-a-real-id/measurements", params={"target": "cranium"})
    assert response.status_code == 404


def test_cohort_mean_shape_unknown_result_id_is_a_404(client):
    response = client.get("/api/cohort/mean-shape/not-a-real-id/mesh")
    assert response.status_code == 404
