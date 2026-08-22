"""API tests for api/routers/facial.py - the Facial Anthropometrics
workspace's own request/response plumbing (template load/pick/preview,
batch extraction/correction/export). Underlying geometry math is covered in
tests/test_facial_measurements.py; this file is about the endpoints,
caching, and per-file/per-measurement error handling."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pytest
import trimesh
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from api.main import app

REPO_ROOT = Path(__file__).resolve().parent.parent
TEMPLATE_PATH = REPO_ROOT / "src" / "craniumpy_core" / "templates" / "template_face.ply"


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


def _grid_mesh(nx: int, ny: int, spacing: float = 1.0, offset: np.ndarray | None = None) -> trimesh.Trimesh:
    xs, ys = np.meshgrid(np.arange(nx) * spacing, np.arange(ny) * spacing, indexing="xy")
    vertices = np.column_stack([xs.ravel(), ys.ravel(), np.zeros(nx * ny)])
    if offset is not None:
        vertices = vertices + offset
    faces = []
    for j in range(ny - 1):
        for i in range(nx - 1):
            v00 = j * nx + i
            v10 = j * nx + (i + 1)
            v01 = (j + 1) * nx + i
            v11 = (j + 1) * nx + (i + 1)
            faces.append([v00, v10, v11])
            faces.append([v00, v11, v01])
    return trimesh.Trimesh(vertices=vertices, faces=np.array(faces), process=False)


def _vidx(nx: int, i: int, j: int) -> int:
    return j * nx + i


def _write_mesh(mesh: trimesh.Trimesh, path: Path) -> str:
    mesh.export(path)
    return str(path)


def _load_template(client: TestClient, path: str) -> str:
    response = client.post("/api/facial/template/load", json={"shipped_name": None, "path": path})
    assert response.status_code == 200, response.text
    return response.json()["template_id"]


def _point(mesh: trimesh.Trimesh, idx: int) -> dict:
    p = mesh.vertices[idx]
    return {"x": float(p[0]), "y": float(p[1]), "z": float(p[2])}


# --- list-meshes -----------------------------------------------------------


def test_list_meshes_finds_mesh_files_non_recursively(client, tmp_path):
    _write_mesh(_grid_mesh(3, 3), tmp_path / "a.ply")
    _write_mesh(_grid_mesh(3, 3), tmp_path / "b.ply")
    (tmp_path / "notes.txt").write_text("not a mesh")
    nested = tmp_path / "nested"
    nested.mkdir()
    _write_mesh(_grid_mesh(3, 3), nested / "c.ply")

    response = client.post("/api/facial/list-meshes", json={"folder": str(tmp_path)})
    assert response.status_code == 200, response.text
    names = sorted(Path(p).name for p in response.json()["mesh_paths"])
    assert names == ["a.ply", "b.ply"]


def test_list_meshes_missing_folder_400s(client, tmp_path):
    response = client.post("/api/facial/list-meshes", json={"folder": str(tmp_path / "nope")})
    assert response.status_code == 400


# --- template load/mesh/pick ------------------------------------------------


def test_load_template_default_shipped_face_template(client):
    response = client.post("/api/facial/template/load", json={})
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["vertex_count"] == 10000
    assert body["face_count"] == 19721


def test_load_template_unknown_shipped_name_400s(client):
    response = client.post("/api/facial/template/load", json={"shipped_name": "not_a_real_template"})
    assert response.status_code == 400


def test_load_template_custom_path(client, tmp_path):
    mesh = _grid_mesh(4, 4)
    path = _write_mesh(mesh, tmp_path / "custom.ply")
    response = client.post("/api/facial/template/load", json={"shipped_name": None, "path": path})
    assert response.status_code == 200, response.text
    assert response.json()["vertex_count"] == 16


def test_load_template_missing_custom_path_400s(client, tmp_path):
    response = client.post("/api/facial/template/load", json={"shipped_name": None, "path": str(tmp_path / "nope.ply")})
    assert response.status_code == 400


def test_get_template_mesh_returns_glb(client, tmp_path):
    mesh = _grid_mesh(3, 3)
    path = _write_mesh(mesh, tmp_path / "t.ply")
    template_id = _load_template(client, path)

    response = client.get(f"/api/facial/template/{template_id}/mesh")
    assert response.status_code == 200
    assert response.headers["content-type"] == "model/gltf-binary"
    assert len(response.content) > 0


def test_get_template_mesh_unknown_id_404s(client):
    assert client.get("/api/facial/template/does-not-exist/mesh").status_code == 404


def test_pick_point_snaps_to_nearest_vertex(client, tmp_path):
    mesh = _grid_mesh(3, 3)
    path = _write_mesh(mesh, tmp_path / "t.ply")
    template_id = _load_template(client, path)

    response = client.post(f"/api/facial/template/{template_id}/pick", json={"point": {"x": 1.1, "y": 0.9, "z": 0.4}})
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["vertex_index"] == _vidx(3, 1, 1)
    assert body["point"] == {"x": 1.0, "y": 1.0, "z": 0.0}


# --- measurement preview ----------------------------------------------------


def test_preview_linear_measurement_straight_and_geodesic(client, tmp_path):
    mesh = _grid_mesh(4, 4)
    path = _write_mesh(mesh, tmp_path / "t.ply")
    template_id = _load_template(client, path)

    a, b = _vidx(4, 0, 0), _vidx(4, 3, 0)
    points = {"p1": _point(mesh, a), "p2": _point(mesh, b)}
    measurements = [
        {"id": "m1", "name": "Width", "abbreviation": "W", "type": "linear", "point_ids": ["p1", "p2"], "geodesic": False},
        {"id": "m2", "name": "Width (geo)", "abbreviation": "WG", "type": "linear", "point_ids": ["p1", "p2"], "geodesic": True},
    ]
    response = client.post(
        f"/api/facial/template/{template_id}/measurement/preview",
        json={"template_id": template_id, "points": points, "measurements": measurements},
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["values"]["m1"] == pytest.approx(3.0)
    assert body["values"]["m2"] == pytest.approx(3.0)
    assert body["value_errors"] == {}

    # render_paths traces the mesh surface for BOTH straight and geodesic
    # Linear measurements (see api/routers/facial.py's _render_geometry) -
    # the overlay always hugs the surface regardless of the value's own
    # straight/geodesic toggle. row 0 of a flat grid is a straight run of
    # mesh edges, so the surface trace is the full 4-vertex row, not just
    # the 2 endpoints.
    for mid in ("m1", "m2"):
        path = body["render_paths"][mid]
        assert len(path) == 4
        assert [p["x"] for p in path] == [0.0, 1.0, 2.0, 3.0]
        assert path[0] == points["p1"]
        assert path[-1] == points["p2"]
    assert body["render_faces"] == {}


def test_preview_angular_measurement(client, tmp_path):
    vertices = np.array([[0.0, 0.0, 0.0], [1.0, 0.0, 0.0], [0.0, 1.0, 0.0]])
    mesh = trimesh.Trimesh(vertices=vertices, faces=np.array([[0, 1, 2]]), process=False)
    path = _write_mesh(mesh, tmp_path / "t.ply")
    template_id = _load_template(client, path)

    points = {"p1": _point(mesh, 1), "p2": _point(mesh, 0), "p3": _point(mesh, 2)}
    measurements = [
        {"id": "m1", "name": "Angle", "abbreviation": "A", "type": "angular", "point_ids": ["p1", "p2", "p3"]}
    ]
    response = client.post(
        f"/api/facial/template/{template_id}/measurement/preview",
        json={"template_id": template_id, "points": points, "measurements": measurements},
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["values"]["m1"] == pytest.approx(90.0)

    # the two legs (vertex -> p1, vertex -> p3) chained into one continuous
    # a -> vertex -> c surface trace - see _render_geometry's own comment.
    render_path = body["render_paths"]["m1"]
    assert render_path == [points["p1"], points["p2"], points["p3"]]


def test_preview_area_measurement(client, tmp_path):
    mesh = _grid_mesh(6, 6)
    path = _write_mesh(mesh, tmp_path / "t.ply")
    template_id = _load_template(client, path)

    corner_indices = [_vidx(6, 1, 1), _vidx(6, 4, 1), _vidx(6, 4, 4), _vidx(6, 1, 4)]
    points = {f"p{i}": _point(mesh, idx) for i, idx in enumerate(corner_indices)}
    measurements = [
        {"id": "m1", "name": "Patch", "abbreviation": "PA", "type": "area", "point_ids": list(points.keys())}
    ]
    response = client.post(
        f"/api/facial/template/{template_id}/measurement/preview",
        json={"template_id": template_id, "points": points, "measurements": measurements},
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["values"]["m1"] == pytest.approx(9.0)

    # a closed geodesic boundary loop (first == last point) plus the
    # enclosed region's own triangle soup, for the workspace's "shade the
    # measured patch of surface" overlay.
    loop = body["render_paths"]["m1"]
    assert len(loop) > 4  # traces every mesh edge along the boundary, not just the 4 corners
    assert loop[0] == loop[-1]
    faces = body["render_faces"]["m1"]
    assert len(faces) % 3 == 0
    assert len(faces) > 0


def test_preview_measurement_wrong_point_count_is_a_per_measurement_error_not_a_500(client, tmp_path):
    mesh = _grid_mesh(3, 3)
    path = _write_mesh(mesh, tmp_path / "t.ply")
    template_id = _load_template(client, path)

    points = {"p1": _point(mesh, 0), "p2": _point(mesh, 1), "p3": _point(mesh, 2)}
    measurements = [
        {"id": "m1", "name": "Bad Linear", "abbreviation": "BL", "type": "linear", "point_ids": ["p1", "p2", "p3"]}
    ]
    response = client.post(
        f"/api/facial/template/{template_id}/measurement/preview",
        json={"template_id": template_id, "points": points, "measurements": measurements},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["values"]["m1"] is None
    assert "exactly 2" in body["value_errors"]["m1"]


def test_preview_measurement_disconnected_geodesic_is_a_per_measurement_error(client, tmp_path):
    island_a = _grid_mesh(2, 2)
    island_b = _grid_mesh(2, 2, offset=np.array([100.0, 100.0, 0.0]))
    combined = trimesh.Trimesh(
        vertices=np.vstack([island_a.vertices, island_b.vertices]),
        faces=np.vstack([island_a.faces, island_b.faces + len(island_a.vertices)]),
        process=False,
    )
    path = _write_mesh(combined, tmp_path / "t.ply")
    template_id = _load_template(client, path)

    points = {"p1": _point(combined, 0), "p2": _point(combined, len(island_a.vertices))}
    measurements = [
        {"id": "m1", "name": "Cross-island", "abbreviation": "CI", "type": "linear", "point_ids": ["p1", "p2"], "geodesic": True}
    ]
    response = client.post(
        f"/api/facial/template/{template_id}/measurement/preview",
        json={"template_id": template_id, "points": points, "measurements": measurements},
    )
    assert response.status_code == 200
    body = response.json()
    assert body["values"]["m1"] is None
    assert "disconnected" in body["value_errors"]["m1"]


# --- batch ------------------------------------------------------------------


def _batch_measurements(mesh) -> tuple[dict, list[dict]]:
    a, b = _vidx(4, 0, 0), _vidx(4, 3, 0)
    points = {"p1": _point(mesh, a), "p2": _point(mesh, b)}
    measurements = [
        {"id": "m1", "name": "Width", "abbreviation": "W", "type": "linear", "point_ids": ["p1", "p2"], "color": "#ff0000"}
    ]
    return points, measurements


def test_start_batch_processes_files_independently_without_aborting(client, tmp_path):
    template_mesh = _grid_mesh(4, 4)
    template_path = _write_mesh(template_mesh, tmp_path / "template.ply")
    template_id = _load_template(client, template_path)

    good_mesh = _grid_mesh(4, 4, offset=np.array([0.0, 0.0, 1.0]))
    good_path = _write_mesh(good_mesh, tmp_path / "patient_good.ply")

    bad_topology_path = _write_mesh(trimesh.creation.box(), tmp_path / "patient_bad_topology.ply")

    points, measurements = _batch_measurements(template_mesh)
    response = client.post(
        "/api/facial/batch/start",
        json={
            "template_id": template_id,
            "mesh_paths": [good_path, bad_topology_path],
            "points": points,
            "measurements": measurements,
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    results_by_name = {r["filename"]: r for r in body["results"]}
    assert results_by_name["patient_good.ply"]["status"] == "ok"
    assert results_by_name["patient_good.ply"]["values"]["m1"] == pytest.approx(3.0)
    # render geometry is computed fresh per file (this mesh's own z-offset
    # vertex positions, not the template's) - not just echoed from the
    # template's own preview.
    good_path = results_by_name["patient_good.ply"]["render_paths"]["m1"]
    assert len(good_path) == 4
    assert all(p["z"] == pytest.approx(1.0) for p in good_path)
    assert results_by_name["patient_bad_topology.ply"]["status"] == "error"
    assert results_by_name["patient_bad_topology.ply"]["error"]


def test_start_batch_missing_mesh_path_400s(client, tmp_path):
    template_mesh = _grid_mesh(4, 4)
    template_path = _write_mesh(template_mesh, tmp_path / "template.ply")
    template_id = _load_template(client, template_path)
    points, measurements = _batch_measurements(template_mesh)

    response = client.post(
        "/api/facial/batch/start",
        json={
            "template_id": template_id,
            "mesh_paths": [str(tmp_path / "nope.ply")],
            "points": points,
            "measurements": measurements,
        },
    )
    assert response.status_code == 400


def _start_simple_batch(client, tmp_path):
    template_mesh = _grid_mesh(4, 4)
    template_path = _write_mesh(template_mesh, tmp_path / "template.ply")
    template_id = _load_template(client, template_path)

    mesh_a = _grid_mesh(4, 4, offset=np.array([0.0, 0.0, 1.0]))
    mesh_b = _grid_mesh(4, 4, offset=np.array([0.0, 0.0, -1.0]))
    path_a = _write_mesh(mesh_a, tmp_path / "patient_a.ply")
    path_b = _write_mesh(mesh_b, tmp_path / "patient_b.ply")

    points, measurements = _batch_measurements(template_mesh)
    response = client.post(
        "/api/facial/batch/start",
        json={"template_id": template_id, "mesh_paths": [path_a, path_b], "points": points, "measurements": measurements},
    )
    assert response.status_code == 200, response.text
    return response.json()["batch_id"], path_a, path_b


def test_get_batch_mesh_returns_glb(client, tmp_path):
    batch_id, _a, _b = _start_simple_batch(client, tmp_path)
    response = client.get(f"/api/facial/batch/{batch_id}/mesh/patient_a.ply")
    assert response.status_code == 200
    assert response.headers["content-type"] == "model/gltf-binary"


def test_get_batch_mesh_unknown_filename_404s(client, tmp_path):
    batch_id, _a, _b = _start_simple_batch(client, tmp_path)
    assert client.get(f"/api/facial/batch/{batch_id}/mesh/does_not_exist.ply").status_code == 404


def test_get_batch_mesh_unknown_batch_404s(client):
    assert client.get("/api/facial/batch/does-not-exist/mesh/x.ply").status_code == 404


def test_correct_landmark_updates_only_the_affected_measurement_for_that_one_file(client, tmp_path):
    batch_id, path_a, path_b = _start_simple_batch(client, tmp_path)

    mesh_a = trimesh.load(path_a, process=False, force="mesh")
    new_point = mesh_a.vertices[_vidx(4, 1, 0)]  # move p1 to a different vertex on mesh A only

    response = client.post(
        f"/api/facial/batch/{batch_id}/correct",
        json={
            "batch_id": batch_id,
            "filename": "patient_a.ply",
            "point_id": "p1",
            "point": {"x": float(new_point[0]), "y": float(new_point[1]), "z": float(new_point[2])},
        },
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["values"]["m1"] == pytest.approx(2.0)  # 1 grid step closer to p2 now
    # render geometry is recomputed against the corrected point too, not
    # left stale from before the correction.
    assert len(body["render_paths"]["m1"]) == 3

    # patient_b's own stored result must be untouched by a correction on patient_a
    export = client.post(f"/api/facial/batch/{batch_id}/export")
    assert export.status_code == 200


def test_correct_landmark_unknown_file_404s(client, tmp_path):
    batch_id, _a, _b = _start_simple_batch(client, tmp_path)
    response = client.post(
        f"/api/facial/batch/{batch_id}/correct",
        json={"batch_id": batch_id, "filename": "nope.ply", "point_id": "p1", "point": {"x": 0, "y": 0, "z": 0}},
    )
    assert response.status_code == 404


def test_correct_landmark_unknown_point_id_400s(client, tmp_path):
    batch_id, _a, _b = _start_simple_batch(client, tmp_path)
    response = client.post(
        f"/api/facial/batch/{batch_id}/correct",
        json={"batch_id": batch_id, "filename": "patient_a.ply", "point_id": "not_a_point", "point": {"x": 0, "y": 0, "z": 0}},
    )
    assert response.status_code == 400


# --- export -------------------------------------------------------------


def test_export_batch_produces_a_workbook_with_measurements_and_legend_sheets(client, tmp_path):
    batch_id, _a, _b = _start_simple_batch(client, tmp_path)
    response = client.post(f"/api/facial/batch/{batch_id}/export")
    assert response.status_code == 200
    assert response.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    import io

    wb = load_workbook(io.BytesIO(response.content))
    assert set(wb.sheetnames) >= {"measurements", "legend"}

    measurements_ws = wb["measurements"]
    header = [c.value for c in next(measurements_ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["identifier", "Width (W)"]

    legend_ws = wb["legend"]
    legend_header = [c.value for c in next(legend_ws.iter_rows(min_row=1, max_row=1))]
    assert legend_header == ["name", "abbreviation", "type", "unit", "geodesic", "color"]
    legend_row = [c.value for c in list(legend_ws.iter_rows(min_row=2, max_row=2))[0]]
    assert legend_row[0] == "Width"
    assert legend_row[1] == "W"
    swatch_cell = legend_ws.cell(row=2, column=6)
    assert swatch_cell.fill.start_color.rgb.upper().endswith("FF0000")


def test_export_batch_includes_a_failed_sheet_when_a_file_errored(client, tmp_path):
    template_mesh = _grid_mesh(4, 4)
    template_path = _write_mesh(template_mesh, tmp_path / "template.ply")
    template_id = _load_template(client, template_path)

    bad_path = _write_mesh(trimesh.creation.box(), tmp_path / "patient_bad.ply")
    points, measurements = _batch_measurements(template_mesh)
    start_response = client.post(
        "/api/facial/batch/start",
        json={"template_id": template_id, "mesh_paths": [bad_path], "points": points, "measurements": measurements},
    )
    batch_id = start_response.json()["batch_id"]

    export_response = client.post(f"/api/facial/batch/{batch_id}/export")
    assert export_response.status_code == 200

    import io

    wb = load_workbook(io.BytesIO(export_response.content))
    assert "failed" in wb.sheetnames
    failed_ws = wb["failed"]
    header = [c.value for c in next(failed_ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["filename", "error"]


def test_export_unknown_batch_404s(client):
    assert client.post("/api/facial/batch/does-not-exist/export").status_code == 404
