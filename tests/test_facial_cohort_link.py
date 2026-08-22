"""unit tests for src/craniumpy_core/facial_cohort_link.py and
facial_measurements_io.py - pure function tests, no FastAPI needed, same
style as test_cohort.py. router-level plumbing (the /facial-measurements/
load endpoint) is covered separately in test_facial_api.py."""

from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook

from craniumpy_core.facial_cohort_link import resolve_cohort_ids_by_filename
from craniumpy_core.facial_measurements_io import load_measurement_export_xlsx


def _mapping_row(cohort_id: str, file_path: str = "", file_name: str = "") -> dict:
    return {"cohort_id": cohort_id, "patient_id": "", "file_name": file_name, "file_path": file_path, "date_of_birth": "", "date_of_intervention": ""}


# --- resolve_cohort_ids_by_filename -----------------------------------------


def test_resolves_a_matched_filename():
    mapping_rows = [_mapping_row("C00001", file_path="C:/scans/patient_a.ply")]
    result = resolve_cohort_ids_by_filename(mapping_rows, ["patient_a.ply"])
    assert result.matched == {"patient_a.ply": "C00001"}
    assert result.unmatched == []
    assert result.ambiguous == {}


def test_reports_an_unmatched_filename_not_silently_dropped():
    mapping_rows = [_mapping_row("C00001", file_path="C:/scans/patient_a.ply")]
    result = resolve_cohort_ids_by_filename(mapping_rows, ["patient_a.ply", "patient_b.ply"])
    assert result.matched == {"patient_a.ply": "C00001"}
    assert result.unmatched == ["patient_b.ply"]


def test_reports_ambiguous_when_two_different_cohort_ids_share_a_basename():
    mapping_rows = [
        _mapping_row("C00001", file_path="C:/site_a/patient.ply"),
        _mapping_row("C00002", file_path="C:/site_b/patient.ply"),
    ]
    result = resolve_cohort_ids_by_filename(mapping_rows, ["patient.ply"])
    assert result.matched == {}
    assert result.ambiguous == {"patient.ply": ["C00001", "C00002"]}


def test_same_row_matched_via_file_path_or_file_name_is_not_ambiguous():
    # a genuine duplicate (the SAME cohort_id appearing twice, e.g. an
    # id-mapping file with both a file_path and file_name entry for the
    # same patient) must not falsely register as ambiguous.
    mapping_rows = [_mapping_row("C00001", file_path="C:/scans/patient_a.ply"), _mapping_row("C00001", file_name="patient_a.ply")]
    result = resolve_cohort_ids_by_filename(mapping_rows, ["patient_a.ply"])
    assert result.matched == {"patient_a.ply": "C00001"}
    assert result.ambiguous == {}


def test_falls_back_to_file_name_when_file_path_is_blank():
    mapping_rows = [_mapping_row("C00001", file_name="patient_a.ply")]
    result = resolve_cohort_ids_by_filename(mapping_rows, ["patient_a.ply"])
    assert result.matched == {"patient_a.ply": "C00001"}


def test_rows_without_a_cohort_id_are_ignored():
    mapping_rows = [_mapping_row("", file_path="C:/scans/patient_a.ply")]
    result = resolve_cohort_ids_by_filename(mapping_rows, ["patient_a.ply"])
    assert result.matched == {}
    assert result.unmatched == ["patient_a.ply"]


def test_empty_measurement_filenames_returns_empty_result():
    result = resolve_cohort_ids_by_filename([_mapping_row("C00001", file_path="C:/scans/a.ply")], [])
    assert result.matched == {}
    assert result.unmatched == []
    assert result.ambiguous == {}


# --- load_measurement_export_xlsx -------------------------------------------


def _write_measurement_export(path_or_buffer) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "measurements"
    ws.append(["identifier", "Width (W)"])
    ws.append(["patient_a.ply", "12.34"])
    ws.append(["patient_b.ply", ""])
    legend_ws = wb.create_sheet("legend")
    legend_ws.append(["name", "abbreviation", "type", "unit", "geodesic", "color"])
    legend_ws.append(["Width", "W", "linear", "mm", "no", ""])
    wb.save(path_or_buffer)


def test_load_measurement_export_xlsx_round_trips_from_a_real_path(tmp_path):
    path = tmp_path / "facial_measurements.xlsx"
    _write_measurement_export(path)

    columns, rows, legend_rows = load_measurement_export_xlsx(path)

    assert columns == ["identifier", "Width (W)"]
    assert rows == [
        {"identifier": "patient_a.ply", "Width (W)": "12.34"},
        {"identifier": "patient_b.ply", "Width (W)": ""},
    ]
    assert legend_rows == [{"name": "Width", "abbreviation": "W", "type": "linear", "unit": "mm", "geodesic": "no", "color": ""}]


def test_load_measurement_export_xlsx_accepts_an_in_memory_buffer():
    buffer = io.BytesIO()
    _write_measurement_export(buffer)
    buffer.seek(0)

    columns, rows, _legend = load_measurement_export_xlsx(buffer)
    assert columns == ["identifier", "Width (W)"]
    assert len(rows) == 2


def test_load_measurement_export_xlsx_rejects_a_file_with_no_measurements_sheet(tmp_path):
    path = tmp_path / "not_a_measurement_export.xlsx"
    wb = Workbook()
    wb.active.append(["some", "other", "sheet"])
    wb.save(path)

    import pytest

    with pytest.raises(ValueError, match="measurements"):
        load_measurement_export_xlsx(path)


def test_load_measurement_export_xlsx_tolerates_a_missing_legend_sheet(tmp_path):
    path = tmp_path / "no_legend.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "measurements"
    ws.append(["identifier", "Width (W)"])
    ws.append(["patient_a.ply", "1.0"])
    wb.save(path)

    columns, rows, legend_rows = load_measurement_export_xlsx(path)
    assert columns == ["identifier", "Width (W)"]
    assert legend_rows == []
