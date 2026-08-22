"""API test for POST /api/cohort/facial-measurements/load
(api/routers/cohort.py's load_facial_measurements) - the endpoint that
attaches a Facial Anthropometrics batch export (api/routers/facial.py's
export_batch) to an existing cohort as a lazily-joined dataset. Pure request/
response plumbing; the join logic itself is unit-tested directly in
tests/test_facial_cohort_link.py."""

from __future__ import annotations

from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook

from api.main import app
from api.results_bundle import _id_mapping_path

REPO_ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture
def client() -> TestClient:
    return TestClient(app)


def _mapping_row(cohort_id: str, file_path: str) -> dict:
    return {
        "cohort_id": cohort_id,
        "patient_id": "",
        "file_name": "",
        "file_path": file_path,
        "date_of_birth": "",
        "date_of_intervention": "",
    }


def _write_id_mapping(cohort_path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    header = ["cohort_id", "patient_id", "file_name", "file_path", "date_of_birth", "date_of_intervention"]
    ws.append(header)
    for row in rows:
        ws.append([row[key] for key in header])
    wb.save(_id_mapping_path(cohort_path))


def _write_measurement_export(path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "measurements"
    ws.append(["identifier", "Width (W)", "Angle (A)"])
    for row in rows:
        ws.append([row["identifier"], row["Width (W)"], row["Angle (A)"]])
    legend_ws = wb.create_sheet("legend")
    legend_ws.append(["name", "abbreviation", "type", "unit", "geodesic", "color"])
    legend_ws.append(["Width", "W", "linear", "mm", "no", "4ADE80"])
    legend_ws.append(["Angle", "A", "angular", "deg", "no", "F97316"])
    wb.save(path)


def test_load_facial_measurements_matches_by_filename(client: TestClient, tmp_path: Path):
    cohort_path = tmp_path / "cohort.xlsx"
    _write_id_mapping(
        cohort_path,
        [
            _mapping_row("C00001", "C:/scans/patient_a.ply"),
            _mapping_row("C00002", "C:/scans/patient_b.ply"),
        ],
    )
    measurement_path = tmp_path / "facial_measurements.xlsx"
    _write_measurement_export(
        measurement_path,
        [
            {"identifier": "patient_a.ply", "Width (W)": "12.34", "Angle (A)": "90.0"},
            {"identifier": "patient_b.ply", "Width (W)": "15.0", "Angle (A)": "88.5"},
            {"identifier": "patient_c.ply", "Width (W)": "9.9", "Angle (A)": "91.2"},
        ],
    )

    response = client.post(
        "/api/cohort/facial-measurements/load",
        json={"cohort_path": str(cohort_path), "measurement_file_path": str(measurement_path)},
    )
    assert response.status_code == 200, response.text
    body = response.json()

    assert body["columns"] == ["Width (W)", "Angle (A)"]
    assert body["rows_by_cohort_id"]["C00001"] == {"Width (W)": "12.34", "Angle (A)": "90.0"}
    assert body["rows_by_cohort_id"]["C00002"] == {"Width (W)": "15.0", "Angle (A)": "88.5"}
    assert "C00003" not in body["rows_by_cohort_id"]
    assert body["unmatched"] == ["patient_c.ply"]
    assert body["ambiguous"] == {}
    assert body["legend"][0]["name"] == "Width"


def test_load_facial_measurements_reports_ambiguous_filenames(client: TestClient, tmp_path: Path):
    cohort_path = tmp_path / "cohort.xlsx"
    _write_id_mapping(
        cohort_path,
        [
            _mapping_row("C00001", "C:/site_a/patient.ply"),
            _mapping_row("C00002", "C:/site_b/patient.ply"),
        ],
    )
    measurement_path = tmp_path / "facial_measurements.xlsx"
    _write_measurement_export(measurement_path, [{"identifier": "patient.ply", "Width (W)": "1.0", "Angle (A)": "2.0"}])

    response = client.post(
        "/api/cohort/facial-measurements/load",
        json={"cohort_path": str(cohort_path), "measurement_file_path": str(measurement_path)},
    )
    assert response.status_code == 200, response.text
    body = response.json()
    assert body["rows_by_cohort_id"] == {}
    assert body["ambiguous"] == {"patient.ply": ["C00001", "C00002"]}


def test_load_facial_measurements_404s_on_missing_file(client: TestClient, tmp_path: Path):
    cohort_path = tmp_path / "cohort.xlsx"
    response = client.post(
        "/api/cohort/facial-measurements/load",
        json={"cohort_path": str(cohort_path), "measurement_file_path": str(tmp_path / "nope.xlsx")},
    )
    assert response.status_code == 400


def test_load_facial_measurements_400s_on_non_measurement_export(client: TestClient, tmp_path: Path):
    cohort_path = tmp_path / "cohort.xlsx"
    bad_path = tmp_path / "not_an_export.xlsx"
    wb = Workbook()
    wb.active.append(["some", "other", "sheet"])
    wb.save(bad_path)

    response = client.post(
        "/api/cohort/facial-measurements/load",
        json={"cohort_path": str(cohort_path), "measurement_file_path": str(bad_path)},
    )
    assert response.status_code == 400
    assert "measurements" in response.json()["detail"]
