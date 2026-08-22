"""builds the results you get after an analysis - meshes, a json report, and
a measurement figure, packaged as CP_{stem}_{C|F}_{3|4}[_CoM]/ (see
results_folder_name). this is basically the same stuff the old app wrote
next to your file (_rg.ply, _metrics.json etc), just organized into one
folder now instead of a pile of sibling files.

two ways this gets delivered: zipped for a browser download
(build_results_bundle), or written straight to disk next to the original
mesh when we know a real filesystem path for it (write_results_to_folder,
desktop app only - see api/routers/mesh.py's /save endpoint). both go
through _build_report_files so the folder layout and filenames match
either way.
"""

from __future__ import annotations

import io
import json
import textwrap
import uuid
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable

import numpy as np
import trimesh
from matplotlib.backends.backend_agg import FigureCanvasAgg
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib.figure import Figure
from matplotlib.tri import Triangulation
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from shapely import concave_hull
from shapely.geometry import MultiPoint

from craniumpy_core.craniometrics import CranioMeasurements, FrontalBossingResult, hc_slice_polygon
from craniumpy_core.asymmetry import AsymmetryResult
from craniumpy_core.cohort import GroupMeasurements, SagittalMidlineBand, SpreadBand
from craniumpy_core.metopic import MetopicResult


def shorten_stem(stem: str) -> str:
    """collapses any embedded dot within an underscore-separated segment of
    the stem, keeping only what's before it - some scanners name files like
    "1016510_20210730.000112_edited", where that middle segment is a
    date-plus-subversion nobody wants spelled out in a results folder name.
    "1016510_20210730.000112_edited" -> "1016510_20210730_edited"."""
    return "_".join(part.split(".", 1)[0] for part in stem.split("_"))


def stem_from_filename(filename: str) -> str:
    """the shortened stem to build a results folder/file names from, given
    an original mesh filename (with its real extension still on it)."""
    stem = filename.rsplit(".", 1)[0] if "." in filename else filename
    return shorten_stem(stem)


def results_folder_name(original_filename: str, target: str, config: dict) -> str:
    """CP_{C|F}[4]_{stem}[_CoM] - what actually went into this run, not a
    generic "_results" suffix: landmark count (a "4" folded into the target
    letter when an alt_frontal_landmark was given, see
    pipeline.analyze_cranial) and whether center-of-mass correction ran. two
    runs with different settings on the same file land in different folders
    instead of one overwriting the other, and the name tells you which is
    which without opening report.json."""
    stem = stem_from_filename(original_filename)
    target_suffix = ("C" if target == "cranium" else "F") + ("4" if config.get("alt_frontal_landmark") else "")
    com_suffix = "_CoM" if config.get("com_translation") else ""
    return f"CP_{target_suffix}_{stem}{com_suffix}"


# figures are drawn twice: once into a standalone Figure that gets saved as
# a PNG (the exported _measurements.png etc.), and once directly into a
# region of a PDF page, where the result stays fully vector - no rasterized
# image is ever placed on a PDF page, so zooming the report never hits a
# resolution limit. that's why each builder below is split into a
# _draw_*(target, rect, ...) that draws into an arbitrary rectangle of an
# arbitrary Figure, plus a thin *_figure(...) wrapper that gives it a whole
# figure of its own and returns PNG bytes.
#
# rect is (left, bottom, width, height) in the target figure's own
# coordinates; _sub below maps a sub-rectangle expressed relative to it
# (0..1 within the rect) into those same figure coordinates, so the drawing
# code can lay itself out as if it owned the whole page.
FIGURE_PNG_DPI = 220


def _sub(rect: tuple[float, float, float, float], left: float, bottom: float, width: float, height: float):
    rl, rb, rw, rh = rect
    return (rl + left * rw, rb + bottom * rh, width * rw, height * rh)


def _rect_text(fig: Figure, rect, x: float, y: float, text: str, **kwargs) -> None:
    """fig.text, positioned relative to rect rather than the whole figure."""
    rl, rb, rw, rh = rect
    fig.text(rl + x * rw, rb + y * rh, text, **kwargs)


def _draw_measurements(
    fig: Figure, rect, mesh: trimesh.Trimesh, measurements: CranioMeasurements, spread_band: SpreadBand | None = None
) -> None:
    """top-down outline of the HC slice, red line style like the old app
    used to draw, with OFD/BPD spans marked on it too.

    spread_band (optional - a craniumpy_core.cohort.SpreadBand, see
    cohort.hc_ring_band) shades a +/-1 SD ring around the HC slice - how
    much the ring itself varies patient-to-patient, across the cohort
    group this figure's own mesh is the MEAN shape of. only meaningful
    (and only ever passed) for a cohort mean-shape report, where
    `measurements` itself describes the mean rather than a real patient -
    a real patient has no "spread" of their own to show."""
    polygon = hc_slice_polygon(mesh, measurements.slice_height)

    ax = fig.add_axes(_sub(rect, 0.12, 0.20, 0.82, 0.72))

    if spread_band is not None:
        outer_x, outer_z = spread_band.outer[:, 0], spread_band.outer[:, 2]
        inner_x, inner_z = spread_band.inner[:, 0], spread_band.inner[:, 2]
        ax.fill(
            np.concatenate([outer_x, inner_x[::-1]]), np.concatenate([outer_z, inner_z[::-1]]),
            color="#d1453d", alpha=0.5, linewidth=0, label="+/-1 SD across group",
        )
        # a real inter-patient radius SD is often only ~1-2mm against an
        # ~80mm mean radius - true to scale, that's sub-pixel at any normal
        # render resolution, so the fill alone (and a thin boundary stroke)
        # is indistinguishable from the solid HC line drawn right on top of
        # it a few lines down. a deliberately bold boundary stroke - well
        # past what the data's own mm-scale width would justify - is what
        # actually keeps a real but small band visible as a distinct
        # colored edge either side of that line, rather than pretending
        # the underlying spread is bigger than it is (the fill itself, and
        # the numbers in the report/UI, are still drawn to true scale).
        ax.plot(
            np.append(outer_x, outer_x[0]), np.append(outer_z, outer_z[0]), color="#d1453d", linewidth=2.2, alpha=0.8
        )
        ax.plot(
            np.append(inner_x, inner_x[0]), np.append(inner_z, inner_z[0]), color="#d1453d", linewidth=2.2, alpha=0.8
        )

    if polygon is not None and len(polygon) > 2:
        closed = np.vstack([polygon, polygon[0]])
        ax.plot(closed[:, 0], closed[:, 2], color="#d1453d", linewidth=2, label="HC slice")

    ax.plot(
        [measurements.lh_opt[0], measurements.rh_opt[0]],
        [measurements.lh_opt[2], measurements.rh_opt[2]],
        color="#2563eb", linewidth=1.5, linestyle="--", marker="o", markersize=4, label="BPD (breadth)",
    )
    ax.plot(
        [measurements.occ_opt[0], measurements.front_opt[0]],
        [measurements.occ_opt[2], measurements.front_opt[2]],
        color="#16a34a", linewidth=1.5, linestyle="--", marker="o", markersize=4, label="OFD (depth)",
    )

    ax.set_aspect("equal")
    ax.set_xlabel("x (mm)")
    ax.set_ylabel("z (mm)")
    ax.set_title(
        f"OFD {measurements.depth_mm}mm  BPD {measurements.breadth_mm}mm  "
        f"CI {measurements.cephalic_index}  HC {measurements.circumference_cm}cm"
    )
    ax.grid(alpha=0.2)

    handles, labels = ax.get_legend_handles_labels()
    rl, rb, rw, rh = rect
    fig.legend(
        handles, labels, loc="lower center",
        bbox_to_anchor=(rl + 0.5 * rw, rb + 0.05 * rh), bbox_transform=fig.transFigure,
        ncol=3, fontsize=8, frameon=False,
    )
    _rect_text(
        fig, rect, 0.5, 0.02,
        "OFD = occipitofrontal diameter (depth)   BPD = biparietal diameter (breadth)   "
        "CI = cephalic index   HC = head circumference",
        ha="center", va="bottom", fontsize=6, color="#666666",
    )


def _measurement_figure(mesh: trimesh.Trimesh, measurements: CranioMeasurements) -> bytes:
    fig = Figure(figsize=(6, 6.4), dpi=FIGURE_PNG_DPI)
    canvas = FigureCanvasAgg(fig)
    _draw_measurements(fig, (0, 0, 1, 1), mesh, measurements)
    buf = io.BytesIO()
    canvas.print_png(buf)
    return buf.getvalue()


def _silhouette_polygon(points_2d: np.ndarray, ratio: float = 0.3) -> np.ndarray:
    """the outer 2D silhouette of a point cloud - a concave hull, not the
    convex hull (a plain convex hull would erase real concave features like
    the notch behind an ear). used to draw a head/face outline that traces
    the shape's true projected boundary, unlike a mesh's own clip-boundary
    loop (the open edge left behind by clipping), which only incidentally
    resembles an outline and can cut across the interior depending on where
    the clip plane happens to sit. ratio trades tightness for smoothness -
    0.3 hugs real bumps without turning ordinary vertex-density jitter into
    a jagged line; falls back to the plain convex hull on the rare point set
    concave_hull can't return a single polygon for."""
    hull = concave_hull(MultiPoint(points_2d), ratio=ratio)
    if hull.geom_type != "Polygon":
        hull = hull.convex_hull
    return np.asarray(hull.exterior.coords)


# (horizontal_idx, vertical_idx, horizontal_label, vertical_label, show_silhouette)
# per _draw_asymmetry view - see that function's docstring for what each one
# actually shows.
_ASYMMETRY_VIEWS: dict[str, tuple[int, int, str, str, bool]] = {
    "frontal": (0, 1, "x (mm)", "y (mm)", True),
    "top": (0, 2, "x (mm)", "z (mm)", True),
    "sagittal": (2, 1, "z (mm, depth)", "y (mm, height)", False),
}


def _draw_asymmetry(
    fig: Figure, rect, mesh: trimesh.Trimesh, asymmetry: AsymmetryResult, *, label: str, view: str = "frontal"
) -> None:
    """the asymmetry heatmap, same blue(dent)/white/red(protruded) diverging
    scale as the live viewer, with an mm colorbar. label is the title prefix
    ("cranial"/"facial"); view picks which two axes stand in for the plot's
    horizontal/vertical and whether a silhouette outline gets drawn (see
    _ASYMMETRY_VIEWS):
      "frontal"  - x horizontal, y vertical (facial's usual view - see
                   registration.rigid, the face-target frame puts x as
                   left/right and y as up/down).
      "top"      - x horizontal, z vertical, looking straight down
                   (cranial's usual view - a frontal view would just show
                   the back of the scalp foreshortened into almost nothing).
      "sagittal" - z (depth) horizontal, y (height) vertical, side-on.
                   restricted to the half of the mesh that actually carries
                   the heatmap (see craniumpy_core.asymmetry's module
                   docstring - the other half is always zeroed): collapsing
                   left/right into a side view makes both halves' surfaces
                   overlap in the same (z, y) footprint, which without this
                   restriction renders as noise rather than a clean profile.
                   no silhouette either, for the same reason - a silhouette
                   of the WHOLE head/face wouldn't match a plot that only
                   ever draws one half of it. the x-axis is inverted below
                   (unlike _draw_frontal_bossing's own profile view, which
                   plots z left-to-right as-is) - this view only ever shows
                   the LEFT half's own surface (x >= 0 - see below), so
                   plotting z as-is would face the profile rightward, which
                   reads as the RIGHT side of the face; inverting it faces
                   the (correctly left) half leftward instead, matching
                   what it actually is.
    """
    vertices = np.asarray(mesh.vertices)
    faces = np.asarray(mesh.faces)
    heatmap = asymmetry.heatmap
    max_abs = max(float(np.abs(heatmap).max()), 1e-6)
    horizontal_idx, vertical_idx, horizontal_label, vertical_label, show_silhouette = _ASYMMETRY_VIEWS[view]

    if view == "sagittal":
        # heatmap is zeroed on x < 0 (see calculate_asymmetry) - keep only
        # faces entirely on the x >= 0 side so the triangulation doesn't
        # also carry the empty (all-zero, all-white) mirror half into the
        # same 2D footprint.
        data_half = vertices[:, 0] >= 0
        faces = faces[data_half[faces].all(axis=1)]

    ax = fig.add_axes(_sub(rect, 0.10, 0.09, 0.72, 0.84))
    cax = fig.add_axes(_sub(rect, 0.86, 0.14, 0.035, 0.74))

    triangulation = Triangulation(vertices[:, horizontal_idx], vertices[:, vertical_idx], faces)
    mesh_plot = ax.tripcolor(triangulation, heatmap, cmap="bwr", vmin=-max_abs, vmax=max_abs, shading="gouraud")

    if show_silhouette:
        silhouette = _silhouette_polygon(vertices[:, [horizontal_idx, vertical_idx]])
        ax.plot(silhouette[:, 0], silhouette[:, 1], color="#999999", linewidth=0.8, zorder=3)

    if view == "sagittal":
        ax.invert_xaxis()

    ax.set_aspect("equal")
    ax.set_xlabel(horizontal_label)
    ax.set_ylabel(vertical_label)
    title_suffix = " (sagittal)" if view == "sagittal" else ""
    ax.set_title(f"{label} asymmetry{title_suffix} - mean asymmetry index {asymmetry.mean_asymmetry_index:.2f}")
    fig.colorbar(mesh_plot, cax=cax, label="deviation from mirrored half (mm)")


def _asymmetry_figure(mesh: trimesh.Trimesh, asymmetry: AsymmetryResult, *, label: str, view: str = "frontal") -> bytes:
    fig = Figure(figsize=(6.6, 6), dpi=FIGURE_PNG_DPI)
    canvas = FigureCanvasAgg(fig)
    _draw_asymmetry(fig, (0, 0, 1, 1), mesh, asymmetry, label=label, view=view)
    buf = io.BytesIO()
    canvas.print_png(buf)
    return buf.getvalue()


def _draw_metopic(fig: Figure, rect, metopic: MetopicResult, spread_band: SpreadBand | None = None) -> None:
    """the forehead contour at the HC slice height, plus the fitted parabola,
    frontal-angle construction, central/temporal regions, and the phi(s)/
    kappa(s)/d_P(s) profiles - see craniumpy_core.metopic for what each of
    these actually is. main panel uses this module's own (x, z) convention,
    same as _measurement_figure above (x = left-right, z = depth).

    spread_band (optional - a craniumpy_core.cohort.SpreadBand, see
    cohort.metopic_band) shades a +/-1 SD band around the forehead contour
    - same "only meaningful for a cohort mean-shape report" reasoning as
    _draw_measurements' own spread_band param."""
    contour = metopic.contour
    x, z = contour[:, 0], contour[:, 1]
    u = metopic.normalized_arc_length
    s = metopic.arc_length
    M, L, R = metopic.frontal_angle_points

    # explicit rects instead of a gridspec: the same layout has to work
    # inside an arbitrary sub-rectangle of a PDF page as well as on a figure
    # of its own, and add_gridspec always divides the whole figure.
    # the 3 side panels are shorter (0.177 vs an even 3-way split's 0.21)
    # and further apart (0.095 gap vs 0.045) than an even split of the same
    # vertical span would give them - a plain even split left each panel's
    # title sitting on top of the tick labels of the panel above it, since
    # those are sized by real font metrics rather than this fixed fraction
    # (see tests/test_reporting.py's metopic overlap test, tuned against
    # this exact spacing).
    ax_main = fig.add_axes(_sub(rect, 0.06, 0.20, 0.50, 0.72))
    ax_phi = fig.add_axes(_sub(rect, 0.68, 0.744, 0.29, 0.177))
    ax_kappa = fig.add_axes(_sub(rect, 0.68, 0.472, 0.29, 0.177))
    ax_dev = fig.add_axes(_sub(rect, 0.68, 0.20, 0.29, 0.177))

    central = (u >= metopic.central_window[0]) & (u <= metopic.central_window[1])
    left_t = (u >= metopic.left_temporal_window[0]) & (u <= metopic.left_temporal_window[1])
    right_t = (u >= metopic.right_temporal_window[0]) & (u <= metopic.right_temporal_window[1])

    x_fit = np.linspace(x.min(), x.max(), 200)
    z_fit = metopic.parabola_a * x_fit**2 + metopic.parabola_c

    if spread_band is not None:
        ax_main.fill_between(
            spread_band.mean[:, 0], spread_band.inner[:, 2], spread_band.outer[:, 2],
            color="#3a3a3a", alpha=0.5, linewidth=0, label="+/-1 SD across group",
        )
        # see _draw_measurements' own comment on the matching HC-ring band -
        # a deliberately bold boundary stroke keeps a real but small band
        # visible either side of the forehead-contour line drawn over it
        # just below, in the same color, rather than the two becoming
        # indistinguishable.
        ax_main.plot(spread_band.mean[:, 0], spread_band.inner[:, 2], color="#3a3a3a", linewidth=2.2, alpha=0.8)
        ax_main.plot(spread_band.mean[:, 0], spread_band.outer[:, 2], color="#3a3a3a", linewidth=2.2, alpha=0.8)

    ax_main.plot(x, z, color="#3a3a3a", linewidth=1.5, label="forehead contour")
    ax_main.plot(x_fit, z_fit, color="#2563eb", linewidth=1.5, linestyle="--", label="ideal (parabola)")
    ax_main.axvline(0, color="#999999", linewidth=1, linestyle=":", label="midline")

    if central.any():
        ax_main.fill_between(x[central], z[central], metopic.parabola_a * x[central] ** 2 + metopic.parabola_c,
                              color="#d1453d", alpha=0.25, label="ridge area")
    for mask, color in ((left_t, "#0891b2"), (right_t, "#0891b2")):
        if mask.any():
            ax_main.fill_between(x[mask], z[mask], metopic.parabola_a * x[mask] ** 2 + metopic.parabola_c,
                                  color=color, alpha=0.2)

    ax_main.plot([L[0], M[0], R[0]], [L[1], M[1], R[1]], color="#16a34a", linewidth=1.2, marker="o", markersize=4,
                 label=f"frontal angle {metopic.frontal_angle_deg:.1f} deg")
    ax_main.scatter([M[0]], [M[1]], color="#d1453d", zorder=5, s=30, label=f"ridge protrusion {metopic.ridge_protrusion_mm:.1f}mm")

    for mask, side_label, depth in (
        (left_t, "L", metopic.left_max_temporal_depth_mm),
        (right_t, "R", metopic.right_max_temporal_depth_mm),
    ):
        if mask.any():
            idx = np.where(mask)[0][np.argmax(-metopic.deviation_profile[mask])]
            ax_main.scatter([x[idx]], [z[idx]], color="#0891b2", zorder=5, s=25, marker="v")
            ax_main.annotate(f"{side_label} depth {depth:.1f}mm", (x[idx], z[idx]), fontsize=7, color="#0891b2",
                              textcoords="offset points", xytext=(4, -8))

    ax_main.set_aspect("equal")
    ax_main.set_xlabel("x (mm)")
    ax_main.set_ylabel("z (mm)")
    ax_main.set_title(f"metopic/frontal shape - MCC {metopic.midline_curvature_concentration:.2f}  PDI {metopic.parabolic_deviation_index:.2f}mm")
    ax_main.grid(alpha=0.2)

    # the same fitted parabola drawn as "ideal parabola" in the main panel,
    # restated as phi/kappa/d_P so each small profile has something to be
    # read against instead of just its own raw shape - a patient line that
    # tracks the dashed one closely is close to a parabolic forehead at
    # that point along the contour; a patient line that pulls away from it
    # is exactly where the shape deviates, which is the plain-numbers-only
    # version of what these three panels previously showed. computed
    # directly from x rather than u/s: parametrizing the ideal parabola by
    # x (dx/dx=1, dz/dx=2ax, d2z/dx2=2a) gives a tangent ANGLE and
    # curvature that are already parametrization-invariant - same formulas
    # _gradient_and_curvature uses, just with the parabola's own exact
    # derivatives standing in for the numerically-differentiated contour
    # ones, and the same sign flip on curvature so it reads on the same
    # axis as curvature_profile.
    slope = 2 * metopic.parabola_a * x
    phi_ideal = np.arctan2(slope, 1.0)
    kappa_ideal = -2 * metopic.parabola_a / np.power(1.0 + slope**2, 1.5)

    # unlabeled - solid vs. dashed is self-evident within each small panel,
    # and a "patient" legend entry would push the shared legend below to a
    # third row, tall enough to collide with ax_dev's own xlabel (see the
    # PDF-page-scale overlap tests in tests/test_reporting.py this is
    # tuned against). "ideal (parabola)" alone, deduped against ax_main's
    # own identically-labeled line below, adds no extra row.
    ax_phi.plot(u, metopic.gradient_profile, color="#2563eb", linewidth=1.2)
    ax_phi.plot(u, phi_ideal, color="#2563eb", linewidth=1.2, linestyle="--", alpha=0.6, label="ideal (parabola)")
    ax_phi.set_ylabel("phi (rad)")
    ax_phi.set_title("gradient / tangent angle", fontsize=9)
    ax_phi.grid(alpha=0.2)

    ax_kappa.plot(u, metopic.curvature_profile, color="#16a34a", linewidth=1.2)
    ax_kappa.plot(u, kappa_ideal, color="#2563eb", linewidth=1.2, linestyle="--", alpha=0.6)
    ax_kappa.set_ylabel("kappa (1/mm)")
    ax_kappa.set_title("signed curvature", fontsize=9)
    ax_kappa.grid(alpha=0.2)

    ax_dev.plot(u, metopic.deviation_profile, color="#d1453d", linewidth=1.2)
    # the fitted parabola is what deviation_profile is measured against in
    # the first place, so its own "ideal" line is always exactly zero -
    # drawn the same dashed style as the other two panels' ideal line
    # rather than the plain gray it used to be, so all three panels read
    # the same way: solid = patient, dashed blue = ideal parabola.
    ax_dev.axhline(0, color="#2563eb", linewidth=1.2, linestyle="--", alpha=0.6)
    ax_dev.set_ylabel("d_P (mm)")
    ax_dev.set_xlabel("normalized arc length u")
    ax_dev.set_title("deviation from parabola", fontsize=9)
    ax_dev.grid(alpha=0.2)

    # legend below the plots rather than floating inside ax_main, where it
    # tends to sit on top of the contour/parabola themselves - handles/
    # labels are gathered from ax_main plus one profile panel ("patient"
    # and "ideal (parabola)" are the same two lines, styled identically,
    # on all three profile panels - only one copy is worth showing, so
    # ax_phi stands in for all three) rather than all three, which would
    # just repeat the same two entries three times over. deduped by label
    # (keeping the first occurrence) since ax_main's own "ideal (parabola)"
    # line - the dashed parabola drawn over the contour itself - uses the
    # exact same label as the profile panels' reference line, on purpose:
    # it's the same fit, just restated in three different ways.
    all_handles, all_labels = ax_main.get_legend_handles_labels()
    profile_handles, profile_labels = ax_phi.get_legend_handles_labels()
    all_handles, all_labels = all_handles + profile_handles, all_labels + profile_labels
    by_label = dict(zip(all_labels, all_handles))
    handles, labels = list(by_label.values()), list(by_label.keys())
    rl, rb, rw, rh = rect
    fig.legend(
        handles, labels, loc="lower center",
        bbox_to_anchor=(rl + 0.5 * rw, rb + 0.04 * rh), bbox_transform=fig.transFigure,
        ncol=3, fontsize=7, frameon=False,
    )
    _rect_text(
        fig, rect, 0.5, 0.015,
        "MCC = midline curvature concentration   PDI = parabolic deviation index   "
        "d_P = signed deviation from the fitted parabola   phi = tangent (gradient) angle   "
        "kappa = signed curvature   u = normalized arc length   L/R = left/right",
        ha="center", va="bottom", fontsize=6, color="#666666",
    )


def _metopic_figure(metopic: MetopicResult) -> bytes:
    fig = Figure(figsize=(11, 8), dpi=FIGURE_PNG_DPI)
    canvas = FigureCanvasAgg(fig)
    _draw_metopic(fig, (0, 0, 1, 1), metopic)
    buf = io.BytesIO()
    canvas.print_png(buf)
    return buf.getvalue()


def _draw_frontal_bossing(
    fig: Figure, rect, result: FrontalBossingResult, sagittal_band: SagittalMidlineBand | None = None
) -> None:
    """side-profile view of the sagittal contour through sellion (z=depth
    on the x-axis, y=height on the y-axis, matching how a side-view photo
    reads) with the angle construction that produced angle_deg - shared by
    cranial and facial exports alike, see
    craniumpy_core.craniometrics.frontal_bossing.

    the horizontal reference is drawn along result.horizontal rather than
    along +z: for a display frame reached via a secondary frontal landmark
    those are two different directions, and the angle was measured against
    the former (see craniumpy_core.pipeline.measure_cranial).

    sagittal_band (optional - a craniumpy_core.cohort.SagittalMidlineBand)
    adds a shaded +/-1 SD ribbon around the profile - how much this same
    sagittal depth varies patient-to-patient at each height, across the
    cohort group this figure's own mesh is the MEAN shape of. only
    meaningful (and only ever passed) for a cohort mean-shape report, where
    `result` itself describes the mean rather than a real patient - a real
    patient has no "spread" of their own to show."""
    profile = result.profile
    z, y = profile[:, 2], profile[:, 1]
    sellion, frontal = result.sellion, result.frontal_point

    ax = fig.add_axes(_sub(rect, 0.14, 0.19, 0.80, 0.74))

    if sagittal_band is not None:
        inner_z = sagittal_band.mean_z - sagittal_band.sd_z
        outer_z = sagittal_band.mean_z + sagittal_band.sd_z
        ax.fill_betweenx(
            sagittal_band.y, inner_z, outer_z,
            color="#3a3a3a", alpha=0.5, linewidth=0, label="+/-1 SD across group",
        )
        # see _draw_measurements' own comment on the matching HC-ring band -
        # a deliberately bold boundary stroke keeps a real but small band
        # visible either side of the sagittal-profile line drawn over it
        # just below, in the same color, rather than the two becoming
        # indistinguishable.
        ax.plot(inner_z, sagittal_band.y, color="#3a3a3a", linewidth=2.2, alpha=0.8)
        ax.plot(outer_z, sagittal_band.y, color="#3a3a3a", linewidth=2.2, alpha=0.8)

    ax.plot(z, y, color="#3a3a3a", linewidth=1.5, label="sagittal profile")
    ax.axhline(
        result.slice_height, color="#9ca3af", linewidth=1, linestyle=":",
        label=f"HC slice height ({result.slice_height:.1f} mm)",
    )
    horizontal = np.asarray(result.horizontal, dtype=np.float64)
    ref_length = 1.15 * float(np.linalg.norm(frontal - sellion))
    ref_end = sellion + horizontal * ref_length
    ax.plot([sellion[2], ref_end[2]], [sellion[1], ref_end[1]],
            color="#2563eb", linewidth=1.2, linestyle="--", label="horizontal")
    ax.plot([sellion[2], frontal[2]], [sellion[1], frontal[1]], color="#16a34a", linewidth=1.5, marker="o", markersize=4,
             label=f"bossing angle {result.angle_deg:.1f} deg")
    ax.scatter([sellion[2]], [sellion[1]], color="#d1453d", zorder=5, s=35, label="sellion")

    ax.set_aspect("equal")
    ax.set_xlabel("z (mm, depth)")
    ax.set_ylabel("y (mm, height)")
    ax.set_title(f"frontal bossing - {result.angle_deg:.1f} deg from horizontal")
    ax.grid(alpha=0.2)

    handles, labels = ax.get_legend_handles_labels()
    rl, rb, rw, rh = rect
    fig.legend(
        handles, labels, loc="lower center",
        bbox_to_anchor=(rl + 0.5 * rw, rb + 0.035 * rh), bbox_transform=fig.transFigure,
        ncol=3, fontsize=7, frameon=False,
    )
    _rect_text(
        fig, rect, 0.5, 0.015,
        "angle measured between horizontal (the sellion-tragus plane's depth axis through sellion) and the "
        "sellion -> forehead vector - smaller = more prominent/bossed forehead, larger = flatter or receding",
        ha="center", va="bottom", fontsize=6, color="#666666",
    )


def _frontal_bossing_figure(result: FrontalBossingResult) -> bytes:
    fig = Figure(figsize=(5, 6.5), dpi=FIGURE_PNG_DPI)
    canvas = FigureCanvasAgg(fig)
    _draw_frontal_bossing(fig, (0, 0, 1, 1), result)
    buf = io.BytesIO()
    canvas.print_png(buf)
    return buf.getvalue()


def _fmt(value: float | None) -> str:
    """formats a metric scalar for a CSV cell / PDF table - fixed 2 decimal
    places (not python's full float repr) and an empty string when the
    value doesn't apply to this session's target, so every row of a cohort
    CSV stays short and column-aligned regardless of which metrics ran."""
    return "" if value is None else f"{value:.2f}"


_METOPIC_ROW_KEYS = (
    "frontal_angle_deg",
    "midline_curvature_concentration",
    "midline_max_curvature",
    "midline_max_curvature_position",
    "ridge_protrusion_mm",
    "ridge_area_mm2",
    "ridge_area_normalized",
    "left_temporal_hollowing",
    "right_temporal_hollowing",
    "mean_temporal_hollowing",
    "left_max_temporal_depth_mm",
    "right_max_temporal_depth_mm",
    "parabolic_deviation_index",
)

# every column _metrics_row fills with _fmt(...) - a real measurement, so
# it's the one _write_xlsx_rows converts back to a numeric Excel cell
# (rather than text) when building a spreadsheet. metadata/settings columns
# are deliberately excluded, even the ones that happen to look numeric
# (a patient_id like "0042" must keep its leading zero, not become 42).
_NUMERIC_ROW_KEYS = (
    "depth_mm", "breadth_mm", "cephalic_index", "circumference_cm", "mesh_volume_cc",
    "cranial_asymmetry_index", "mean_asymmetry_index", "frontal_bossing_angle_deg",
) + tuple(f"metopic_{key}" for key in _METOPIC_ROW_KEYS)


def _nicp_template_name(nicp: dict | None) -> str:
    """the shipped template name, or just the filename (not the full local
    path - that's already in the JSON report's settings block verbatim for
    whoever needs it) of a custom template file. empty when nicp is None
    (no fit ran) or carries neither (shouldn't happen - see NicpConfig)."""
    if not nicp:
        return ""
    if nicp.get("template"):
        return str(nicp["template"])
    if nicp.get("custom_template_path"):
        return Path(str(nicp["custom_template_path"])).name
    return ""


def _nicp_mesh_path(results_dir: Path, original_filename: str, target: str, nicp_mesh: trimesh.Trimesh | None) -> str:
    """the absolute path of the NICP-fitted mesh this save actually wrote
    (or previously wrote, if this call skipped re-writing meshes - see
    write_analysis_to_folder's include_meshes) - "" if no fit ran, or if
    one did but the file genuinely isn't on disk (meshes were never saved
    at all, e.g. the "meshes" export checkbox was off on a session's very
    first export). filename pattern must match _build_mesh_files' own
    f"{stem}_rg_{target_suffix}N.ply" exactly, since that's what actually
    wrote it - checked against the real filesystem rather than assumed, so
    a cohort spreadsheet never ends up pointing at a file that isn't
    there."""
    if nicp_mesh is None:
        return ""
    stem = stem_from_filename(original_filename)
    target_suffix = "C" if target == "cranium" else "F"
    path = results_dir / "meshes" / f"{stem}_rg_{target_suffix}N.ply"
    return str(path.resolve()) if path.is_file() else ""


def _metrics_row(
    target: str,
    metadata: dict[str, str],
    config: dict,
    craniometrics: CranioMeasurements | None,
    asymmetry: AsymmetryResult | None,
    metopic: MetopicResult | None,
    frontal_bossing: FrontalBossingResult | None,
    nicp_mesh_path: str = "",
) -> dict[str, str]:
    """one flat {column: value} row - the shared source for the per-session
    summary spreadsheet, a cohort spreadsheet's accumulated rows (see
    _summary_xlsx/_upsert_cohort_xlsx), and the numbers table on the PDF
    report's pages. every metric that doesn't apply to this session (wrong
    target, a group that didn't compute) comes through as an empty string
    rather than being left out of the row entirely - a cohort spreadsheet
    needs the same columns on every row regardless of target to stay easy
    to stratify/filter later.

    config is whatever settings dict the JSON report's own "settings" block
    got (see _build_analysis_files) - com_correction/nicp_used/nicp_template
    below are pulled from the exact same source, so the spreadsheet and the
    JSON report can never disagree about what actually ran.

    nicp_mesh_path is the absolute path the NICP-fitted mesh (the third
    "_rg_{C|F}N.ply" file - see _build_mesh_files) actually got written to,
    when NICP ran and this row was built by one of the two writers that
    know a real dest_dir (write_analysis_to_folder/write_results_to_folder)
    - blank otherwise, same "still a column, just empty" convention as
    every other field here. this is the join key the cohort workspace uses
    to find same-template patients' meshes for a mean-shape computation -
    without it, a cohort spreadsheet has no way to locate the meshes behind
    its own rows at all."""
    nicp = config.get("nicp")
    row = {
        "file_name": metadata.get("file_name", ""),
        "file_path": metadata.get("file_path", ""),
        "patient_id": metadata.get("patient_id", ""),
        "date_of_birth": metadata.get("date_of_birth", ""),
        "diagnosis": metadata.get("diagnosis", ""),
        "sex": metadata.get("sex", ""),
        "date_imaging": metadata.get("date_imaging", ""),
        "age_imaging": metadata.get("age_imaging", ""),
        "image_timing": metadata.get("image_timing", ""),
        "surgical_status": metadata.get("surgical_status", ""),
        "treatment": metadata.get("treatment", ""),
        "date_of_intervention": metadata.get("date_of_intervention", ""),
        "age_intervention_months": metadata.get("age_intervention_months", ""),
        "free_variable": metadata.get("free_variable", ""),
        "target": target,
        "com_correction": "yes" if config.get("com_translation") else "no",
        "nicp_used": "yes" if nicp else "no",
        "nicp_template": _nicp_template_name(nicp),
        "nicp_mesh_path": nicp_mesh_path,
        "depth_mm": _fmt(craniometrics.depth_mm if craniometrics else None),
        "breadth_mm": _fmt(craniometrics.breadth_mm if craniometrics else None),
        "cephalic_index": _fmt(craniometrics.cephalic_index if craniometrics else None),
        "circumference_cm": _fmt(craniometrics.circumference_cm if craniometrics else None),
        "mesh_volume_cc": _fmt(float(craniometrics.mesh_volume_cc) if craniometrics else None),
        # separate columns rather than one shared "asymmetry index" - now
        # that both targets compute one (see _draw_asymmetry's view
        # param), a shared column would leave a mixed cohort file unable
        # to tell a cranial-cap asymmetry number from a facial one without
        # cross-referencing the target column for every row.
        "cranial_asymmetry_index": _fmt(asymmetry.mean_asymmetry_index if asymmetry and target == "cranium" else None),
        "mean_asymmetry_index": _fmt(asymmetry.mean_asymmetry_index if asymmetry and target == "face" else None),
        "frontal_bossing_angle_deg": _fmt(frontal_bossing.angle_deg if frontal_bossing else None),
    }
    for key in _METOPIC_ROW_KEYS:
        row[f"metopic_{key}"] = _fmt(getattr(metopic, key) if metopic else None)
    return row


def _write_xlsx_rows(rows: list[dict[str, str]], fieldnames: list[str]) -> bytes:
    """one sheet, formatted as an actual Excel table (banded rows, header
    dropdown filters, frozen header) rather than a plain range of appended
    cells - a spreadsheet meant to be opened and read, not just parsed back
    out. columns in _NUMERIC_ROW_KEYS get real numeric cells (2-decimal
    number format) instead of text, so Excel can sort/filter/average them
    directly instead of every value needing a manual "convert to number"
    pass first; every other column (patient/visit fields, settings,
    target) stays plain text, deliberately - a patient_id like "0042" has
    to keep its leading zero, which a numeric cell would silently drop.
    column widths are sized to content so nothing needs manual widening to
    read. this is also, incidentally, why the export is .xlsx and not
    .csv: a plain CSV opened by double-click lands as one unsplit column
    for any Excel whose locale uses semicolon (not comma) as its list
    separator - which is most non-US locales - defeating the entire point
    of a per-column export. openpyxl sidesteps that ambiguity outright."""
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    ws.append(fieldnames)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    display_values: list[list[str]] = []
    for row in rows:
        cells = []
        for name in fieldnames:
            value = row.get(name, "")
            if name in _NUMERIC_ROW_KEYS and value != "":
                cells.append(float(value))
            elif name in _NUMERIC_ROW_KEYS:
                cells.append(None)
            else:
                cells.append(value)
        ws.append(cells)
        display_values.append([str(c) if c is not None else "" for c in cells])

    for i, name in enumerate(fieldnames, start=1):
        if name in _NUMERIC_ROW_KEYS:
            for cell in ws[get_column_letter(i)][1:]:
                cell.number_format = "0.00"
        widest = max([len(name)] + [len(values[i - 1]) for values in display_values])
        ws.column_dimensions[get_column_letter(i)].width = min(max(widest + 2, 10), 40)

    if rows:
        last_col = get_column_letter(len(fieldnames))
        table = Table(displayName="summary_table", ref=f"A1:{last_col}{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False
        )
        ws.add_table(table)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _summary_xlsx(row: dict[str, str]) -> bytes:
    """header + this one row, as a real .xlsx - always generated (see
    _build_analysis_files), every column present even when its value is
    empty, so a batch of these can be combined later with nothing to
    reconcile."""
    return _write_xlsx_rows([row], list(row.keys()))


def _read_xlsx_rows(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    """(header, rows) for an existing .xlsx written by _write_xlsx_rows -
    every cell read back as a string (openpyxl hands back the underlying
    Python type, float for a numeric cell) since every row dict elsewhere
    in this module is string-keyed/valued and gets re-typed on its way
    back out anyway (see _write_xlsx_rows' own numeric-column handling).
    ([], []) if path doesn't exist - "nothing to merge with yet" rather
    than an error, since both callers below treat a missing file as the
    normal first-write case."""
    if not path.exists():
        return [], []
    wb = load_workbook(path)
    ws = wb.active
    header = [str(cell.value) if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        {header[i]: ("" if v is None else str(v)) for i, v in enumerate(excel_row) if i < len(header)}
        for excel_row in ws.iter_rows(min_row=2, values_only=True)
    ]
    return header, rows


def _row_key(row: dict[str, str]) -> str:
    """the identity a cohort/mapping row is matched and upserted on -
    file_path, falling back to file_name if blank (the closest thing to a
    stable per-mesh identity available without asking the user to type one
    in). shared by _upsert_cohort_xlsx and its id-mapping companion so a
    session's cohort row and mapping row are always keyed the same way."""
    return row.get("file_path") or row.get("file_name") or ""


def _upsert_rows(path: Path, row: dict[str, str], key_fn: Callable[[dict[str, str]], str] = _row_key) -> None:
    """creates path with a header + this row if it doesn't exist yet;
    otherwise replaces the row matching key_fn(row) or appends a new one.
    unions the existing file's columns with this row's, so an older file
    (fewer columns, from before some feature existed) doesn't lose data -
    its rows just stay blank in any new column, and this row stays blank
    in any column only the old file had.

    key_fn defaults to _row_key (file_path/file_name) - the ordinary case,
    and what the id-mapping file always uses (see _upsert_cohort_xlsx).
    the cohort file itself passes a different key_fn (matching on
    cohort_id instead), since file_path/file_name are deliberately never
    columns in that file at all - see _upsert_cohort_xlsx's own docstring
    for why."""
    key = key_fn(row)
    header, existing_rows = _read_xlsx_rows(path)

    fieldnames = list(row.keys())
    for name in header:
        if name not in fieldnames:
            fieldnames.append(name)

    replaced = False
    if key:
        for i, existing in enumerate(existing_rows):
            if key_fn(existing) == key:
                existing_rows[i] = row
                replaced = True
                break
    if not replaced:
        existing_rows.append(row)

    path.write_bytes(_write_xlsx_rows(existing_rows, fieldnames))


def _id_mapping_path(cohort_path: Path) -> Path:
    """the local-only companion file to a cohort spreadsheet - see
    _upsert_cohort_xlsx for why this is a separate file rather than a
    column in the cohort file itself."""
    return cohort_path.with_name(f"{cohort_path.stem}_id_mapping{cohort_path.suffix}")


def _next_cohort_id(existing_ids: list[str]) -> str:
    """C00001, C00002, ... - the next unused id for a cohort file, one more
    than the highest numeric suffix already in use (not just len(rows)+1,
    so a manually edited/pruned cohort file can't hand out a duplicate)."""
    highest = 0
    for value in existing_ids:
        if value.startswith("C") and value[1:].isdigit():
            highest = max(highest, int(value[1:]))
    return f"C{highest + 1:05d}"


def prepare_new_cohort_path(picked_path: str) -> str:
    """redirects a freshly-picked "create new cohort file" save location
    (desktop/app.py's pick_excel_file, save=True - defaults to a bare
    "cohort.xlsx" suggestion) into its own dedicated cohort_{uniqueID}/
    folder, with the cohort file itself carrying that same uniqueID in its
    own filename too (and so will its id-mapping companion, since
    _id_mapping_path just derives from this file's own stem) - so two
    cohorts both saved with the default suggested name, or just given the
    same name by two different users, never collide or silently overwrite
    each other, and each cohort's own folder is self-describing on disk
    rather than an indistinguishable bare "cohort.xlsx" sitting wherever it
    was saved. only meant to be called once, right when "create new cohort
    file..." is picked (see desktop/app.py's pick_excel_file) - the
    resulting path is what the frontend then holds onto (App.jsx's
    cohortPath) for every later row this session upserts into the same
    file, so the folder isn't reminted on every save."""
    original = Path(picked_path)
    unique_id = uuid.uuid4().hex[:8]
    folder = original.parent / f"cohort_{unique_id}"
    folder.mkdir(parents=True, exist_ok=True)
    return str(folder / f"cohort_{unique_id}.xlsx")


def list_cohort_patients(cohort_path: Path) -> list[dict[str, str]]:
    """the Per-patient sidebar's "load from cohort" dropdown (see
    PatientMetadataForm.jsx) - one entry per unique patient_id already in
    this cohort, reconstructed by joining the id-mapping file's own
    identity fields (patient_id/date_of_birth/date_of_intervention - see
    _id_mapping_path) back to the shared cohort file's own non-identifying
    ones (sex/diagnosis/treatment - still present there, only the fields in
    _COHORT_XLSX_EXCLUDED_COLUMNS were ever stripped out) via cohort_id. a
    patient with more than one row (a baseline plus follow-ups) appears
    once, using their MOST RECENT row - sex/diagnosis/date_of_birth
    shouldn't actually differ across a patient's own rows, but if one was
    ever corrected, the latest value is the one worth offering back. [] for
    a cohort with no mapping file yet (a freshly created one nothing's been
    saved into) - _read_xlsx_rows already treats a missing file as "nothing
    to read" rather than an error."""
    _, mapping_rows = _read_xlsx_rows(_id_mapping_path(cohort_path))
    _, cohort_rows = _read_xlsx_rows(cohort_path)
    cohort_by_id = {r.get("cohort_id", ""): r for r in cohort_rows}

    by_patient: dict[str, dict[str, str]] = {}
    for m in mapping_rows:
        patient_id = m.get("patient_id", "")
        if not patient_id:
            continue
        cohort_row = cohort_by_id.get(m.get("cohort_id", ""), {})
        by_patient[patient_id] = {
            "patient_id": patient_id,
            "date_of_birth": m.get("date_of_birth", ""),
            "date_of_intervention": m.get("date_of_intervention", ""),
            "sex": cohort_row.get("sex", ""),
            "diagnosis": cohort_row.get("diagnosis", ""),
            "treatment": cohort_row.get("treatment", ""),
        }
    return sorted(by_patient.values(), key=lambda p: p["patient_id"])


# columns that never appear in the shared cohort file itself - anything
# that identifies the patient (directly, like patient_id/date_of_birth, or
# indirectly, like a file_name/file_path that might embed a name or MRN)
# stays only in the id-mapping companion file (see _id_mapping_path).
# age_imaging/age_intervention_months are deliberately NOT in this set -
# they're derived FROM date_of_birth, but a plain age-in-months number
# isn't itself patient-identifying the way the underlying dates are, and
# the whole point of a shared cohort file is to carry exactly that kind of
# de-identified derived measurement.
_COHORT_XLSX_EXCLUDED_COLUMNS = {"patient_id", "file_name", "file_path", "date_of_birth", "date_of_intervention"}


def _upsert_cohort_xlsx(cohort_path: Path, row: dict[str, str]) -> None:
    """upserts this session's row into a shared, cross-center-ready cohort
    spreadsheet - a second export of the same source file overwrites its
    existing row instead of duplicating it, exactly like _upsert_rows in
    general, but this one additionally assigns every distinct file a
    cohort_id: a short, sequential, center-issued identifier that can
    travel with the row when this file is shared with another center, in
    place of every patient-identifying field this function strips out
    before writing (see _COHORT_XLSX_EXCLUDED_COLUMNS) - patient_id (a
    locally meaningful identifier, e.g. a hospital MRN or local study
    number), file_name/file_path (which can embed a patient name or MRN in
    the filename itself), and date_of_birth/date_of_intervention (real
    calendar dates, identifying in combination with the rest of a row even
    though the derived age-in-months fields that stay behind aren't). the
    id is stable across re-exports of the same file (reused, not
    reassigned, when a row is overwritten), and recorded - alongside every
    field just stripped out - in a SEPARATE file next to this one (see
    _id_mapping_path). that companion file, never this one, is what should
    be treated as sensitive and kept off anything shared onward; matching
    "is this the same file as before" (so re-exporting reuses the same
    cohort_id) has to key off THAT file's own rows now, since the fields
    _row_key needs for that are exactly the ones this file no longer
    carries."""
    mapping_path = _id_mapping_path(cohort_path)
    _, existing_mapping_rows = _read_xlsx_rows(mapping_path)
    _, existing_cohort_rows = _read_xlsx_rows(cohort_path)
    key = _row_key(row)
    existing_id = next((r.get("cohort_id", "") for r in existing_mapping_rows if _row_key(r) == key), "")
    # minted from whichever of the two files' own ids goes higher, not just
    # the mapping file's - a cohort file that ever ends up without a
    # matching (or fully in-sync) mapping file next to it (hand-edited,
    # migrated from an older version, mapping file deleted...) still can't
    # collide with an id already sitting in the cohort file itself.
    all_existing_ids = [r.get("cohort_id", "") for r in existing_mapping_rows] + [
        r.get("cohort_id", "") for r in existing_cohort_rows
    ]
    cohort_id = existing_id or _next_cohort_id(all_existing_ids)

    cohort_row = {k: v for k, v in row.items() if k not in _COHORT_XLSX_EXCLUDED_COLUMNS}
    cohort_row = {"cohort_id": cohort_id, **cohort_row}
    _upsert_rows(cohort_path, cohort_row, key_fn=lambda r: r.get("cohort_id", ""))

    mapping_row = {
        "cohort_id": cohort_id,
        "patient_id": row.get("patient_id", ""),
        "file_name": row.get("file_name", ""),
        "file_path": row.get("file_path", ""),
        "date_of_birth": row.get("date_of_birth", ""),
        "date_of_intervention": row.get("date_of_intervention", ""),
    }
    _upsert_rows(mapping_path, mapping_row)


# short, parent-facing one-liners for the metrics actually shown on the PDF
# report - keyed the same as _metrics_row's columns. deliberately a curated
# subset of the full CSV row (skips near-duplicate/intermediate-math fields
# like mean_temporal_hollowing, ridge_area_normalized, midline_max_curvature
# and its position) - the CSV is for research/batch use and keeps everything,
# the PDF is for a parent conversation and only needs what's actually
# discussable.
METRIC_EXPLAINERS: dict[str, str] = {
    "depth_mm": "Head length, measured front to back.",
    "breadth_mm": "Head width, measured side to side.",
    "cephalic_index": "Ratio of width to length - describes the overall head shape.",
    "circumference_cm": "Head circumference, measured around the widest part.",
    "mesh_volume_cc": "Estimated mesh volume.",
    "mean_asymmetry_index": "Average left-right difference across the face. Lower means more symmetric.",
    "cranial_asymmetry_index": "Average left-right difference across the head. Lower means more symmetric.",
    "frontal_bossing_angle_deg": (
        "How much the forehead curves forward. A smaller angle means a more prominent forehead; "
        "a larger angle means flatter or more receding."
    ),
    "metopic_frontal_angle_deg": (
        "The angle at the forehead's central ridge point, between two points 30% of this forehead's own "
        "contour length to either side of the midline - a sharper (smaller) angle can suggest a more "
        "pointed/ridged forehead shape."
    ),
    "metopic_ridge_protrusion_mm": (
        "How far the center sticks out (positive) or falls short (negative) of the ideal parabola - a "
        "self-referential baseline fit to this same forehead's own sides, not a healthy-head template."
    ),
    "metopic_ridge_area_mm2": (
        "Net area between the contour and the ideal parabola across the central window: positive means the "
        "center sticks out on net, negative means it falls short on net (flatter/recessed relative to what "
        "the sides predict). Near zero doesn't mean 'normal' - a real forehead was never exactly a parabola."
    ),
    "metopic_left_temporal_hollowing": "How sunken in the left temple is compared to the surrounding forehead shape.",
    "metopic_right_temporal_hollowing": "How sunken in the right temple is compared to the surrounding forehead shape.",
    "metopic_midline_curvature_concentration": (
        "How tightly the curve of the forehead is focused around the centerline, versus spread out evenly."
    ),
    "metopic_parabolic_deviation_index": (
        "An overall score for how much the forehead's shape differs from a smooth, evenly curved reference shape."
    ),
}

# (row_key, display label, unit) per metric group, in display order - the
# curated subset METRIC_EXPLAINERS has text for. group keys match the PNG
# filenames _build_analysis_files produces (see figure_names in _report_pdf).
# the two asymmetry groups (and their sagittal companions, which don't add
# any new metric - same numbers as their primary group, just a second
# figure) are last on purpose, so the report always ends on the asymmetry
# section regardless of target.
_PDF_METRIC_FIELDS: dict[str, list[tuple[str, str, str]]] = {
    "craniometrics": [
        ("depth_mm", "Head length (OFD)", "mm"),
        ("breadth_mm", "Head width (BPD)", "mm"),
        ("cephalic_index", "Cephalic index", ""),
        ("circumference_cm", "Head circumference", "cm"),
        ("mesh_volume_cc", "Estimated mesh volume", "cc"),
    ],
    "frontal_bossing": [
        ("frontal_bossing_angle_deg", "Frontal bossing angle", "deg"),
    ],
    "metopic": [
        ("metopic_frontal_angle_deg", "Frontal angle", "deg"),
        ("metopic_ridge_protrusion_mm", "Ridge protrusion", "mm"),
        ("metopic_ridge_area_mm2", "Ridge area", "mm2"),
        ("metopic_left_temporal_hollowing", "Left temple hollowing", ""),
        ("metopic_right_temporal_hollowing", "Right temple hollowing", ""),
        ("metopic_midline_curvature_concentration", "Midline curvature concentration", ""),
        ("metopic_parabolic_deviation_index", "Overall shape deviation index", "mm"),
    ],
    "cranial_asymmetry": [
        ("cranial_asymmetry_index", "Cranial asymmetry index", "mm"),
    ],
    "cranial_asymmetry_sagittal": [],
    "asymmetry": [
        ("mean_asymmetry_index", "Facial asymmetry index", "mm"),
    ],
    "asymmetry_sagittal": [],
}


# A4 portrait, in inches - the page every PDF sheet below is laid out on.
PAGE_W_IN, PAGE_H_IN = 8.27, 11.69
PAGE_H_PT = PAGE_H_IN * 72.0

# line advance as a multiple of font size, and the extra gap between one
# metric's block and the next, in points. every line of text on a PDF page
# advances by exactly fontsize * LINE_SPACING points regardless of what the
# line actually contains - a fixed distance, so no two lines can ever land
# on top of each other. this replaces an earlier measure-the-rendered-bbox
# approach: measuring is only as good as the renderer agreeing with the
# final PDF backend about text extents, and when it disagreed the lines
# overlapped. a fixed advance is verifiable by arithmetic instead.
#
# every string passed to _draw_line is pre-wrapped to a single visual line
# (see _wrap), which is what makes a per-line constant sufficient - a
# multi-line string would need its own line count factored in, and there
# deliberately aren't any.
LINE_SPACING = 1.7
BLOCK_GAP_PT = 9.0
# left margin and the y of the first body line, in figure fractions
TEXT_X = 0.10
LABEL_VALUE_X = 0.46

# guaranteed blank space at the physical bottom of a metric-fields page - a
# fixed 8pt explainer size (see _draw_metric_fields) can genuinely run a
# long enough group (metopic's own 7 fields, especially with a spread-band
# caption tacked on - see mean_shape_report_pdf) right off the page with
# nothing to stop it otherwise, since text_y itself never gets checked
# against the page's own physical bottom edge.
TEXT_BOTTOM_MARGIN_PT = 20.0


def _pt_to_frac(points: float) -> float:
    """points -> figure-fraction of an A4 portrait page's height."""
    return points / PAGE_H_PT


def _draw_line(page: Figure, y_top: float, text: str, fontsize: float = 10, x: float = TEXT_X, **kwargs) -> float:
    """draws one pre-wrapped line with its top edge at y_top (figure
    fraction, measured top-down) and returns the y_top for the next line,
    advanced by a fixed fontsize * LINE_SPACING. no renderer needed, and no
    dependence on font metrics: the spacing is the same whether the line is
    empty or full, bold or regular."""
    page.text(x, y_top, text, va="top", fontsize=fontsize, **kwargs)
    return y_top - _pt_to_frac(fontsize * LINE_SPACING)


def _wrap(text: str, width: int) -> list[str]:
    """textwrap, but never returns an empty list for empty input - callers
    iterate the result and an empty string still needs to occupy its line."""
    return textwrap.wrap(text, width=width) or [""]


def _draw_metric_fields(
    page: Figure,
    text_y: float,
    fields: list[tuple[str, str, str]],
    row: dict[str, str],
    extra_captions: list[str] = (),
    wrap_width: int = 105,
) -> float:
    """one metric group's text block: a bold label:value line per field
    that actually has a value, each followed by its wrapped one-line-per-
    sentence-ish explainer (see METRIC_EXPLAINERS), then any extra plain
    caption lines drawn after all of them (the spread-band notes
    mean_shape_report_pdf adds - e.g. "Shaded band on the figure above:
    ..."). returns the y position after the last line, same convention as
    _draw_line, for a caller that wants to keep drawing below it (none do
    right now, but nothing here assumes otherwise).

    the explainer/caption font size shrinks (never grows past the normal
    8pt) just enough to keep the WHOLE block above TEXT_BOTTOM_MARGIN_PT
    from the page's own physical bottom edge - a fixed size regardless of
    how much text a group has is what let a long enough group (metopic's
    own 7 fields, worse with a spread-band caption on top) run text right
    off the page, since nothing ever checked text_y against where the page
    actually ends. only the explainer text shrinks, never the bold label
    line - that's the actual number/name being reported, more important to
    keep at a normal, easily readable size than the explanation under it.
    line COUNTS (from _wrap) don't depend on font size (a fixed character
    width, not a fixed physical width), so the fit-to-budget math below is
    exact, not an approximation that then needs re-checking after the fact."""
    present = [(key, label, unit) for key, label, unit in fields if row.get(key, "")]
    explainer_lines = [_wrap(METRIC_EXPLAINERS.get(key, ""), width=wrap_width) for key, _label, _unit in present]
    caption_lines = [_wrap(c, width=wrap_width) for c in extra_captions]

    label_and_gap_total_pt = len(present) * (10 * LINE_SPACING + BLOCK_GAP_PT)
    explainer_line_count = sum(len(lines) for lines in explainer_lines) + sum(len(lines) for lines in caption_lines)
    explainer_fontsize = 8.0
    if explainer_line_count > 0:
        budget_pt = text_y * PAGE_H_PT - TEXT_BOTTOM_MARGIN_PT
        max_fontsize = (budget_pt - label_and_gap_total_pt) / (explainer_line_count * LINE_SPACING)
        explainer_fontsize = max(6.0, min(8.0, max_fontsize))

    for (key, label, unit), lines in zip(present, explainer_lines):
        display = f"{row[key]} {unit}".strip()
        text_y = _draw_line(page, text_y, f"{label}: {display}", fontsize=10, weight="bold")
        for line in lines:
            text_y = _draw_line(page, text_y, line, fontsize=explainer_fontsize, color="#555555")
        text_y -= _pt_to_frac(BLOCK_GAP_PT)

    for lines in caption_lines:
        for line in lines:
            text_y = _draw_line(page, text_y, line, fontsize=explainer_fontsize, color="#555555")

    return text_y


def _report_pdf(
    original_filename: str,
    target: str,
    metadata: dict[str, str],
    row: dict[str, str],
    draw_figure: dict[str, Callable[[Figure, tuple[float, float, float, float]], None]],
) -> bytes:
    """multi-page PDF, plain-language enough to hand to parents during a
    visit: a title/patient-visit page, then one page per metric group that
    actually ran for this session, each showing that group's figure with a
    short table of its numbers and a one-line explainer per metric (see
    METRIC_EXPLAINERS).

    draw_figure maps a metric-group key (the keys of _PDF_METRIC_FIELDS) to
    a callable that draws that group's figure into a given rectangle of a
    given page - the same _draw_* functions the exported PNGs are built
    from. drawing them straight onto the page instead of embedding the
    already-rendered PNG is what keeps the whole report vector: text and
    lines stay sharp at any zoom, where a placed raster image visibly
    breaks down (an A4 page can only hold so many pixels, and a figure
    scaled into half of one had far fewer than a reader zooming in
    expects).

    a group missing from draw_figure just doesn't get a page, which is how
    target-specific groups (craniometrics for cranial, asymmetry/metopic
    for facial) are skipped."""
    buf = io.BytesIO()

    with PdfPages(buf) as pdf:
        title_page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))  # A4 portrait
        FigureCanvasAgg(title_page)
        title_page.text(0.5, 0.92, "CraniumPy v2.0 Analysis Report", ha="center", fontsize=20, weight="bold")
        title_page.text(
            0.5, 0.885,
            f"{'Cranial' if target == 'cranium' else 'Facial'} analysis - {original_filename}",
            ha="center", fontsize=11, color="#666666",
        )

        info_lines = [
            ("File", metadata.get("file_name", "") or original_filename),
            ("Patient ID", metadata.get("patient_id", "")),
            ("Date of birth", metadata.get("date_of_birth", "")),
            ("Diagnosis", metadata.get("diagnosis", "")),
            ("Imaging date", metadata.get("date_imaging", "")),
            ("Age at imaging (months)", metadata.get("age_imaging", "")),
            ("Image timing", metadata.get("image_timing", "")),
            ("Surgical status", metadata.get("surgical_status", "")),
            ("Sex", metadata.get("sex", "")),
            ("Treatment", metadata.get("treatment", "")),
            ("Date of intervention", metadata.get("date_of_intervention", "")),
            ("Age at intervention (months)", metadata.get("age_intervention_months", "")),
            ("Notes", metadata.get("free_variable", "")),
        ]
        y = 0.80
        for label, value in info_lines:
            # label and value share one line, so the label is drawn without
            # advancing and the value's own _draw_line carries the advance
            # for both - keeping every row exactly one fixed step apart.
            title_page.text(TEXT_X, y, f"{label}:", va="top", fontsize=10, weight="bold")
            wrapped = _wrap(value or "-", width=54)
            for line in wrapped:
                y = _draw_line(title_page, y, line, fontsize=10, x=LABEL_VALUE_X)
        y -= _pt_to_frac(BLOCK_GAP_PT)
        _draw_line(
            title_page, y, f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d')}",
            fontsize=8, color="#999999",
        )
        pdf.savefig(title_page)

        for group, fields in _PDF_METRIC_FIELDS.items():
            draw = draw_figure.get(group)
            if draw is None:
                continue

            page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
            FigureCanvasAgg(page)
            draw(page, (0.04, 0.46, 0.92, 0.50))

            _draw_metric_fields(page, 0.42, fields, row)
            pdf.savefig(page)

    return buf.getvalue()


def mean_shape_report_pdf(
    mesh: trimesh.Trimesh,
    target: str,
    group_label: str,
    source_count: int,
    measurements: GroupMeasurements,
    sagittal_band: SagittalMidlineBand | None = None,
    hc_ring_band: SpreadBand | None = None,
    metopic_band: SpreadBand | None = None,
) -> bytes:
    """same multi-page layout as _report_pdf (a title page, then one page
    per metric group that actually applies) - built for a cohort MEAN
    shape (see craniumpy_core.cohort.measure_mean_shape) instead of one
    patient. measurements carries the exact same CranioMeasurements/
    AsymmetryResult/MetopicResult/FrontalBossingResult dataclasses a real
    patient's own report is drawn from, so every _draw_* figure function
    below applies to a mean shape completely unchanged - the only real
    difference from _report_pdf is the title page (a group description
    instead of one patient's metadata) and the three optional spread bands.

    group_label is free text describing what this group actually is (e.g.
    "trigonocephaly, pre-op, surgical" - the same kind of string the
    frontend's naming.js builds for the mesh-download filename), since
    there's no single patient/file name to put in its place.

    sagittal_band/hc_ring_band/metopic_band (optional - see
    craniumpy_core.cohort's sagittal_midline_band/hc_ring_band/
    metopic_band), when given, each add a +/-1 SD shaded band to their own
    page - the frontal_bossing, craniometrics, and metopic pages
    respectively. every number in this report still describes the group's
    MEAN shape; these bands are the only real patient-to-patient spread
    shown, each measured by re-running the relevant piece of the single-
    patient measurement pipeline (a slice search, a forehead contour...)
    against every individual mesh in the group, not derived from the mean
    shape's own surface (which has no spread of its own left to show)."""
    row = _metrics_row(
        target, {}, {"com_translation": True, "nicp": None},
        measurements.craniometrics, measurements.asymmetry, measurements.metopic, measurements.frontal_bossing,
    )

    is_cranial_asymmetry = measurements.asymmetry is not None and target == "cranium"
    asymmetry_view = "top" if is_cranial_asymmetry else "frontal"
    asymmetry_label = "cranial" if is_cranial_asymmetry else "facial"
    asymmetry_group = "cranial_asymmetry" if is_cranial_asymmetry else "asymmetry"
    asymmetry_sagittal_group = "cranial_asymmetry_sagittal" if is_cranial_asymmetry else "asymmetry_sagittal"

    draw_figure: dict[str, Callable[[Figure, tuple[float, float, float, float]], None]] = {}
    if measurements.craniometrics is not None:
        draw_figure["craniometrics"] = lambda fig, rect: _draw_measurements(
            fig, rect, mesh, measurements.craniometrics, spread_band=hc_ring_band
        )
    if measurements.asymmetry is not None:
        draw_figure[asymmetry_group] = lambda fig, rect: _draw_asymmetry(
            fig, rect, mesh, measurements.asymmetry, label=asymmetry_label, view=asymmetry_view
        )
        draw_figure[asymmetry_sagittal_group] = lambda fig, rect: _draw_asymmetry(
            fig, rect, mesh, measurements.asymmetry, label=asymmetry_label, view="sagittal"
        )
    if measurements.metopic is not None:
        draw_figure["metopic"] = lambda fig, rect: _draw_metopic(fig, rect, measurements.metopic, spread_band=metopic_band)
    if measurements.frontal_bossing is not None:
        draw_figure["frontal_bossing"] = lambda fig, rect: _draw_frontal_bossing(
            fig, rect, measurements.frontal_bossing, sagittal_band=sagittal_band
        )

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        title_page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
        FigureCanvasAgg(title_page)
        title_page.text(0.5, 0.92, "CraniumPy v2.0 Cohort Mean Shape Report", ha="center", fontsize=20, weight="bold")
        title_page.text(
            0.5, 0.885,
            f"{'Cranial' if target == 'cranium' else 'Facial'} mean shape - {group_label}",
            ha="center", fontsize=11, color="#666666",
        )

        info_lines = [
            ("Group", group_label),
            ("Target", "Cranium" if target == "cranium" else "Face"),
            ("Patients averaged", str(source_count)),
        ]
        y = 0.80
        for label, value in info_lines:
            title_page.text(TEXT_X, y, f"{label}:", va="top", fontsize=10, weight="bold")
            for line in _wrap(value or "-", width=54):
                y = _draw_line(title_page, y, line, fontsize=10, x=LABEL_VALUE_X)
        y -= _pt_to_frac(BLOCK_GAP_PT)
        y = _draw_line(
            title_page, y,
            "Every number in this report describes the AVERAGE shape of this group, not any single real patient.",
            fontsize=9, color="#999999",
        )
        _draw_line(
            title_page, y, f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", fontsize=8, color="#999999",
        )
        pdf.savefig(title_page)

        for group, fields in _PDF_METRIC_FIELDS.items():
            draw = draw_figure.get(group)
            if draw is None:
                continue

            page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
            FigureCanvasAgg(page)
            draw(page, (0.04, 0.46, 0.92, 0.50))

            extra_captions = []
            if group == "frontal_bossing" and sagittal_band is not None:
                extra_captions.append("Shaded band on the figure above: +/-1 SD sagittal depth across the group.")
            if group == "craniometrics" and hc_ring_band is not None:
                extra_captions.append("Shaded ring on the figure above: +/-1 SD head-circumference radius across the group.")
            if group == "metopic" and metopic_band is not None:
                extra_captions.append("Shaded band on the figure above: +/-1 SD forehead depth across the group.")

            _draw_metric_fields(page, 0.42, fields, row, extra_captions=extra_captions)
            pdf.savefig(page)

    return buf.getvalue()


def longitudinal_comparison_report_pdf(
    mesh_a: trimesh.Trimesh,
    mesh_b: trimesh.Trimesh,
    target: str,
    label_a: str,
    label_b: str,
    measurements_a: GroupMeasurements,
    measurements_b: GroupMeasurements,
    diff_heatmap: np.ndarray | None = None,
) -> bytes:
    """two-timepoint comparison PDF for the Longitudinal/Follow-up
    workspace - a title page, then Timepoint A's own full report section
    (one page per metric group that applies, exactly _report_pdf's own
    per-group page layout), then Timepoint B's section the same way, then
    (when diff_heatmap is given) a final page showing the per-vertex
    change from mesh_a to mesh_b as a heatmap - same diverging blue/red
    scale an asymmetry heatmap uses, reusing _draw_asymmetry unchanged via
    a synthetic AsymmetryResult wrapping diff_heatmap (see
    craniumpy_core.cohort.reference_diff for what diff_heatmap actually is
    - not a real asymmetry measurement, just the same "signed mm
    displacement, diverging colormap" shape).

    each section reuses the exact same full-page-width rect
    mean_shape_report_pdf's own pages already use, rather than splitting
    the page in half for both timepoints at once - every _draw_* figure
    function's title/legend/colorbar text is sized in fixed points, not
    scaled to the rect it's given, so halving the page width just runs
    that text into the other half instead of shrinking to fit. two
    full-width sections avoids that outright, at the cost of a longer PDF
    instead of a literally side-by-side one."""
    row_a = _metrics_row(
        target, {}, {"com_translation": True, "nicp": None},
        measurements_a.craniometrics, measurements_a.asymmetry, measurements_a.metopic, measurements_a.frontal_bossing,
    )
    row_b = _metrics_row(
        target, {}, {"com_translation": True, "nicp": None},
        measurements_b.craniometrics, measurements_b.asymmetry, measurements_b.metopic, measurements_b.frontal_bossing,
    )

    is_cranial = target == "cranium"
    asymmetry_view = "top" if is_cranial else "frontal"
    asymmetry_label = "cranial" if is_cranial else "facial"
    asymmetry_group = "cranial_asymmetry" if is_cranial else "asymmetry"
    asymmetry_sagittal_group = "cranial_asymmetry_sagittal" if is_cranial else "asymmetry_sagittal"

    def draw_figure_for(mesh: trimesh.Trimesh, measurements: GroupMeasurements) -> dict[str, Callable]:
        draw_figure: dict[str, Callable[[Figure, tuple[float, float, float, float]], None]] = {}
        if measurements.craniometrics is not None:
            draw_figure["craniometrics"] = lambda fig, rect: _draw_measurements(fig, rect, mesh, measurements.craniometrics)
        if measurements.asymmetry is not None:
            draw_figure[asymmetry_group] = lambda fig, rect: _draw_asymmetry(
                fig, rect, mesh, measurements.asymmetry, label=asymmetry_label, view=asymmetry_view
            )
            draw_figure[asymmetry_sagittal_group] = lambda fig, rect: _draw_asymmetry(
                fig, rect, mesh, measurements.asymmetry, label=asymmetry_label, view="sagittal"
            )
        if measurements.metopic is not None:
            draw_figure["metopic"] = lambda fig, rect: _draw_metopic(fig, rect, measurements.metopic)
        if measurements.frontal_bossing is not None:
            draw_figure["frontal_bossing"] = lambda fig, rect: _draw_frontal_bossing(fig, rect, measurements.frontal_bossing)
        return draw_figure

    sections = [
        (label_a, "#2563eb", row_a, draw_figure_for(mesh_a, measurements_a)),
        (label_b, "#d1453d", row_b, draw_figure_for(mesh_b, measurements_b)),
    ]

    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        title_page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
        FigureCanvasAgg(title_page)
        title_page.text(0.5, 0.92, "CraniumPy v2.0 Longitudinal Comparison Report", ha="center", fontsize=20, weight="bold")
        title_page.text(
            0.5, 0.885, f"{'Cranial' if target == 'cranium' else 'Facial'} comparison", ha="center", fontsize=11, color="#666666",
        )

        info_lines = [
            ("Timepoint A", label_a),
            ("Timepoint B", label_b),
            ("Target", "Cranium" if target == "cranium" else "Face"),
        ]
        y = 0.80
        for label, value in info_lines:
            title_page.text(TEXT_X, y, f"{label}:", va="top", fontsize=10, weight="bold")
            for line in _wrap(value or "-", width=54):
                y = _draw_line(title_page, y, line, fontsize=10, x=LABEL_VALUE_X)
        y -= _pt_to_frac(BLOCK_GAP_PT)
        y = _draw_line(
            title_page, y,
            "Each timepoint's own analysis follows as its own section, in the same order as a single-patient report.",
            fontsize=9, color="#999999",
        )
        if diff_heatmap is not None:
            y = _draw_line(
                title_page, y,
                "Includes a per-vertex change map on the final page (both timepoints share point correspondence).",
                fontsize=9, color="#999999",
            )
        _draw_line(
            title_page, y, f"Generated {datetime.now(timezone.utc).strftime('%Y-%m-%d')}", fontsize=8, color="#999999",
        )
        pdf.savefig(title_page)

        for section_label, color, row, draw_figure in sections:
            for group, fields in _PDF_METRIC_FIELDS.items():
                draw = draw_figure.get(group)
                if draw is None:
                    continue

                page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
                FigureCanvasAgg(page)
                page.text(0.5, 0.975, section_label, ha="center", fontsize=12, weight="bold", color=color)
                draw(page, (0.04, 0.46, 0.92, 0.50))
                _draw_metric_fields(page, 0.42, fields, row)
                pdf.savefig(page)

        if diff_heatmap is not None:
            page = Figure(figsize=(PAGE_W_IN, PAGE_H_IN))
            FigureCanvasAgg(page)
            page.text(0.5, 0.96, f"Change from {label_a} to {label_b}", ha="center", fontsize=13, weight="bold")
            synthetic = AsymmetryResult(heatmap=diff_heatmap, mean_asymmetry_index=float(np.mean(np.abs(diff_heatmap))))
            view = "top" if target == "cranium" else "frontal"
            _draw_asymmetry(page, (0.04, 0.06, 0.92, 0.84), mesh_b, synthetic, label="net", view=view)
            pdf.savefig(page)

    return buf.getvalue()


def _build_mesh_files(
    original_filename: str,
    registered_mesh: trimesh.Trimesh,
    final_mesh: trimesh.Trimesh,
    target: str,
    config: dict,
    nicp_mesh: trimesh.Trimesh | None = None,
) -> tuple[str, dict[str, bytes]]:
    """(folder_name, {filename: bytes}) - the two mesh files, {stem}_rg.ply
    / _rg_{C|F}.ply, plus a third _rg_{C|F}N.ply when nicp_mesh is given -
    the topology-consistent template fit ("fit template" in the UI), kept
    alongside the patient's own clipped/resampled mesh rather than in place
    of it. {C|F} is cranial/facial, same convention as results_folder_name."""
    stem = stem_from_filename(original_filename)
    folder = results_folder_name(original_filename, target, config)

    # bare geometry only, no visual/UV - registered_mesh still carries
    # texture data at this point (register() deliberately keeps it, for the
    # live viewer), but trimesh's PLY writer stores UV as double-precision
    # while vertex positions stay float, and never writes a "TextureFile"
    # comment pointing at the actual image - so the UV is orphaned data with
    # no texture resolvable from the file, and the mixed float/double
    # vertex record makes some tools (Meshmixer) read the file as empty.
    # final_mesh never has this problem since repair_mesh already strips
    # visual data before repair runs, further up the pipeline.
    registered_export = trimesh.Trimesh(
        vertices=registered_mesh.vertices, faces=registered_mesh.faces, process=False
    )
    target_suffix = "C" if target == "cranium" else "F"
    files = {
        f"{stem}_rg.ply": registered_export.export(file_type="ply"),
        f"{stem}_rg_{target_suffix}.ply": final_mesh.export(file_type="ply"),
    }
    if nicp_mesh is not None:
        files[f"{stem}_rg_{target_suffix}N.ply"] = nicp_mesh.export(file_type="ply")
    return folder, files


def _build_analysis_files(
    original_filename: str,
    final_mesh: trimesh.Trimesh,
    landmarks: np.ndarray,
    target: str,
    craniometrics: CranioMeasurements | None,
    asymmetry: AsymmetryResult | None,
    config: dict,
    sellion_mesh: trimesh.Trimesh | None = None,
    sellion_landmarks: np.ndarray | None = None,
    metopic: MetopicResult | None = None,
    frontal_bossing: FrontalBossingResult | None = None,
    metadata: dict[str, str] | None = None,
) -> dict[str, bytes]:
    """{filename: bytes} - {stem}_report.json / _measurements.png (cranium
    target only) / _asymmetry.png (face target only) / _summary.xlsx /
    _report.pdf (both always, regardless of target - see _metrics_row/
    _summary_xlsx/_report_pdf).

    sellion_mesh/sellion_landmarks are only given for a cranial analysis that
    used an alt_frontal_landmark - final_mesh and landmarks are the ALT
    frame in that case (deliberately, that's what gets exported as the
    mesh - see analyze_cranial), but the saved 2D figure still has to come
    from the sellion-frame mesh: it's a static picture, so it can't "rotate
    with the viewer" the way the display frame does, and craniometrics was
    computed against sellion in the first place. left None (the default),
    this is just final_mesh/landmarks again - the ordinary single-frame
    case.

    metadata is the patient/visit fields the sidebar form collects (see
    api/schemas.py's PatientMetadata) - None (the default) is treated the
    same as an empty dict, so every field just comes through blank."""
    stem = stem_from_filename(original_filename)
    # "_cranial"/"_frontal" rather than reusing results_folder_name's
    # "C"/"F" - these files often end up standing alone (attached
    # somewhere, dropped into a shared cohort folder) away from the
    # CP_..._C_.../CP_..._F_... folder name that would otherwise be the
    # only thing saying which target a given _report/_summary came from.
    target_suffix = "cranial" if target == "cranium" else "frontal"
    figure_mesh = sellion_mesh if sellion_mesh is not None else final_mesh
    report_landmarks = sellion_landmarks if sellion_landmarks is not None else landmarks
    metadata = metadata or {}

    report = {
        "generated": datetime.now(timezone.utc).isoformat(),
        "source_file": original_filename,
        "target": target,
        # always the landmarks craniometrics was actually computed from
        # (sellion, always - see analyze_cranial) - NOT necessarily the
        # frame the exported mesh is in.
        "landmarks": {
            "sellion": report_landmarks[0].tolist(),
            "left_tragus": report_landmarks[1].tolist(),
            "right_tragus": report_landmarks[2].tolist(),
        },
        "settings": config,
    }
    if sellion_landmarks is not None:
        # the exported mesh's own frame, only present when it differs from
        # the sellion one above
        report["display_frontal_landmark"] = landmarks[0].tolist()
    if craniometrics is not None:
        report["craniometrics"] = {
            "depth_mm": craniometrics.depth_mm,
            "breadth_mm": craniometrics.breadth_mm,
            "cephalic_index": craniometrics.cephalic_index,
            "circumference_cm": craniometrics.circumference_cm,
            "mesh_volume_cc": float(craniometrics.mesh_volume_cc),
            "slice_height": craniometrics.slice_height,
        }
    if asymmetry is not None:
        report["asymmetry"] = {"mean_asymmetry_index": asymmetry.mean_asymmetry_index}
    if metopic is not None:
        report["metopic"] = {
            "frontal_angle_deg": metopic.frontal_angle_deg,
            "midline_curvature_concentration": metopic.midline_curvature_concentration,
            "midline_max_curvature": metopic.midline_max_curvature,
            "midline_max_curvature_position": metopic.midline_max_curvature_position,
            "ridge_protrusion_mm": metopic.ridge_protrusion_mm,
            "ridge_area_mm2": metopic.ridge_area_mm2,
            "ridge_area_normalized": metopic.ridge_area_normalized,
            "left_temporal_hollowing": metopic.left_temporal_hollowing,
            "right_temporal_hollowing": metopic.right_temporal_hollowing,
            "mean_temporal_hollowing": metopic.mean_temporal_hollowing,
            "left_max_temporal_depth_mm": metopic.left_max_temporal_depth_mm,
            "right_max_temporal_depth_mm": metopic.right_max_temporal_depth_mm,
            "parabolic_deviation_index": metopic.parabolic_deviation_index,
        }
    if frontal_bossing is not None:
        report["frontal_bossing"] = {"angle_deg": frontal_bossing.angle_deg}

    # cranial asymmetry uses figure_mesh (always the sellion-frame mesh,
    # same as craniometrics/frontal_bossing's own figures) since asymmetry
    # was computed against that exact mesh's vertices (see
    # pipeline.measure_cranial) - final_mesh could be a different frame
    # entirely (the alt-frontal display mesh), which wouldn't line up with
    # the heatmap array at all. facial has no such split - sellion_mesh is
    # always None there, so figure_mesh already equals final_mesh.
    is_cranial_asymmetry = asymmetry is not None and target == "cranium"
    asymmetry_view = "top" if is_cranial_asymmetry else "frontal"
    asymmetry_label = "cranial" if is_cranial_asymmetry else "facial"
    asymmetry_mesh = figure_mesh if is_cranial_asymmetry else final_mesh
    asymmetry_group = "cranial_asymmetry" if is_cranial_asymmetry else "asymmetry"
    asymmetry_sagittal_group = "cranial_asymmetry_sagittal" if is_cranial_asymmetry else "asymmetry_sagittal"
    asymmetry_png_name = f"{stem}_cranial_asymmetry.png" if is_cranial_asymmetry else f"{stem}_asymmetry.png"
    asymmetry_sagittal_png_name = (
        f"{stem}_cranial_asymmetry_sagittal.png" if is_cranial_asymmetry else f"{stem}_asymmetry_sagittal.png"
    )

    files = {f"{stem}_report_{target_suffix}.json": json.dumps(report, indent=2).encode("utf-8")}
    if craniometrics is not None:
        files[f"{stem}_measurements.png"] = _measurement_figure(figure_mesh, craniometrics)
    if asymmetry is not None:
        files[asymmetry_png_name] = _asymmetry_figure(asymmetry_mesh, asymmetry, label=asymmetry_label, view=asymmetry_view)
        files[asymmetry_sagittal_png_name] = _asymmetry_figure(
            asymmetry_mesh, asymmetry, label=asymmetry_label, view="sagittal"
        )
    if metopic is not None:
        files[f"{stem}_metopic.png"] = _metopic_figure(metopic)
    if frontal_bossing is not None:
        files[f"{stem}_frontal_bossing.png"] = _frontal_bossing_figure(frontal_bossing)

    # the PDF redraws each figure as vector straight onto its page rather
    # than embedding the PNGs above - same _draw_* code, so the two can't
    # drift apart, but nothing on the report is rasterized. see _report_pdf.
    draw_figure: dict[str, Callable[[Figure, tuple[float, float, float, float]], None]] = {}
    if craniometrics is not None:
        draw_figure["craniometrics"] = lambda fig, rect: _draw_measurements(fig, rect, figure_mesh, craniometrics)
    if asymmetry is not None:
        draw_figure[asymmetry_group] = lambda fig, rect: _draw_asymmetry(
            fig, rect, asymmetry_mesh, asymmetry, label=asymmetry_label, view=asymmetry_view
        )
        draw_figure[asymmetry_sagittal_group] = lambda fig, rect: _draw_asymmetry(
            fig, rect, asymmetry_mesh, asymmetry, label=asymmetry_label, view="sagittal"
        )
    if metopic is not None:
        draw_figure["metopic"] = lambda fig, rect: _draw_metopic(fig, rect, metopic)
    if frontal_bossing is not None:
        draw_figure["frontal_bossing"] = lambda fig, rect: _draw_frontal_bossing(fig, rect, frontal_bossing)

    row = _metrics_row(target, metadata, config, craniometrics, asymmetry, metopic, frontal_bossing)
    files[f"{stem}_summary_{target_suffix}.xlsx"] = _summary_xlsx(row)
    files[f"{stem}_report_{target_suffix}.pdf"] = _report_pdf(original_filename, target, metadata, row, draw_figure)
    return files


def _zip_files(files_by_prefix: dict[str, dict[str, bytes]]) -> bytes:
    """{zip-path-prefix: {filename: bytes}} -> zip bytes - shared by every
    zip-producing function below, so meshes-only/analysis-only/combined
    bundles all build the same way."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for prefix, files in files_by_prefix.items():
            for name, content in files.items():
                zf.writestr(f"{prefix}/{name}" if prefix else name, content)
    return buf.getvalue()


def build_meshes_bundle(
    original_filename: str,
    registered_mesh: trimesh.Trimesh,
    final_mesh: trimesh.Trimesh,
    target: str,
    config: dict,
    nicp_mesh: trimesh.Trimesh | None = None,
) -> bytes:
    """zip bytes for the mesh files (two, or three when a NICP fit exists -
    see _build_mesh_files) - the browser-download side of "save meshes"
    (part 9). nested under {folder}/meshes/, matching write_meshes_to_folder's
    own on-disk layout."""
    folder, files = _build_mesh_files(original_filename, registered_mesh, final_mesh, target, config, nicp_mesh)
    return _zip_files({f"{folder}/meshes": files})


def write_meshes_to_folder(
    dest_dir: Path,
    original_filename: str,
    registered_mesh: trimesh.Trimesh,
    final_mesh: trimesh.Trimesh,
    target: str,
    config: dict,
    nicp_mesh: trimesh.Trimesh | None = None,
) -> Path:
    """writes the mesh files (two, or three when a NICP fit exists) into
    dest_dir/{folder}/meshes/ - the desktop side of "save meshes" (part 9),
    sibling to write_analysis_to_folder's own analysis/ subfolder. returns
    the meshes/ folder written."""
    folder, files = _build_mesh_files(original_filename, registered_mesh, final_mesh, target, config, nicp_mesh)
    results_dir = dest_dir / folder / "meshes"
    results_dir.mkdir(parents=True, exist_ok=True)
    for name, content in files.items():
        (results_dir / name).write_bytes(content)
    return results_dir


def build_analysis_bundle(
    original_filename: str,
    registered_mesh: trimesh.Trimesh,
    final_mesh: trimesh.Trimesh,
    landmarks: np.ndarray,
    target: str,
    craniometrics: CranioMeasurements | None,
    asymmetry: AsymmetryResult | None,
    config: dict,
    sellion_mesh: trimesh.Trimesh | None = None,
    sellion_landmarks: np.ndarray | None = None,
    nicp_mesh: trimesh.Trimesh | None = None,
    metopic: MetopicResult | None = None,
    frontal_bossing: FrontalBossingResult | None = None,
    metadata: dict[str, str] | None = None,
    include_meshes: bool = True,
) -> bytes:
    """zip bytes for the mesh files (two, or three - see _build_mesh_files)
    plus a nested analysis/ subfolder - the browser-download side of
    "export analysis" (part 10). self-contained (always includes the
    meshes) since there's no persistent folder to add an analysis/
    subfolder onto across separate browser downloads, unlike
    write_analysis_to_folder below. no cohort_xlsx_path here for the same
    reason - see _upsert_cohort_xlsx, desktop-only.

    include_meshes=False (the "meshes" export checkbox unticked - see
    App.jsx's AnalysisPanel) skips the mesh files entirely, for a
    report-only zip. craniometrics/asymmetry being None already skips
    their own sections wherever they're used (see _build_analysis_files) -
    that's how the "measurements"/"asymmetry" checkboxes work, no separate
    flag needed here; the router just passes None for whichever's unticked."""
    folder = results_folder_name(original_filename, target, config)
    mesh_files = (
        _build_mesh_files(original_filename, registered_mesh, final_mesh, target, config, nicp_mesh)[1]
        if include_meshes
        else {}
    )
    analysis_files = _build_analysis_files(
        original_filename, final_mesh, landmarks, target, craniometrics, asymmetry, config, sellion_mesh, sellion_landmarks,
        metopic, frontal_bossing, metadata,
    )
    files_by_prefix = {f"{folder}/analysis": analysis_files}
    if mesh_files:
        files_by_prefix[f"{folder}/meshes"] = mesh_files
    return _zip_files(files_by_prefix)


def write_analysis_to_folder(
    dest_dir: Path,
    original_filename: str,
    registered_mesh: trimesh.Trimesh,
    final_mesh: trimesh.Trimesh,
    landmarks: np.ndarray,
    target: str,
    craniometrics: CranioMeasurements | None,
    asymmetry: AsymmetryResult | None,
    config: dict,
    sellion_mesh: trimesh.Trimesh | None = None,
    sellion_landmarks: np.ndarray | None = None,
    nicp_mesh: trimesh.Trimesh | None = None,
    metopic: MetopicResult | None = None,
    frontal_bossing: FrontalBossingResult | None = None,
    metadata: dict[str, str] | None = None,
    cohort_xlsx_path: Path | None = None,
    include_meshes: bool = True,
) -> Path:
    """writes the report/figures into dest_dir/{folder}/analysis/ - the
    desktop side of "export analysis" (part 10). if the mesh folder doesn't
    exist yet (meshes were never separately saved) AND include_meshes is
    True (the "meshes" export checkbox - see App.jsx's AnalysisPanel),
    writes those first - see write_meshes_to_folder; with it unticked, only
    the analysis/ subfolder gets written, even into a dest_dir that has
    never had meshes saved into it at all. when cohort_xlsx_path is given,
    also upserts this session's row into that external file (see
    _upsert_cohort_xlsx) - browser mode has no equivalent, there's no
    persistent file across separate zip downloads to append to. returns
    the analysis folder written."""
    folder_name = results_folder_name(original_filename, target, config)
    results_dir = dest_dir / folder_name
    if include_meshes and not (results_dir / "meshes").exists():
        write_meshes_to_folder(dest_dir, original_filename, registered_mesh, final_mesh, target, config, nicp_mesh)

    analysis_dir = results_dir / "analysis"
    analysis_dir.mkdir(parents=True, exist_ok=True)
    files = _build_analysis_files(
        original_filename, final_mesh, landmarks, target, craniometrics, asymmetry, config, sellion_mesh, sellion_landmarks,
        metopic, frontal_bossing, metadata,
    )
    for name, content in files.items():
        (analysis_dir / name).write_bytes(content)
    if cohort_xlsx_path is not None:
        row = _metrics_row(
            target, metadata or {}, config, craniometrics, asymmetry, metopic, frontal_bossing,
            nicp_mesh_path=_nicp_mesh_path(results_dir, original_filename, target, nicp_mesh),
        )
        _upsert_cohort_xlsx(cohort_xlsx_path, row)
    return analysis_dir


def build_results_bundle(
    original_filename: str,
    registered_mesh: trimesh.Trimesh,
    final_mesh: trimesh.Trimesh,
    landmarks: np.ndarray,
    target: str,
    craniometrics: CranioMeasurements | None,
    asymmetry: AsymmetryResult | None,
    config: dict,
    sellion_mesh: trimesh.Trimesh | None = None,
    sellion_landmarks: np.ndarray | None = None,
    nicp_mesh: trimesh.Trimesh | None = None,
    metopic: MetopicResult | None = None,
    frontal_bossing: FrontalBossingResult | None = None,
    metadata: dict[str, str] | None = None,
) -> bytes:
    """zip bytes for the results folder (see results_folder_name), meshes
    (two, or three - see _build_mesh_files) and analysis flat together -
    the original "everything in one shot" browser-download path (pre-dates
    the separate save-meshes/export-analysis split - part 9/10 - and keeps
    its existing flat layout for backward compatibility rather than
    adopting their nested analysis/ convention)."""
    folder, mesh_files = _build_mesh_files(original_filename, registered_mesh, final_mesh, target, config, nicp_mesh)
    analysis_files = _build_analysis_files(
        original_filename, final_mesh, landmarks, target, craniometrics, asymmetry, config, sellion_mesh, sellion_landmarks,
        metopic, frontal_bossing, metadata,
    )
    return _zip_files({folder: {**mesh_files, **analysis_files}})


def write_results_to_folder(
    dest_dir: Path,
    original_filename: str,
    registered_mesh: trimesh.Trimesh,
    final_mesh: trimesh.Trimesh,
    landmarks: np.ndarray,
    target: str,
    craniometrics: CranioMeasurements | None,
    asymmetry: AsymmetryResult | None,
    config: dict,
    sellion_mesh: trimesh.Trimesh | None = None,
    sellion_landmarks: np.ndarray | None = None,
    nicp_mesh: trimesh.Trimesh | None = None,
    metopic: MetopicResult | None = None,
    frontal_bossing: FrontalBossingResult | None = None,
    metadata: dict[str, str] | None = None,
    cohort_xlsx_path: Path | None = None,
) -> Path:
    """writes the results folder (see results_folder_name) straight into
    dest_dir, meshes and analysis flat together - the original "everything
    in one shot" desktop-save path (see build_results_bundle for why this
    stays flat rather than nesting analysis/ like part 9/10's split
    functions do). cohort_xlsx_path behaves the same as on
    write_analysis_to_folder - see _upsert_cohort_xlsx."""
    folder, mesh_files = _build_mesh_files(original_filename, registered_mesh, final_mesh, target, config, nicp_mesh)
    analysis_files = _build_analysis_files(
        original_filename, final_mesh, landmarks, target, craniometrics, asymmetry, config, sellion_mesh, sellion_landmarks,
        metopic, frontal_bossing, metadata,
    )
    results_dir = dest_dir / folder
    results_dir.mkdir(parents=True, exist_ok=True)
    for name, content in {**mesh_files, **analysis_files}.items():
        (results_dir / name).write_bytes(content)
    if cohort_xlsx_path is not None:
        row = _metrics_row(
            target, metadata or {}, config, craniometrics, asymmetry, metopic, frontal_bossing,
            nicp_mesh_path=_nicp_mesh_path(results_dir, original_filename, target, nicp_mesh),
        )
        _upsert_cohort_xlsx(cohort_xlsx_path, row)
    return results_dir
