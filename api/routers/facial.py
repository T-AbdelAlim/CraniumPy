"""Facial Anthropometrics workspace endpoints: template load/pick/measurement-
preview (the definition phase, against the template mesh only), batch
extraction across many already-NICP-registered patient meshes, per-mesh
landmark correction, and the batch's own Excel export. See
craniumpy_core.facial_measurements for the actual geometry - this is just
the request/response plumbing and the caching, the same split every other
router in this app already uses (compare api/routers/cohort.py against
craniumpy_core.cohort).

caching here mirrors api/routers/cohort.py's own _mean_shape_cache exactly:
plain dicts, FIFO-evicted at a small cap, populated synchronously - no
job-queue/session machinery, since even a large (~200 file) batch is fast
enough to process inline (see facial_measurements.py's own module docstring
for why: geodesic distance and enclosed-area topology are computed ONCE per
template and reused unchanged for every batch mesh, so per-mesh work is
just cheap vectorized numeric recomputation).

BATCH MESHES ARE NEVER HELD IN MEMORY beyond the single request that needs
them - _process_one_mesh loads, measures, and lets its `trimesh.Trimesh` go
out of scope; only the resulting FacialBatchFileResult (landmark points +
scalar values, no mesh data) is cached under batch_id. the review panel's
own GET .../mesh/{filename} reloads a mesh fresh from disk every time it's
opened, on the same reasoning - a 50-200 file batch never needs more than
one full mesh in memory at once.
"""

from __future__ import annotations

import io
import uuid
from pathlib import Path

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from scipy.spatial import cKDTree

from craniumpy_core.cohort import _validate_same_topology
from craniumpy_core.facial_measurements import (
    BoundaryTopology,
    build_area_boundary,
    build_topology,
    compute_measurement,
    geodesic_path_vertices,
    nearest_vertex_index,
)
from craniumpy_core.io import load_mesh, mesh_to_glb
from craniumpy_core.template_registry import load_shipped_template
from api.routers.cohort import _EXPORT_HEADER_FILL, _EXPORT_HEADER_FONT, _sheet_name, _write_generic_sheet
from api.schemas import (
    CohortExportSheet,
    FacialBatchFileResult,
    FacialBatchResponse,
    FacialBatchStartRequest,
    FacialCorrectionRequest,
    FacialCorrectionResponse,
    FacialListMeshesRequest,
    FacialListMeshesResponse,
    FacialMeasurementDef,
    FacialMeasurementPreviewRequest,
    FacialMeasurementPreviewResponse,
    FacialPickRequest,
    FacialPickResponse,
    FacialTemplateLoadRequest,
    FacialTemplateResponse,
    LandmarkPoint,
)

_MESH_EXTENSIONS = {".ply", ".obj", ".stl"}

router = APIRouter(prefix="/api/facial", tags=["facial"])

_UNIT_BY_TYPE = {"linear": "mm", "angular": "deg", "area": "mm2"}

# loaded templates - small (a handful at most, one per template the user's
# actually opened this session), FIFO-evicted, same shape as cohort.py's
# own _mean_shape_cache. holds the one genuinely reusable-across-a-whole-
# batch object set: the mesh itself, its cKDTree (for landmark picking/
# snapping), and its MeshTopology (edge graph - see facial_measurements.py).
_template_cache: dict[str, dict] = {}
_TEMPLATE_CACHE_MAX = 5

# BoundaryTopology per (template_id, point_ids tuple, vertex_indices tuple) -
# keying on the RESOLVED vertex indices (not just which points) means a
# landmark move naturally invalidates the old entry (different indices ->
# different key -> cache miss -> fresh boundary) with no explicit
# invalidation logic needed at all.
_boundary_cache: dict[tuple, BoundaryTopology] = {}
_BOUNDARY_CACHE_MAX = 50

# batches - small (landmark points + scalar values per file, never mesh
# data), FIFO-evicted.
_batch_cache: dict[str, dict] = {}
_BATCH_CACHE_MAX = 10


def _evict(cache: dict, max_size: int) -> None:
    while len(cache) > max_size:
        cache.pop(next(iter(cache)))


def _get_template(template_id: str) -> dict:
    entry = _template_cache.get(template_id)
    if entry is None:
        raise HTTPException(status_code=404, detail=f"no template {template_id!r} (or it's since been evicted - reload it)")
    return entry


def _get_batch(batch_id: str) -> dict:
    batch = _batch_cache.get(batch_id)
    if batch is None:
        raise HTTPException(status_code=404, detail=f"no batch {batch_id!r} (or it's since been evicted - re-run it)")
    return batch


def _resolve_vertex_index(entry: dict, points: dict[str, LandmarkPoint], point_id: str) -> int:
    point = points.get(point_id)
    if point is None:
        raise ValueError(f"unknown point id {point_id!r}")
    return nearest_vertex_index(entry["mesh"], [point.x, point.y, point.z], kdtree=entry["kdtree"])


def _get_or_build_boundary(template_id: str, entry: dict, point_ids: list[str], vertex_indices: list[int]) -> BoundaryTopology:
    key = (template_id, tuple(point_ids), tuple(vertex_indices))
    cached = _boundary_cache.get(key)
    if cached is not None:
        return cached
    boundary = build_area_boundary(entry["mesh"], entry["topology"], vertex_indices)
    _boundary_cache[key] = boundary
    _evict(_boundary_cache, _BOUNDARY_CACHE_MAX)
    return boundary


@router.post("/list-meshes", response_model=FacialListMeshesResponse)
def list_meshes(request: FacialListMeshesRequest) -> FacialListMeshesResponse:
    folder = Path(request.folder)
    if not folder.is_dir():
        raise HTTPException(status_code=400, detail=f"folder not found: {request.folder}")
    paths = sorted(
        str(p) for p in folder.iterdir() if p.is_file() and p.suffix.lower() in _MESH_EXTENSIONS
    )
    return FacialListMeshesResponse(mesh_paths=paths)


@router.post("/template/load", response_model=FacialTemplateResponse)
def load_template(request: FacialTemplateLoadRequest) -> FacialTemplateResponse:
    if request.path:
        path = Path(request.path)
        if not path.is_file():
            raise HTTPException(status_code=400, detail=f"file not found: {request.path}")
        try:
            mesh = load_mesh(path)
        except Exception as exc:  # noqa: BLE001 - want the real reason surfaced
            raise HTTPException(status_code=400, detail=f"could not read template mesh: {exc}") from exc
    else:
        shipped_name = request.shipped_name or "template_face"
        try:
            mesh = load_shipped_template(shipped_name)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc

    template_id = str(uuid.uuid4())
    _template_cache[template_id] = {
        "mesh": mesh,
        "kdtree": cKDTree(mesh.vertices),
        "topology": build_topology(mesh),
    }
    _evict(_template_cache, _TEMPLATE_CACHE_MAX)
    return FacialTemplateResponse(template_id=template_id, vertex_count=len(mesh.vertices), face_count=len(mesh.faces))


@router.get("/template/{template_id}/mesh")
def get_template_mesh(template_id: str):
    entry = _get_template(template_id)
    return Response(content=mesh_to_glb(entry["mesh"]), media_type="model/gltf-binary")


@router.post("/template/{template_id}/pick", response_model=FacialPickResponse)
def pick_point(template_id: str, request: FacialPickRequest) -> FacialPickResponse:
    entry = _get_template(template_id)
    idx = nearest_vertex_index(entry["mesh"], [request.point.x, request.point.y, request.point.z], kdtree=entry["kdtree"])
    snapped = entry["mesh"].vertices[idx]
    return FacialPickResponse(vertex_index=idx, point=LandmarkPoint(x=float(snapped[0]), y=float(snapped[1]), z=float(snapped[2])))


def _measure_one(
    mesh, topology, m: FacialMeasurementDef, vertex_index_by_point: dict[str, int], boundary: BoundaryTopology | None
) -> float:
    vertex_indices = [vertex_index_by_point[pid] for pid in m.point_ids]
    return compute_measurement(mesh, topology, m.type, vertex_indices, geodesic=m.geodesic, boundary=boundary)


def _points_from_indices(mesh, indices) -> list[LandmarkPoint]:
    verts = mesh.vertices[indices]
    return [LandmarkPoint(x=float(p[0]), y=float(p[1]), z=float(p[2])) for p in verts]


def _points_from_array(arr) -> list[LandmarkPoint]:
    return [LandmarkPoint(x=float(p[0]), y=float(p[1]), z=float(p[2])) for p in arr]


def _render_geometry(
    mesh, topology, m: FacialMeasurementDef, vertex_index_by_point: dict[str, int], boundary: BoundaryTopology | None
) -> tuple[list[LandmarkPoint] | None, list[LandmarkPoint] | None]:
    """the VISUAL overlay geometry for one measurement - always traces along
    the mesh surface, regardless of whether the measurement's own computed
    VALUE is a straight or geodesic distance (a straight chord between two
    points on a curved face reads as floating off the surface, so the line
    the viewer draws always hugs the mesh even for a "straight distance"
    Linear measurement - the number shown is still the real straight-line
    value, this only changes what gets drawn). returns (path_points,
    boundary_face_triangle_points) - Area only ever returns the second,
    Linear/Angular only the first. (None, None) - never raised - if a path
    can't be traced (a disconnected mesh): the frontend falls back to a
    plain straight connector between the raw landmark points in that case,
    same as it already does before this round trip returns."""
    try:
        if m.type == "linear":
            a, b = (vertex_index_by_point[pid] for pid in m.point_ids)
            return _points_from_indices(mesh, geodesic_path_vertices(mesh, topology, a, b)), None
        if m.type == "angular":
            a, vertex, c = (vertex_index_by_point[pid] for pid in m.point_ids)
            leg_a = geodesic_path_vertices(mesh, topology, vertex, a)
            leg_c = geodesic_path_vertices(mesh, topology, vertex, c)
            path = list(reversed(leg_a)) + leg_c[1:]  # a -> vertex -> c, one continuous surface trace
            return _points_from_indices(mesh, path), None
        if m.type == "area" and boundary is not None:
            loop_points = _points_from_indices(mesh, boundary.boundary_vertex_loop)
            face_verts = mesh.vertices[mesh.faces[boundary.face_indices]].reshape(-1, 3)
            return loop_points, _points_from_array(face_verts)
    except ValueError:
        pass
    return None, None


@router.post("/template/{template_id}/measurement/preview", response_model=FacialMeasurementPreviewResponse)
def preview_measurements(template_id: str, request: FacialMeasurementPreviewRequest) -> FacialMeasurementPreviewResponse:
    """live values while still defining measurements - computed directly on
    the template mesh itself (no batch involved yet)."""
    entry = _get_template(template_id)
    mesh, topology = entry["mesh"], entry["topology"]
    values: dict[str, float | None] = {}
    errors: dict[str, str] = {}
    render_paths: dict[str, list[LandmarkPoint]] = {}
    render_faces: dict[str, list[LandmarkPoint]] = {}

    for m in request.measurements:
        try:
            vertex_index_by_point = {pid: _resolve_vertex_index(entry, request.points, pid) for pid in m.point_ids}
            boundary = None
            if m.type == "area":
                vertex_indices = [vertex_index_by_point[pid] for pid in m.point_ids]
                boundary = _get_or_build_boundary(template_id, entry, m.point_ids, vertex_indices)
            values[m.id] = _measure_one(mesh, topology, m, vertex_index_by_point, boundary)
        except ValueError as exc:
            values[m.id] = None
            errors[m.id] = str(exc)
            continue
        path, faces = _render_geometry(mesh, topology, m, vertex_index_by_point, boundary)
        if path is not None:
            render_paths[m.id] = path
        if faces is not None:
            render_faces[m.id] = faces

    return FacialMeasurementPreviewResponse(values=values, value_errors=errors, render_paths=render_paths, render_faces=render_faces)


def _process_one_mesh(
    path: Path,
    entry: dict,
    vertex_index_by_point: dict[str, int],
    boundary_by_measurement: dict[str, BoundaryTopology],
    boundary_errors: dict[str, str],
    measurements: list[FacialMeasurementDef],
) -> FacialBatchFileResult:
    """load, validate, measure, and let the mesh go - never kept around
    past this one function call. every failure (bad file, wrong topology,
    a single bad measurement) is caught and recorded per-file/per-
    measurement, never raised past this function - a batch never aborts on
    one bad file."""
    filename = path.name
    try:
        mesh = load_mesh(path)
    except Exception as exc:  # noqa: BLE001
        return FacialBatchFileResult(filename=filename, status="error", error=f"could not load mesh: {exc}")

    try:
        _validate_same_topology(mesh, entry["mesh"], filename, "the selected template")
    except ValueError as exc:
        return FacialBatchFileResult(filename=filename, status="error", error=str(exc))

    topology = entry["topology"]
    landmark_points = {}
    for pid, vidx in vertex_index_by_point.items():
        p = mesh.vertices[vidx]
        landmark_points[pid] = LandmarkPoint(x=float(p[0]), y=float(p[1]), z=float(p[2]))

    values: dict[str, float | None] = {}
    value_errors: dict[str, str] = {}
    render_paths: dict[str, list[LandmarkPoint]] = {}
    render_faces: dict[str, list[LandmarkPoint]] = {}
    for m in measurements:
        if m.id in boundary_errors:
            values[m.id] = None
            value_errors[m.id] = boundary_errors[m.id]
            continue
        try:
            boundary = boundary_by_measurement.get(m.id)
            values[m.id] = _measure_one(mesh, topology, m, vertex_index_by_point, boundary)
        except (KeyError, ValueError) as exc:
            values[m.id] = None
            value_errors[m.id] = str(exc)
            continue
        path, faces = _render_geometry(mesh, topology, m, vertex_index_by_point, boundary)
        if path is not None:
            render_paths[m.id] = path
        if faces is not None:
            render_faces[m.id] = faces

    return FacialBatchFileResult(
        filename=filename,
        status="ok",
        landmark_points=landmark_points,
        values=values,
        value_errors=value_errors,
        render_paths=render_paths,
        render_faces=render_faces,
    )


@router.post("/batch/start", response_model=FacialBatchResponse)
def start_batch(request: FacialBatchStartRequest) -> FacialBatchResponse:
    entry = _get_template(request.template_id)
    mesh, topology = entry["mesh"], entry["topology"]

    missing = [p for p in request.mesh_paths if not Path(p).is_file()]
    if missing:
        raise HTTPException(status_code=400, detail=f"mesh file(s) not found: {', '.join(missing)}")

    # every point resolved to a TEMPLATE vertex index exactly once - this is
    # what every batch mesh's own vertices[vertex_index] transfers via, no
    # per-mesh re-snapping (see this module's own docstring).
    vertex_index_by_point: dict[str, int] = {}
    for pid, point in request.points.items():
        vertex_index_by_point[pid] = nearest_vertex_index(mesh, [point.x, point.y, point.z], kdtree=entry["kdtree"])

    boundary_by_measurement: dict[str, BoundaryTopology] = {}
    boundary_errors: dict[str, str] = {}
    for m in request.measurements:
        if m.type != "area":
            continue
        try:
            vertex_indices = [vertex_index_by_point[pid] for pid in m.point_ids]
            boundary_by_measurement[m.id] = _get_or_build_boundary(request.template_id, entry, m.point_ids, vertex_indices)
        except (KeyError, ValueError) as exc:
            boundary_errors[m.id] = str(exc)

    results = [
        _process_one_mesh(Path(p), entry, vertex_index_by_point, boundary_by_measurement, boundary_errors, request.measurements)
        for p in request.mesh_paths
    ]

    measurements_by_point: dict[str, list[str]] = {}
    for m in request.measurements:
        for pid in m.point_ids:
            measurements_by_point.setdefault(pid, []).append(m.id)

    batch_id = str(uuid.uuid4())
    _batch_cache[batch_id] = {
        "template_id": request.template_id,
        "mesh_paths_by_filename": {Path(p).name: p for p in request.mesh_paths},
        "vertex_index_by_point": vertex_index_by_point,
        "measurements": {m.id: m for m in request.measurements},
        "measurements_by_point": measurements_by_point,
        "vertex_overrides": {},  # {filename: {point_id: vertex_index}} - per-mesh corrections
        "results": results,
    }
    _evict(_batch_cache, _BATCH_CACHE_MAX)
    return FacialBatchResponse(batch_id=batch_id, results=results)


@router.get("/batch/{batch_id}/mesh/{filename}")
def get_batch_mesh(batch_id: str, filename: str):
    batch = _get_batch(batch_id)
    path_str = batch["mesh_paths_by_filename"].get(filename)
    if path_str is None:
        raise HTTPException(status_code=404, detail=f"no file {filename!r} in this batch")
    try:
        mesh = load_mesh(Path(path_str))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=404, detail=f"could not load {filename!r}: {exc}") from exc
    return Response(content=mesh_to_glb(mesh), media_type="model/gltf-binary")


@router.post("/batch/{batch_id}/correct", response_model=FacialCorrectionResponse)
def correct_landmark(batch_id: str, request: FacialCorrectionRequest) -> FacialCorrectionResponse:
    batch = _get_batch(batch_id)
    path_str = batch["mesh_paths_by_filename"].get(request.filename)
    if path_str is None:
        raise HTTPException(status_code=404, detail=f"no file {request.filename!r} in this batch")
    if request.point_id not in batch["vertex_index_by_point"]:
        raise HTTPException(status_code=400, detail=f"unknown point id {request.point_id!r}")

    result = next((r for r in batch["results"] if r.filename == request.filename), None)
    if result is None:
        raise HTTPException(status_code=404, detail=f"no result for {request.filename!r}")

    try:
        mesh = load_mesh(Path(path_str))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=404, detail=f"could not load {request.filename!r}: {exc}") from exc

    corrected_index = nearest_vertex_index(mesh, [request.point.x, request.point.y, request.point.z])
    snapped = mesh.vertices[corrected_index]
    snapped_point = LandmarkPoint(x=float(snapped[0]), y=float(snapped[1]), z=float(snapped[2]))

    # a per-mesh override, independent of every other file and of the
    # template's own vertex_index_by_point - "corrections are preserved
    # independently per mesh."
    overrides = batch["vertex_overrides"].setdefault(request.filename, {})
    overrides[request.point_id] = corrected_index
    result.landmark_points[request.point_id] = snapped_point

    topology = _template_cache[batch["template_id"]]["topology"]
    affected_ids = batch["measurements_by_point"].get(request.point_id, [])
    values: dict[str, float | None] = {}
    value_errors: dict[str, str] = {}
    render_paths: dict[str, list[LandmarkPoint]] = {}
    render_faces: dict[str, list[LandmarkPoint]] = {}
    for mid in affected_ids:
        m = batch["measurements"][mid]
        try:
            vertex_indices = [overrides.get(pid, batch["vertex_index_by_point"][pid]) for pid in m.point_ids]
            vertex_index_by_point = dict(zip(m.point_ids, vertex_indices))
            boundary = None
            if m.type == "area":
                # never the shared template-derived cache - this ONE mesh's
                # own boundary, since its geometry (and possibly this very
                # correction) is specific to it alone.
                boundary = build_area_boundary(mesh, topology, vertex_indices)
            value = compute_measurement(mesh, topology, m.type, vertex_indices, geodesic=m.geodesic, boundary=boundary)
            values[mid] = value
            result.values[mid] = value
            result.value_errors.pop(mid, None)
            path, faces = _render_geometry(mesh, topology, m, vertex_index_by_point, boundary)
            if path is not None:
                render_paths[mid] = path
                result.render_paths[mid] = path
            else:
                result.render_paths.pop(mid, None)
            if faces is not None:
                render_faces[mid] = faces
                result.render_faces[mid] = faces
            else:
                result.render_faces.pop(mid, None)
        except ValueError as exc:
            values[mid] = None
            value_errors[mid] = str(exc)
            result.values[mid] = None
            result.value_errors[mid] = str(exc)
            result.render_paths.pop(mid, None)
            result.render_faces.pop(mid, None)

    return FacialCorrectionResponse(
        landmark_point=snapped_point, values=values, value_errors=value_errors, render_paths=render_paths, render_faces=render_faces
    )


def _measurement_header(m: FacialMeasurementDef) -> str:
    return f"{m.name} ({m.abbreviation})"


@router.post("/batch/{batch_id}/export")
def export_batch(batch_id: str):
    batch = _get_batch(batch_id)
    measurements: list[FacialMeasurementDef] = list(batch["measurements"].values())
    results: list[FacialBatchFileResult] = batch["results"]

    columns = ["identifier"] + [_measurement_header(m) for m in measurements]
    rows = []
    for r in results:
        row = {"identifier": r.filename}
        for m in measurements:
            value = r.values.get(m.id)
            row[_measurement_header(m)] = "" if value is None else str(round(value, 4))
        rows.append(row)
    measurements_sheet = CohortExportSheet(title="measurements", columns=columns, rows=rows)

    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    ws = wb.create_sheet(_sheet_name("measurements", 0, used_names))
    _write_generic_sheet(ws, measurements_sheet, "facial_measurements_table")

    legend_ws = wb.create_sheet(_sheet_name("legend", 1, used_names))
    legend_ws.append(["name", "abbreviation", "type", "unit", "geodesic", "color"])
    for cell in legend_ws[1]:
        cell.font = _EXPORT_HEADER_FONT
        cell.fill = _EXPORT_HEADER_FILL
    for i, m in enumerate(measurements):
        color_hex = m.color.lstrip("#").upper() or "4ADE80"
        legend_ws.append([m.name, m.abbreviation, m.type, _UNIT_BY_TYPE[m.type], "yes" if m.geodesic else "no", ""])
        swatch_cell = legend_ws.cell(row=i + 2, column=6)
        swatch_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
    legend_ws.freeze_panes = "A2"
    for col, width in zip("ABCDEF", (24, 12, 10, 8, 10, 10)):
        legend_ws.column_dimensions[col].width = width

    failed = [r for r in results if r.status == "error"]
    if failed:
        failed_sheet = CohortExportSheet(
            title="failed", columns=["filename", "error"], rows=[{"filename": r.filename, "error": r.error or ""} for r in failed]
        )
        failed_ws = wb.create_sheet(_sheet_name("failed", 2, used_names))
        _write_generic_sheet(failed_ws, failed_sheet, "facial_failed_table")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="facial_measurements.xlsx"'},
    )
