"""cohort/batch-analysis endpoints: loading an accumulated cohort
spreadsheet, running the one thing that genuinely needs a real numerical
library (inferential statistical tests, via scipy.stats - everything else
in the cohort workspace, descriptive stats and plotting, runs client-side
in JS, see frontend/src/workspaces/cohort/lib/stats.js for why), and
computing/serving a mean 3D shape across same-template NICP-fitted meshes.
See craniumpy_core.cohort for the actual math - this is just the request/
response plumbing around it, the same split every other router in this
app already uses (compare api/routers/mesh.py against craniumpy_core.pipeline).
"""

from __future__ import annotations

import io
import re
import uuid
from pathlib import Path

import numpy as np
from fastapi import APIRouter, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from scipy import stats as scipy_stats

from craniumpy_core import cohort
from craniumpy_core.io import mesh_to_glb
from craniumpy_core.template_registry import load_shipped_template
from api.results_bundle import mean_shape_report_pdf
from api.routers._group_measurements import group_measurements_response
from api.schemas import (
    CohortDataResponse,
    CohortExportRequest,
    CohortExportSheet,
    CohortLoadRequest,
    CohortMeanShapeMeasurementsResponse,
    CohortMeanShapeRequest,
    CohortMeanShapeResponse,
    CohortReferenceDiffResponse,
    CohortReportRequest,
    CohortSagittalBandRequest,
    CohortSagittalBandResponse,
    CohortSpreadBandResponse,
    CohortStatsTestRequest,
    CohortStatsTestResponse,
    LandmarkPoint,
)

router = APIRouter(prefix="/api/cohort", tags=["cohort"])

# computed mean shapes, kept just long enough to be fetched once as a GLB
# right after POST /mean-shape returns its result_id - same "no real
# persistence needed" reasoning as api/sessions.py's SessionStore, just
# without any of the async-job machinery: unlike repair/NICP, averaging a
# handful of already-loaded meshes is fast enough to run synchronously
# inline. capped and FIFO-evicted rather than left to grow unboundedly
# across a long session, since nothing else here ever clears an entry.
_mean_shape_cache: dict[str, cohort.MeanShapeResult] = {}
_MEAN_SHAPE_CACHE_MAX = 20


def _cache_mean_shape(result: cohort.MeanShapeResult) -> str:
    result_id = str(uuid.uuid4())
    _mean_shape_cache[result_id] = result
    while len(_mean_shape_cache) > _MEAN_SHAPE_CACHE_MAX:
        _mean_shape_cache.pop(next(iter(_mean_shape_cache)))
    return result_id


@router.post("/load", response_model=CohortDataResponse)
def load_cohort(request: CohortLoadRequest) -> CohortDataResponse:
    path = Path(request.path)
    if not path.is_file():
        raise HTTPException(status_code=400, detail=f"file not found: {request.path}")
    try:
        columns, rows = cohort.load_cohort_xlsx(path)
    except Exception as exc:  # noqa: BLE001 - want the real reason surfaced, whatever openpyxl raised
        raise HTTPException(status_code=400, detail=f"could not read cohort file: {exc}") from exc
    return CohortDataResponse(columns=columns, rows=rows)


@router.get("/demo", response_model=CohortDataResponse)
def load_demo() -> CohortDataResponse:
    try:
        columns, rows = cohort.load_demo_cohort()
    except FileNotFoundError as exc:
        raise HTTPException(status_code=404, detail=str(exc)) from exc
    return CohortDataResponse(columns=columns, rows=rows)


@router.post("/upload", response_model=CohortDataResponse)
async def upload_cohort(file: UploadFile) -> CohortDataResponse:
    """browser: the same load, from an uploaded file's bytes instead of a
    real path - there's no persistent file on this machine to point at
    from inside a browser tab."""
    contents = await file.read()
    try:
        columns, rows = cohort.load_cohort_xlsx(io.BytesIO(contents))
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"could not read cohort file: {exc}") from exc
    return CohortDataResponse(columns=columns, rows=rows)


def _run_stats_test(groups: dict[str, list[float]]) -> CohortStatsTestResponse:
    """2 groups: Welch's t-test (doesn't assume equal variance) +
    Mann-Whitney U (rank-based, no normality assumption at all). 3+ groups:
    one-way ANOVA + Kruskal-Wallis H, the same parametric/rank-based
    pairing one level up. both numbers come back together rather than the
    backend silently picking one - which one to trust depends on things
    (sample size, actual distribution shape) this endpoint has no way to
    judge on the caller's behalf."""
    labels = list(groups.keys())
    if len(labels) < 2:
        raise HTTPException(status_code=400, detail="need at least 2 groups to compare")
    for label in labels:
        if len(groups[label]) < 2:
            raise HTTPException(status_code=400, detail=f"group {label!r} has fewer than 2 values")

    arrays = [np.asarray(groups[label], dtype=float) for label in labels]
    if len(labels) == 2:
        primary = scipy_stats.ttest_ind(arrays[0], arrays[1], equal_var=False)
        alternative = scipy_stats.mannwhitneyu(arrays[0], arrays[1], alternative="two-sided")
        test_name, alternative_name = "Welch's t-test", "Mann-Whitney U"
    else:
        primary = scipy_stats.f_oneway(*arrays)
        alternative = scipy_stats.kruskal(*arrays)
        test_name, alternative_name = "One-way ANOVA", "Kruskal-Wallis H"

    return CohortStatsTestResponse(
        n_groups=len(labels),
        group_sizes={label: len(groups[label]) for label in labels},
        test_name=test_name,
        statistic=float(primary.statistic),
        p_value=float(primary.pvalue),
        alternative_test_name=alternative_name,
        alternative_statistic=float(alternative.statistic),
        alternative_p_value=float(alternative.pvalue),
    )


@router.post("/stats-test", response_model=CohortStatsTestResponse)
def stats_test(request: CohortStatsTestRequest) -> CohortStatsTestResponse:
    return _run_stats_test(request.values)


@router.post("/mean-shape", response_model=CohortMeanShapeResponse)
def compute_mean_shape(request: CohortMeanShapeRequest) -> CohortMeanShapeResponse:
    paths = [Path(p) for p in request.mesh_paths]
    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        raise HTTPException(status_code=400, detail=f"mesh file(s) not found: {', '.join(missing)}")
    try:
        result = cohort.mean_shape(paths)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    result_id = _cache_mean_shape(result)
    return CohortMeanShapeResponse(
        result_id=result_id,
        vertex_count=len(result.mesh.vertices),
        source_count=result.source_count,
        heatmap=result.variability.tolist(),
    )


@router.get("/mean-shape/{result_id}/mesh")
def get_mean_shape_mesh(result_id: str):
    result = _mean_shape_cache.get(result_id)
    if result is None:
        raise HTTPException(status_code=404, detail=f"no mean-shape result {result_id!r} (or it's since been evicted)")
    glb_bytes = mesh_to_glb(result.mesh)
    return Response(content=glb_bytes, media_type="model/gltf-binary")


_SAFE_FILENAME_RE = re.compile(r"[^A-Za-z0-9_.-]+")


def _sanitize_filename(name: str, extension: str = ".ply", default: str = "mean") -> str:
    """strips anything outside a conservative filename-safe character set
    (letters/digits/underscore/dash/dot) - filename is built client-side
    from the active filter/stratification state (see frontend/src/
    workspaces/cohort/lib/naming.js), which is itself derived from cohort
    spreadsheet cell values (a diagnosis string, a treatment label, ...),
    not free-typed user input - but it still crosses a network boundary
    into a header this response actually sends, so it's sanitized here
    rather than trusted."""
    cleaned = _SAFE_FILENAME_RE.sub("-", name).strip("-.") or (default + extension)
    if not cleaned.lower().endswith(extension):
        cleaned += extension
    return cleaned[:120]


@router.get("/mean-shape/{result_id}/download")
def download_mean_shape_mesh(result_id: str, filename: str = "mean.ply"):
    """the mean shape as a downloadable .ply (not the GLB /mesh serves the
    live viewer - a research export should be the same format every other
    mesh this app produces already is), named by the caller - see this
    module's docstring and the frontend's naming.js for how that name
    reflects the active filters/stratification, so a saved file is
    self-describing (e.g. "trigonocephaly_pre-op_surgical_mean.ply")
    without the user having to remember what they filtered to get it."""
    result = _mean_shape_cache.get(result_id)
    if result is None:
        raise HTTPException(status_code=404, detail=f"no mean-shape result {result_id!r} (or it's since been evicted)")
    safe_name = _sanitize_filename(filename)
    ply_bytes = result.mesh.export(file_type="ply")
    return Response(
        content=ply_bytes,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@router.get("/mean-shape/{result_id}/measurements", response_model=CohortMeanShapeMeasurementsResponse)
def get_mean_shape_measurements(result_id: str, target: str) -> CohortMeanShapeMeasurementsResponse:
    """the same measurement suite the Patients workspace's Analysis tab
    shows, run directly on an already-computed mean shape - see
    craniumpy_core.cohort.measure_mean_shape. target has to be given
    explicitly (a mean shape has no target of its own the way a patient
    session does) - "cranium" or "face", matching whichever the group's
    own rows actually are."""
    result = _mean_shape_cache.get(result_id)
    if result is None:
        raise HTTPException(status_code=404, detail=f"no mean-shape result {result_id!r} (or it's since been evicted)")
    if target not in ("cranium", "face"):
        raise HTTPException(status_code=400, detail=f"target must be 'cranium' or 'face', got {target!r}")

    gm = cohort.measure_mean_shape(result.mesh, target)
    return group_measurements_response(result.mesh, gm)


@router.get("/mean-shape/{result_id}/reference-diff", response_model=CohortReferenceDiffResponse)
def get_reference_diff(result_id: str, template: str) -> CohortReferenceDiffResponse:
    """signed displacement of an already-computed mean shape from a shipped
    reference template (see craniumpy_core.cohort.reference_diff) - the
    frontend's Mean shape tab defaults `template` to the same nicp_template
    the group was fit to (the only choice guaranteed to match topology),
    but any shipped template name is accepted; a mismatch comes back as a
    clear 400 rather than a silently meaningless comparison."""
    result = _mean_shape_cache.get(result_id)
    if result is None:
        raise HTTPException(status_code=404, detail=f"no mean-shape result {result_id!r} (or it's since been evicted)")
    try:
        reference_mesh = load_shipped_template(template)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    try:
        heatmap = cohort.reference_diff(result.mesh, reference_mesh)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return CohortReferenceDiffResponse(heatmap=heatmap.tolist())


@router.post("/sagittal-band", response_model=CohortSagittalBandResponse)
def compute_sagittal_band(request: CohortSagittalBandRequest) -> CohortSagittalBandResponse:
    """mean +/- SD of the group's sagittal midline forehead-to-vertex depth
    profile - see craniumpy_core.cohort.sagittal_midline_band. computed
    independently of any cached /mean-shape result (it needs every
    individual mesh, not just their average), so this takes mesh_paths
    directly, same as POST /mean-shape itself."""
    if request.target not in ("cranium", "face"):
        raise HTTPException(status_code=400, detail=f"target must be 'cranium' or 'face', got {request.target!r}")
    paths = [Path(p) for p in request.mesh_paths]
    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        raise HTTPException(status_code=400, detail=f"mesh file(s) not found: {', '.join(missing)}")
    try:
        band = cohort.sagittal_midline_band(paths, request.target)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return CohortSagittalBandResponse(
        y=band.y.tolist(), mean_z=band.mean_z.tolist(), sd_z=band.sd_z.tolist(), source_count=band.source_count,
    )


def _spread_band_response(band: cohort.SpreadBand) -> CohortSpreadBandResponse:
    def points(arr: np.ndarray) -> list[LandmarkPoint]:
        return [LandmarkPoint(x=p[0], y=p[1], z=p[2]) for p in arr]

    return CohortSpreadBandResponse(
        mean=points(band.mean), inner=points(band.inner), outer=points(band.outer),
        closed=band.closed, source_count=band.source_count,
    )


@router.post("/hc-ring-band", response_model=CohortSpreadBandResponse)
def compute_hc_ring_band(request: CohortSagittalBandRequest) -> CohortSpreadBandResponse:
    """+/-1 SD ribbon around the group's HC ring, as real 3D points - see
    craniumpy_core.cohort.hc_ring_band. same request shape as
    /sagittal-band (mesh_paths + target), computed the same
    independently-of-any-cached-result way."""
    if request.target not in ("cranium", "face"):
        raise HTTPException(status_code=400, detail=f"target must be 'cranium' or 'face', got {request.target!r}")
    paths = [Path(p) for p in request.mesh_paths]
    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        raise HTTPException(status_code=400, detail=f"mesh file(s) not found: {', '.join(missing)}")
    try:
        band = cohort.hc_ring_band(paths, request.target)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _spread_band_response(band)


@router.post("/metopic-band", response_model=CohortSpreadBandResponse)
def compute_metopic_band(request: CohortSagittalBandRequest) -> CohortSpreadBandResponse:
    """+/-1 SD ribbon around the group's metopic forehead contour, as real
    3D points - see craniumpy_core.cohort.metopic_band. same request shape
    as /sagittal-band."""
    if request.target not in ("cranium", "face"):
        raise HTTPException(status_code=400, detail=f"target must be 'cranium' or 'face', got {request.target!r}")
    paths = [Path(p) for p in request.mesh_paths]
    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        raise HTTPException(status_code=400, detail=f"mesh file(s) not found: {', '.join(missing)}")
    try:
        band = cohort.metopic_band(paths, request.target)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return _spread_band_response(band)


@router.post("/report")
def generate_mean_shape_report(request: CohortReportRequest):
    """the same multi-page PDF layout a real patient's own export produces
    (see api/results_bundle.mean_shape_report_pdf), for this group's mean
    shape instead - computes the mean, its measurements, and (optionally)
    every spread band that applies to this target, all in this one call,
    rather than requiring the caller to have already computed a cached
    /mean-shape result (a report is a one-shot download, not something
    that needs to stay around for later requests the way a viewer's own
    mean-shape result does). the sagittal/frontal-bossing band always
    applies (both targets have a forehead); the HC-ring band only for a
    cranium target, the metopic band only for a face target - same
    craniometrics-vs-metopic split measure_mean_shape itself follows."""
    if request.target not in ("cranium", "face"):
        raise HTTPException(status_code=400, detail=f"target must be 'cranium' or 'face', got {request.target!r}")
    paths = [Path(p) for p in request.mesh_paths]
    missing = [str(p) for p in paths if not p.is_file()]
    if missing:
        raise HTTPException(status_code=400, detail=f"mesh file(s) not found: {', '.join(missing)}")

    try:
        result = cohort.mean_shape(paths)
        measurements = cohort.measure_mean_shape(result.mesh, request.target)
        sagittal = cohort.sagittal_midline_band(paths, request.target) if request.include_spread_bands else None
        hc_ring = (
            cohort.hc_ring_band(paths, request.target)
            if request.include_spread_bands and request.target == "cranium"
            else None
        )
        metopic = (
            cohort.metopic_band(paths, request.target)
            if request.include_spread_bands and request.target == "face"
            else None
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    pdf_bytes = mean_shape_report_pdf(
        result.mesh, request.target, request.group_label, result.source_count, measurements,
        sagittal_band=sagittal, hc_ring_band=hc_ring, metopic_band=metopic,
    )
    safe_name = _sanitize_filename(f"{request.group_label}_report", extension=".pdf", default="mean_shape_report")
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


_EXPORT_HEADER_FILL = PatternFill(start_color="2F6FED", end_color="2F6FED", fill_type="solid")
_EXPORT_HEADER_FONT = Font(bold=True, color="FFFFFF")
_INVALID_SHEET_NAME_CHARS = re.compile(r"[\[\]:*?/\\]")


def _looks_numeric(value: str) -> bool:
    try:
        float(value)
    except ValueError:
        return False
    return True


def _sheet_name(title: str, index: int, used: set[str]) -> str:
    """a valid, unique-within-this-workbook Excel sheet name - max 31
    chars, none of []:*?/\\ (Excel's own hard limits, not a style choice).
    a blank/all-invalid title falls back to "Sheet{n}" rather than an
    empty name, which openpyxl itself would reject."""
    cleaned = _INVALID_SHEET_NAME_CHARS.sub("", title).strip()[:31] or f"Sheet{index + 1}"
    name = cleaned
    suffix_n = 2
    while name in used:
        suffix = f" ({suffix_n})"
        name = cleaned[: 31 - len(suffix)] + suffix
        suffix_n += 1
    used.add(name)
    return name


def _build_export_xlsx(sheets: list[CohortExportSheet]) -> bytes:
    """one worksheet per sheet - bold white-on-blue header, frozen header
    row, banded rows (openpyxl's own TableStyleMedium2, same table style
    api/results_bundle.py's own cohort/summary spreadsheets already use),
    and a best-effort numeric column format: a column gets real numeric
    cells (2-decimal format) only when EVERY non-blank cell in it parses
    as a float, never a mix - the same "a patient_id like 0042 must keep
    its leading zero" reasoning api/results_bundle.py's own
    _NUMERIC_ROW_KEYS follows, just determined at export time instead of
    from a fixed known column list (an export sheet can be any columns the
    caller picked - the Stratify tab's own descriptive-stats sheet, for
    instance, has no column that also appears in a patient export)."""
    wb = Workbook()
    wb.remove(wb.active)
    used_names: set[str] = set()

    for i, sheet in enumerate(sheets):
        ws = wb.create_sheet(_sheet_name(sheet.title, i, used_names))
        columns, rows = sheet.columns, sheet.rows
        if not columns:
            continue

        numeric_columns = {
            col
            for col in columns
            if any(row.get(col, "") != "" for row in rows)
            and all(row.get(col, "") == "" or _looks_numeric(row.get(col, "")) for row in rows)
        }

        ws.append(columns)
        for cell in ws[1]:
            cell.font = _EXPORT_HEADER_FONT
            cell.fill = _EXPORT_HEADER_FILL
        ws.freeze_panes = "A2"

        display_values: list[list[str]] = []
        for row in rows:
            cells = []
            for name in columns:
                value = row.get(name, "")
                if name in numeric_columns and value != "":
                    cells.append(float(value))
                elif name in numeric_columns:
                    cells.append(None)
                else:
                    cells.append(value)
            ws.append(cells)
            display_values.append([str(c) if c is not None else "" for c in cells])

        for j, name in enumerate(columns, start=1):
            if name in numeric_columns:
                for cell in ws[get_column_letter(j)][1:]:
                    cell.number_format = "0.00"
            widest = max([len(name)] + [len(values[j - 1]) for values in display_values])
            ws.column_dimensions[get_column_letter(j)].width = min(max(widest + 2, 10), 40)

        if rows:
            last_col = get_column_letter(len(columns))
            table = Table(displayName=f"export_table_{i}", ref=f"A1:{last_col}{len(rows) + 1}")
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@router.post("/export-xlsx")
def export_xlsx(request: CohortExportRequest):
    """a formatted Excel workbook built from whatever sheets the caller
    already assembled client-side (the Table tab's current sorted/filtered
    view, or the Stratify tab's descriptive-stats + test-result sheets) -
    see _build_export_xlsx for the actual formatting."""
    xlsx_bytes = _build_export_xlsx(request.sheets)
    safe_name = _sanitize_filename(request.filename, extension=".xlsx", default="cohort_export")
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )
